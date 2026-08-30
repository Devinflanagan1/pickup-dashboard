"""
The Compton — Daily Pick-Up Report Builder
Run it with:   streamlit run app.py
"""

import io
import numpy as np
import pandas as pd
import streamlit as st

st.set_page_config(page_title="Compton Daily Pick-Up Report", layout="wide")


def _num(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return np.nan


def _read_all_sheets(uploaded_file):
    xls = pd.ExcelFile(uploaded_file, engine="openpyxl")
    return {s: xls.parse(s, header=None) for s in xls.sheet_names}


def detect_and_load(uploaded_files):
    found = {"pcdc": None, "data": None, "mseg": None, "shop": None,
             "shop_chg_1": None, "shop_chg_3": None, "shop_chg_7": None,
             "pcdc_asof": None}

    for f in uploaded_files:
        try:
            sheets = _read_all_sheets(f)
        except Exception as e:
            st.warning(f"Could not open **{f.name}** ({e}).")
            continue
        names = [s.lower() for s in sheets.keys()]

        if any("rate" in n for n in names) and any("overview" in n for n in names):
            raw = sheets[[s for s in sheets if s.lower() == "rates"][0]]
            found["shop"] = _parse_rateshop(raw)
            for tab, key in [("vs. yesterday", "shop_chg_1"),
                             ("vs. 3 days ago", "shop_chg_3"),
                             ("vs. 7 days ago", "shop_chg_7")]:
                match = [s for s in sheets if s.lower() == tab]
                if match:
                    found[key] = _parse_shop_change(sheets[match[0]])

        elif any("changereport" in n for n in names):
            raw = sheets[[s for s in sheets if "changereport" in s.lower()][0]]
            found["pcdc"] = _parse_pcdc(raw)
            crit = [s for s in sheets if s.lower() == "report criteria"]
            if crit:
                found["pcdc_asof"] = _parse_pcdc_asof(sheets[crit[0]])

        elif any("market segment" in n for n in names):
            raw = sheets[[s for s in sheets if s.lower() == "market segment"][0]]
            found["mseg"] = _parse_mseg(raw)

        elif any(n == "property" for n in names):
            raw = sheets[[s for s in sheets if s.lower() == "property"][0]]
            found["data"] = _parse_data(raw)

        else:
            st.warning(f"Didn't recognise **{f.name}** — skipping. "
                       f"(sheets: {', '.join(sheets.keys())})")
    return found


def _parse_data(raw):
    df = raw.copy()
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)
    df["Occupancy Date"] = pd.to_datetime(df["Occupancy Date"], errors="coerce")
    df = df.dropna(subset=["Occupancy Date"])
    return df


def _parse_pcdc(raw):
    df = raw.iloc[3:].copy()
    df[0] = pd.to_datetime(df[0], errors="coerce")
    df = df.dropna(subset=[0]).reset_index(drop=True)
    out = pd.DataFrame({"Occupancy Date": df[0]})
    out["pc_event"]      = df[2]
    out["trans_cur"]     = pd.to_numeric(df[3], errors="coerce")
    out["trans_chg"]     = pd.to_numeric(df[4], errors="coerce")
    out["grp_cur"]       = pd.to_numeric(df[5], errors="coerce")
    out["grp_chg"]       = pd.to_numeric(df[6], errors="coerce")
    out["grp_blocked"]   = pd.to_numeric(df[7], errors="coerce")
    out["grp_avail"]     = pd.to_numeric(df[8], errors="coerce")
    out["grp_pickup"]    = pd.to_numeric(df[9], errors="coerce")
    out["fcst_trans"]    = pd.to_numeric(df[10], errors="coerce")
    out["fcst_grp"]      = pd.to_numeric(df[12], errors="coerce")
    out["rev_trans"]     = pd.to_numeric(df[14], errors="coerce")
    out["rev_grp"]       = pd.to_numeric(df[16], errors="coerce")
    out["adr_trans"]     = pd.to_numeric(df[22], errors="coerce")
    out["adr_grp"]       = pd.to_numeric(df[24], errors="coerce")
    return out


def _parse_pcdc_asof(raw):
    """
    Read the PCDC 'Report Criteria' sheet and return the reference date for
    'Days Left' (i.e. "today" for pace purposes).

    'Generated On' is the actual timestamp the report was run/exported, so it's
    the most reliable proxy for "today". 'Analysis Start Date' and 'Activity
    Start Date' are report-WINDOW parameters the revenue manager can set to any
    date (e.g. to look at pace from a few days back) — they are NOT guaranteed
    to equal today, so they're only used as a fallback if 'Generated On' is
    missing. Returns a datetime.date or None.
    """
    try:
        # find the header row that contains the labels, then the value row below it
        for i in range(len(raw) - 1):
            rowvals = [str(x).strip().lower() for x in raw.iloc[i].tolist()]
            if "analysis start date" in rowvals or "generated on" in rowvals:
                header = raw.iloc[i].tolist()
                values = raw.iloc[i + 1].tolist()
                lut = {str(h).strip().lower(): v for h, v in zip(header, values)}
                for key in ("generated on", "activity start date", "analysis start date"):
                    if key in lut and lut[key] not in (None, ""):
                        d = pd.to_datetime(lut[key], errors="coerce")
                        if pd.notna(d):
                            return d.date()
    except Exception:
        pass
    return None


def _parse_mseg(raw):
    df = raw.copy()
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)
    df["Occupancy Date"] = pd.to_datetime(df["Occupancy Date"], errors="coerce")
    df = df.dropna(subset=["Occupancy Date"])
    df["Occupancy On Books This Year"] = pd.to_numeric(
        df["Occupancy On Books This Year"], errors="coerce").fillna(0)
    df["is_trans"] = df["Market Segment"].astype(str).str.startswith("Transient")
    agg = df.groupby("Occupancy Date").apply(
        lambda g: pd.Series({
            "ms_trans": g.loc[g["is_trans"], "Occupancy On Books This Year"].sum(),
            "ms_group": g.loc[~g["is_trans"], "Occupancy On Books This Year"].sum(),
        })
    ).reset_index()
    return agg


def _clean_shop(x):
    if x is None:
        return None
    v = _num(x)
    if not np.isnan(v):
        return int(v) if float(v).is_integer() else v
    s = str(x).strip()
    return s if s else None


def _parse_rateshop(raw):
    df = raw.iloc[5:].copy()
    df[2] = pd.to_datetime(df[2], errors="coerce")
    df = df.dropna(subset=[2]).reset_index(drop=True)
    out = pd.DataFrame({"Occupancy Date": df[2]})
    out["own_bar"] = df[4].map(_num)
    for key, col in [("c_21c", 5), ("c_motto", 6), ("c_ac", 7), ("c_dt", 8)]:
        out[key] = df[col].map(_clean_shop)
        out[key + "_n"] = df[col].map(_num)
    return out


def _parse_shop_change(raw):
    df = raw.iloc[5:].copy()
    df[2] = pd.to_datetime(df[2], errors="coerce")
    df = df.dropna(subset=[2]).reset_index(drop=True)
    out = pd.DataFrame({"Occupancy Date": df[2]})
    # our own BAR change sits in column 6 (right of our rate in column 5)
    out["own_chg"] = df[6].map(_num) if 6 in df.columns else np.nan
    for key, chg_col in [("c_21c", 8), ("c_motto", 10), ("c_ac", 12), ("c_dt", 14)]:
        out[key + "_chg"] = df[chg_col].map(_num) if chg_col in df.columns else np.nan
    return out


def compute_monthly_recap(raw, as_of):
    """Owner-grade monthly recap. Returns (months, metrics) where months is a list
    of 'Mon YYYY' labels and metrics is a dict keyed by metric name -> list of
    per-month dicts {otb, fcst, budget, stly}. All pulled from the Data Extract."""
    import calendar as _cal
    d = raw.copy()
    d["Occupancy Date"] = pd.to_datetime(d["Occupancy Date"], errors="coerce")
    d = d.dropna(subset=["Occupancy Date"])

    def col(name):
        return pd.to_numeric(d[name], errors="coerce") if name in d.columns else pd.Series(0.0, index=d.index)

    myf = col("My Forecast Revenue This Year")
    fc_rev_src = "My Forecast Revenue This Year" if myf.fillna(0).abs().sum() > 0 else "Forecasted Room Revenue This Year"
    myfo = col("My Forecast Occupancy - Total This Year")
    fc_occ_src = "My Forecast Occupancy - Total This Year" if myfo.fillna(0).abs().sum() > 0 else "Occupancy Forecast - Total This Year"
    fc_kind = "Your forecast" if fc_rev_src.startswith("My") else "System forecast"

    d["_avail"]   = col("Physical Capacity This Year")
    d["_otb_rev"] = col("Booked Room Revenue This Year")
    d["_otb_rms"] = col("Occupancy On Books This Year")
    d["_fc_rev"]  = col(fc_rev_src)
    d["_fc_rms"]  = col(fc_occ_src)
    d["_bud_rev"] = col("Budget Room Revenue This Year")
    d["_bud_rms"] = col("Budget Occupancy - Total This Year")
    d["_sy_rev"]  = col("Booked Room Revenue STLY")
    d["_sy_rms"]  = col("Occupancy On Books STLY")
    d["_ym"] = list(zip(d["Occupancy Date"].dt.year, d["Occupancy Date"].dt.month))

    base = pd.Timestamp(as_of)
    yms = [((base + pd.DateOffset(months=k)).year, (base + pd.DateOffset(months=k)).month) for k in range(3)]
    labels = [f"{_cal.month_abbr[m]} {y}" for y, m in yms]

    def scen(sub, rev_c, rms_c):
        # cast to plain Python float so downstream isinstance()/formatting is reliable
        rev = float(sub[rev_c].sum()); rms = float(sub[rms_c].sum()); avail = float(sub["_avail"].sum())
        return {
            "Rooms Sold": rms,
            "Occupancy %": (rms / avail) if avail else np.nan,
            "ADR": (rev / rms) if rms else np.nan,
            "RevPAR": (rev / avail) if avail else np.nan,
            "Room Revenue": rev,
        }

    metric_names = ["Rooms Sold", "Occupancy %", "ADR", "RevPAR", "Room Revenue"]
    metrics = {mn: [] for mn in metric_names}
    for ym in yms:
        sub = d[d["_ym"] == ym]
        otb = scen(sub, "_otb_rev", "_otb_rms")
        fcst = scen(sub, "_fc_rev", "_fc_rms")
        bud = scen(sub, "_bud_rev", "_bud_rms")
        stly = scen(sub, "_sy_rev", "_sy_rms")
        for mn in metric_names:
            metrics[mn].append({"otb": otb[mn], "fcst": fcst[mn], "budget": bud[mn], "stly": stly[mn]})
    return labels, metrics, fc_kind


def build_report(parts, as_of, comp_method="Average", show_change="None", thr=None,
                 show_own_change=True):
    data, pcdc, mseg, shop = parts["data"], parts["pcdc"], parts["mseg"], parts["shop"]
    if data is None:
        st.error("The **Data Extract** file is required (couldn't find a 'Property' sheet).")
        st.stop()

    df = data.copy()
    for extra in (pcdc, mseg, shop):
        if extra is not None:
            df = df.merge(extra, on="Occupancy Date", how="left")

    chg_key = {"Yesterday": "shop_chg_1", "3 days ago": "shop_chg_3",
               "7 days ago": "shop_chg_7"}.get(show_change)
    chg_df = parts.get(chg_key) if chg_key else None
    if chg_df is not None:
        # avoid duplicate own_chg if we also pull the 7-day own change below
        df = df.merge(chg_df.drop(columns=[c for c in ["own_chg"] if c in chg_df], errors="ignore"),
                      on="Occupancy Date", how="left")

    # Always pull OUR OWN 7-day BAR change (independent of the competitor toggle)
    own7 = parts.get("shop_chg_7")
    if own7 is not None and "own_chg" in own7:
        df = df.merge(own7[["Occupancy Date", "own_chg"]].rename(columns={"own_chg": "bar_chg_7d"}),
                      on="Occupancy Date", how="left")

    def dnum(col):
        return pd.to_numeric(df.get(col), errors="coerce")

    cap        = dnum("Physical Capacity This Year")
    occ_ty     = dnum("Occupancy On Books This Year")
    occ_stly   = dnum("Occupancy On Books STLY")
    tr_ty      = dnum("Rooms Sold - Transient This Year")
    tr_stly    = dnum("Rooms Sold - Transient STLY")
    gr_ty      = dnum("Rooms Sold - Group This Year")
    gr_stly    = dnum("Rooms Sold - Group STLY")
    dem_tot    = dnum("User Total Demand - Total This Year")
    dem_grp    = dnum("User Constrained Total Demand - Group This Year")
    dem_tr     = dnum("User Unconstrained Total Demand - Transient This Year")

    r = pd.DataFrame()
    r["DOW"]  = df["Day of Week"]
    r["Date"] = df["Occupancy Date"].dt.date
    r["Days Left"] = (df["Occupancy Date"] - pd.Timestamp(as_of)).dt.days

    ev = df.get("Special Event This Year")
    ev = ev.fillna("") if ev is not None else ""
    if "pc_event" in df:
        ev = ev.where(ev.astype(str).str.strip() != "", df["pc_event"].fillna(""))
    r["Events"] = pd.Series(ev, index=df.index).replace("nan", "")

    r["Rooms Left to Sell"] = dnum("Remaining Capacity This Year")
    r["BAR"]                = dnum("BAR")
    if show_own_change:
        r["BAR Chg 7d"]     = dnum("bar_chg_7d")     # our own rate move over last 7 days

    if "c_21c_n" in df:
        comps_n = df[["c_21c_n", "c_motto_n", "c_ac_n", "c_dt_n"]]
        r["Comp Set Avg"] = (comps_n.mean(axis=1) if comp_method == "Average"
                             else comps_n.median(axis=1)).round(0)
    else:
        r["Comp Set Avg"] = np.nan

    r["Last Room Value"] = dnum("Last Room Value This Year")

    add_chg = chg_df is not None
    for name, key in [("21c Museum Hotel Bentonville - MGallery", "c_21c"),
                      ("Motto By Hilton Bentonville Downtown",    "c_motto"),
                      ("AC Hotel by Marriott Bentonville",        "c_ac"),
                      ("DoubleTree Suites by Hilton Bentonville", "c_dt")]:
        base = "Competitor Shops | " + name
        if add_chg:
            r[base + " | Current"] = df.get(key)
            r[base + " | Change"]  = df.get(key + "_chg")
        else:
            r[base] = df.get(key)

    r["OOO"]   = dnum("Rooms N/A - Out of Order This Year")
    r["Ovrbk"] = dnum("Overbooking This Year")

    tc, tch = df.get("trans_cur"), df.get("trans_chg")
    gc, gch = df.get("grp_cur"),   df.get("grp_chg")
    r["Rooms Sold | Total Hotel | Current"]   = (tc.fillna(0) + gc.fillna(0)) if tc is not None else occ_ty
    r["Rooms Sold | Total Hotel | Change"]    = (tch.fillna(0) + gch.fillna(0)) if tch is not None else np.nan
    r["Rooms Sold | Total Transient | Current"] = tc
    r["Rooms Sold | Total Transient | Change"]  = tch
    r["Rooms Sold | Total Group | Current"]     = gc
    r["Rooms Sold | Total Group | Change"]      = gch
    r["Rooms Sold | Total Group | Blocked"]     = df.get("grp_blocked")
    r["Rooms Sold | Total Group | P/U"]         = df.get("grp_pickup")
    r["Rooms Sold | Total Group | Remaining"]   = df.get("grp_avail")

    r["Rooms OTB STLY | Total Hotel | Current"]     = occ_stly
    r["Rooms OTB STLY | Total Hotel | Change"]      = occ_ty - occ_stly
    r["Rooms OTB STLY | Total Transient | Current"] = tr_stly
    r["Rooms OTB STLY | Total Transient | Change"]  = tr_ty - tr_stly
    r["Rooms OTB STLY | Total Group | Current"]     = gr_stly
    r["Rooms OTB STLY | Total Group | Change"]      = gr_ty - gr_stly

    r["Remaining Demand | Total Hotel"]     = (dem_tot - occ_ty).clip(lower=0)
    r["Remaining Demand | Total Transient"] = (dem_tr - tr_ty).clip(lower=0)
    r["Remaining Demand | Total Group"]     = (dem_grp - gr_ty).clip(lower=0)

    ft, fg = df.get("fcst_trans"), df.get("fcst_grp")
    fcst_tot = (ft.fillna(0) + fg.fillna(0)) if ft is not None else dnum("Forecasted Room Revenue This Year")*np.nan
    r["Occ Forecast | Total Hotel"]     = fcst_tot.round(1) if ft is not None else np.nan
    r["Occ Forecast | Total Transient"] = ft.round(1) if ft is not None else np.nan
    r["Occ Forecast | Total Group"]     = fg.round(1) if fg is not None else np.nan

    r["Occ Forecast % | Total Hotel"]     = (fcst_tot / cap) if ft is not None else np.nan
    r["Occ Forecast % | Total Transient"] = (ft / cap) if ft is not None else np.nan
    r["Occ Forecast % | Total Group"]     = (fg / cap) if fg is not None else np.nan

    rev_t, rev_g = df.get("rev_trans"), df.get("rev_grp")
    if tc is not None:
        total_rooms = (tc.fillna(0) + gc.fillna(0)).replace(0, np.nan)
        r["Booked ADR | Total Hotel"] = ((rev_t.fillna(0) + rev_g.fillna(0)) / total_rooms).round(2)
    else:
        r["Booked ADR | Total Hotel"] = dnum("ADR On Books This Year")
    r["Booked ADR | Total Transient"] = df.get("adr_trans")
    r["Booked ADR | Total Group"]     = df.get("adr_grp")

    r["Estimated ADR"] = dnum("ADR Forecast This Year")

    # ================================================================= #
    #  PRICING DECISION SIGNALS  (VP-of-Revenue toolkit)
    # ================================================================= #
    bar   = r["BAR"]
    comp  = r["Comp Set Avg"]
    rl    = r["Rooms Left to Sell"]
    rd    = r["Remaining Demand | Total Hotel"]
    ofp   = r["Occ Forecast % | Total Hotel"]
    pace  = r["Rooms OTB STLY | Total Hotel | Change"]      # TY − STLY (+ = ahead)
    dleft = r["Days Left"]
    floor   = dnum("Floor")
    ceiling = dnum("Ceiling")

    # --- Rate Rank badge:  where our BAR sits in the comp set (1 = highest) --- #
    rank_cols = [c for c in ("c_21c_n", "c_motto_n", "c_ac_n", "c_dt_n") if c in df]
    if rank_cols:
        price_df = pd.DataFrame({"own": pd.to_numeric(r["BAR"], errors="coerce").values})
        for cc in rank_cols:
            price_df[cc] = pd.to_numeric(df[cc], errors="coerce").values

        def _rank_row(row):
            own = row["own"]
            allv = [v for v in row.tolist() if pd.notna(v)]
            if pd.isna(own) or not allv:
                return ""
            higher = sum(1 for v in allv if v > own)
            return f"{higher + 1} of {len(allv)}"

        r["Rate Rank"] = [_rank_row(price_df.iloc[i]) for i in range(len(price_df))]
    else:
        r["Rate Rank"] = ""

    # --- Signal thresholds (tunable from the sidebar) --- #
    thr = thr or {}
    lead_buffer   = thr.get("lead_buffer", 20)    # min $ cushion over the #2 to be "safely" #1
    pickup_pctile = thr.get("pickup_pctile", 75)  # percentile of positive pickups = "big"
    pickup_min    = thr.get("pickup_min", 3)      # floor so quiet reports don't flag tiny pickup

    # Strategy: The Compton prices as the market LEADER. The Action column is
    # purely about rate-leadership position — two flags only:
    #   🔴 Not #1        — a competitor is at or above our BAR
    #   🟠 Too close #2  — we're #1, but our cushion over the top competitor
    #                      is thinner than we want (within "lead_buffer" $)

    # top competitor rate per date (max of the 4 shops) — effectively "the #2"
    if rank_cols:
        comp_matrix = pd.DataFrame({cc: pd.to_numeric(df[cc], errors="coerce").values
                                    for cc in rank_cols})
        max_comp = comp_matrix.max(axis=1)
        max_comp.index = r.index
    else:
        max_comp = pd.Series(np.nan, index=r.index)

    # gap to the #2 (top competitor): + = we're above them, 0/− = not #1
    lead_gap = (bar - max_comp)
    r["Lead Gap vs Comp"] = lead_gap.round(0)

    not_leader = pd.notna(max_comp) & (lead_gap <= 0)             # someone matches/beats us
    too_close  = pd.notna(max_comp) & (lead_gap > 0) & (lead_gap < lead_buffer)
    at_ceiling = bar >= ceiling

    def _label(i):
        if pd.isna(dleft.iloc[i]) or dleft.iloc[i] < 0:
            return ""                                     # past date
        if pd.isna(max_comp.iloc[i]):
            return ""                                     # no comp-set data for this date
        if bool(not_leader.iloc[i]):
            return "🔴 Not #1" if not bool(at_ceiling.iloc[i]) else "🟠 Not #1 · at ceiling"
        if bool(too_close.iloc[i]):
            return "🟠 Too close to #2" if not bool(at_ceiling.iloc[i]) else "🟠 At ceiling"
        return "✅ Leading"

    r["⚡ Action"] = [_label(i) for i in range(len(r))]

    # --- 🔥 Spike: flags an unusually large day's pickup ------------- #
    # "Big" is relative, not a fixed number, so this self-calibrates off this
    # property's own pickup pattern: it flags a date when its pickup sits at or
    # above the Nth percentile of all POSITIVE daily pickups in this report
    # (default 90th = top 10%), with a small floor so a quiet report full of
    # 1-2 room days doesn't get flagged just because it's "relatively" high.
    pickup = r["Rooms Sold | Total Hotel | Change"]
    pos_pickup = pickup[pickup > 0]
    if len(pos_pickup) >= 5:
        pctile_val = pos_pickup.quantile(pickup_pctile / 100)
        pickup_threshold = max(pctile_val, pickup_min)
    else:
        pickup_threshold = pickup_min          # not enough data to calibrate — use the floor

    is_big_pickup = (pickup >= pickup_threshold) & (dleft >= 0)

    def _pickup_label(i):
        if bool(is_big_pickup.iloc[i]):
            v = pickup.iloc[i]
            return f"🔥 Spike (+{v:,.0f})"
        return ""

    r["🔥 Spike"] = [_pickup_label(i) for i in range(len(r))]

    # ---- Enforce the exact original report column order ---- #
    lead = ["⚡ Action", "🔥 Spike", "DOW", "Date", "Days Left", "Events",
            "Rooms Left to Sell", "BAR", "BAR Chg 7d", "Comp Set Avg", "Rate Rank",
            "Lead Gap vs Comp", "Last Room Value"]
    comp = [c for c in r.columns if c.startswith("Competitor Shops")]
    mid  = ["OOO", "Ovrbk"]
    groups = ["Rooms Sold", "Rooms OTB STLY", "Remaining Demand",
              "Occ Forecast |", "Occ Forecast %", "Booked ADR"]
    grouped = []
    for g in groups:
        grouped += [c for c in r.columns if c.startswith(g)]
    tail = ["Estimated ADR"]
    order = lead + comp + mid + grouped + tail
    order = [c for c in order if c in r.columns]          # keep only existing
    order += [c for c in r.columns if c not in order]     # safety: append any strays
    r = r[order]

    return r


def _levels(col):
    parts = [p.strip() for p in str(col).split("|")]
    while len(parts) < 3:
        parts.append("")
    return tuple(parts[:3])


def to_excel(r, kpi_ctx=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    lv = [_levels(c) for c in r.columns]
    n = len(lv)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pick-Up Report"

    NAVY, TEAL, ORANGE, PURPLE, GREEN = "16365C", "215967", "E26B0A", "60497A", "76933C"
    white = Font(color="FFFFFF", bold=True, size=9)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

    def PF(hexcol):
        return PatternFill("solid", fgColor=hexcol)

    def seg_color(txt):
        if "Total Hotel" in txt:
            return ORANGE
        if "Total Transient" in txt:
            return PURPLE
        if "Total Group" in txt:
            return GREEN
        return None

    def band_fill(l1, l2):
        return PF(NAVY) if l2 == "" else PF(TEAL)

    def lvl2_fill(l1, l2):
        sc = seg_color(l2)
        if sc:
            return PF(sc)
        if l1 == "Competitor Shops":
            return PF(NAVY)
        return PF(TEAL)

    def lvl3_fill(l1, l2):
        sc = seg_color(l2)
        return PF(sc) if sc else PF(NAVY)

    j = 0
    while j < n:
        l1 = lv[j][0]
        k = j
        while k + 1 < n and lv[k + 1][0] == l1 and lv[j][1]:
            k += 1
        fill = band_fill(l1, lv[j][1])
        if lv[j][1] == "" and l1 in ("Estimated ADR",):
            fill = PF(ORANGE)
        cell = ws.cell(row=1, column=j + 1, value=l1)
        cell.font = white
        cell.alignment = center
        cell.fill = fill
        if lv[j][1] == "":
            ws.merge_cells(start_row=1, start_column=j + 1, end_row=3, end_column=j + 1)
        elif k > j:
            ws.merge_cells(start_row=1, start_column=j + 1, end_row=1, end_column=k + 1)
        j = k + 1

    j = 0
    while j < n:
        l1, l2, l3 = lv[j]
        if l2:
            k = j
            while k + 1 < n and lv[k + 1][0] == l1 and lv[k + 1][1] == l2:
                k += 1
            cell = ws.cell(row=2, column=j + 1, value=l2)
            cell.font = white
            cell.alignment = center
            cell.fill = lvl2_fill(l1, l2)
            if l3 == "":
                ws.merge_cells(start_row=2, start_column=j + 1, end_row=3, end_column=j + 1)
            elif k > j:
                ws.merge_cells(start_row=2, start_column=j + 1, end_row=2, end_column=k + 1)
            j = k + 1
        else:
            j += 1

    for j, (l1, l2, l3) in enumerate(lv, start=1):
        if l3:
            cell = ws.cell(row=3, column=j, value=l3)
            cell.font = white
            cell.alignment = center
            cell.fill = lvl3_fill(l1, l2)

    _hotelnames = ("Bentonville", "MGallery")
    DATA0 = 4
    pct = {i for i, c0 in enumerate(r.columns) if c0.startswith("Occ Forecast %")}
    signed_cols = ("BAR Chg 7d", "Lead Gap vs Comp")
    chg = {i for i, c0 in enumerate(r.columns)
           if (any(h in c0 for h in _hotelnames) and c0.endswith("Change"))
           or c0 in signed_cols}
    money = {i for i, c0 in enumerate(r.columns)
             if (c0 in ("BAR", "Comp Set Avg", "Last Room Value", "Estimated ADR")
                 or "ADR" in c0 or any(h in c0 for h in _hotelnames))
             and i not in chg}
    otb_sold_col = None
    for i, c0 in enumerate(r.columns):
        if c0 == "Rooms Left to Sell":
            otb_sold_col = i

    def _action_fill(txt):
        t = str(txt)
        if "Not #1" in t:          return PatternFill("solid", fgColor="F8D7DA")
        if "Too close" in t:       return PatternFill("solid", fgColor="FFF3CD")
        if "At ceiling" in t:      return PatternFill("solid", fgColor="FFE5B4")
        if "Leading" in t:         return PatternFill("solid", fgColor="D1E7DD")
        return None

    def _pickup_fill(txt):
        return PatternFill("solid", fgColor="FDE7CE") if "Spike" in str(txt) else None

    for ri, (_, row) in enumerate(r.iterrows(), start=DATA0):
        for ci2, col in enumerate(r.columns, start=1):
            v = row[col]
            cell = ws.cell(row=ri, column=ci2, value=(None if pd.isna(v) else v))
            cell.border = thin
            cell.font = Font(size=9)
            # center everything; left-align the long-text "Events" column
            halign = "left" if col == "Events" else "center"
            cell.alignment = Alignment(horizontal=halign, vertical="center")
            if col == "⚡ Action":
                af = _action_fill(v)
                if af is not None:
                    cell.fill = af
            elif col == "🔥 Spike":
                pf = _pickup_fill(v)
                if pf is not None:
                    cell.fill = pf
                    cell.font = Font(size=9, bold=True, color="B45309")
            idx = ci2 - 1
            if idx in pct:
                cell.number_format = "0.0%"
            elif idx in chg:
                cell.number_format = "+#,##0;-#,##0;\"\""
            elif idx in money:
                cell.number_format = "$#,##0"
            elif isinstance(v, float) and float(v).is_integer():
                cell.number_format = "#,##0"

    if otb_sold_col is not None and len(r) > 0:
        L = get_column_letter(otb_sold_col + 1)
        rng = f"{L}{DATA0}:{L}{DATA0 + len(r) - 1}"
        # Rooms Left to Sell: RED = few left (near sell-out), GREEN = lots left
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(start_type="min", start_color="F8696B",         # red (few left)
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",  # yellow
                           end_type="max", end_color="63BE7B"))            # green (lots left)

    ws.freeze_panes = "C4"
    for j in range(1, n + 1):
        ws.column_dimensions[get_column_letter(j)].width = 12
    ws.column_dimensions["A"].width = 16   # ⚡ Action
    ws.column_dimensions["E"].width = 22   # Events

    # Excel AutoFilter so the team can filter/sort right in Excel.
    # Anchored on the bottom header row (row 3) through the last data row.
    last_row = DATA0 + len(r) - 1
    ws.auto_filter.ref = f"A3:{get_column_letter(n)}{last_row}"

    # ================================================================= #
    #  KPI Summary tab (self-contained snapshot for sharing)
    # ================================================================= #
    _build_kpi_sheet(wb, r, NAVY, TEAL, ORANGE, PURPLE, GREEN, kpi_ctx)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _month_key(d):
    return (d.year, d.month)


def _month_label(y, m):
    import calendar
    return f"{calendar.month_abbr[m]} {y}"


def _build_kpi_sheet(wb, r, NAVY, TEAL, ORANGE, PURPLE, GREEN, kpi_ctx=None):
    """Owner-grade monthly recap tab: Rooms / Occupancy% / ADR / RevPAR / Revenue,
    compared across On-the-Books, Forecast, Budget and Last Year, for the current /
    next / following month. Plus a Group vs Transient rooms split."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import calendar
    kpi_ctx = kpi_ctx or {}
    as_of = kpi_ctx.get("as_of", pd.Timestamp.today().date())
    raw = kpi_ctx.get("data_raw")

    ks = wb.create_sheet("KPI Summary", 0)
    white = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    thin = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

    ks.merge_cells("A1:G1")
    t = ks["A1"]; t.value = "The Compton - Monthly Owner Recap"
    t.font = Font(color="FFFFFF", bold=True, size=14); t.alignment = center
    t.fill = PatternFill("solid", fgColor=NAVY); ks.row_dimensions[1].height = 30

    if raw is None or "Occupancy Date" not in getattr(raw, "columns", []):
        ks["A3"] = "Upload the Data Extract with Budget/Forecast columns to populate this recap."
        ks["A3"].font = Font(size=11, italic=True, color="884400")
        return ks

    labels, metrics, fc_kind = compute_monthly_recap(raw, as_of)

    ks.merge_cells("A2:G2")
    s2 = ks["A2"]
    s2.value = f"As of {pd.Timestamp(as_of):%b %d, %Y}   -   Forecast source: {fc_kind}   -   Occupancy & RevPAR use physical capacity"
    s2.font = Font(size=10, color="555555"); s2.alignment = center

    pct_metrics = {"Occupancy %"}
    money_metrics = {"ADR", "RevPAR", "Room Revenue"}

    def numfmt(metric):
        if metric in pct_metrics: return "0.0%"
        if metric in money_metrics: return "$#,##0"
        return "#,##0"

    def var_numfmt(metric):
        if metric in pct_metrics: return "+0.0%;-0.0%"
        if metric in money_metrics: return "$+#,##0;$-#,##0"
        return "+#,##0;-#,##0"

    metric_order = ["Rooms Sold", "Occupancy %", "ADR", "RevPAR", "Room Revenue"]
    heads = ["Metric", "On the Books", "Forecast", "Budget", "Last Year",
             "Fcst vs Budget", "OTB vs Budget"]
    monthcolors = [ORANGE, PURPLE, "3A6EA5"]

    row = 4
    for mi, label in enumerate(labels):
        # month banner
        ks.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        b = ks.cell(row=row, column=1, value=label)
        b.font = Font(color="FFFFFF", bold=True, size=12); b.alignment = center
        b.fill = PatternFill("solid", fgColor=monthcolors[mi % 3])
        ks.row_dimensions[row].height = 22
        row += 1
        # header
        for j, h in enumerate(heads, start=1):
            c = ks.cell(row=row, column=j, value=h)
            c.font = white; c.alignment = center; c.fill = PatternFill("solid", fgColor=TEAL)
        row += 1
        # metric rows
        for metric in metric_order:
            m = metrics[metric][mi]
            fvb = (m["fcst"] - m["budget"]) if pd.notna(m["fcst"]) and pd.notna(m["budget"]) else None
            ovb = (m["otb"] - m["budget"]) if pd.notna(m["otb"]) and pd.notna(m["budget"]) else None
            cells = [metric, m["otb"], m["fcst"], m["budget"], m["stly"], fvb, ovb]
            for j, v in enumerate(cells, start=1):
                c = ks.cell(row=row, column=j,
                            value=(None if (isinstance(v, float) and pd.isna(v)) else v))
                c.border = thin
                if j == 1:
                    c.font = Font(size=10, bold=True); c.alignment = left
                else:
                    c.alignment = center; c.font = Font(size=10)
                    if j in (6, 7):
                        c.number_format = var_numfmt(metric)
                        if isinstance(v, (int, float)) and pd.notna(v):
                            c.font = Font(size=10, bold=True, color=("1a7f37" if v >= 0 else "C00000"))
                    else:
                        c.number_format = numfmt(metric)
            row += 1
        row += 1  # spacer between months

    # ---- Group vs Transient split (rooms) ----
    d = raw.copy()
    d["Occupancy Date"] = pd.to_datetime(d["Occupancy Date"], errors="coerce")
    d = d.dropna(subset=["Occupancy Date"])
    def col(name):
        return pd.to_numeric(d[name], errors="coerce") if name in d.columns else pd.Series(0.0, index=d.index)
    myf_g = col("My Forecast Occupancy - Group This Year")
    fcg = "My Forecast Occupancy - Group This Year" if myf_g.fillna(0).abs().sum() > 0 else "Occupancy Forecast - Group This Year"
    myf_t = col("My Forecast Occupancy - Transient This Year")
    fct = "My Forecast Occupancy - Transient This Year" if myf_t.fillna(0).abs().sum() > 0 else "Occupancy Forecast - Transient This Year"
    d["_otb_g"] = col("Rooms Sold - Group This Year"); d["_otb_t"] = col("Rooms Sold - Transient This Year")
    d["_fc_g"] = col(fcg); d["_fc_t"] = col(fct)
    d["_bud_g"] = col("Budget Occupancy - Group This Year"); d["_bud_t"] = col("Budget Occupancy - Transient This Year")
    d["_ym"] = list(zip(d["Occupancy Date"].dt.year, d["Occupancy Date"].dt.month))
    base = pd.Timestamp(as_of)
    yms = [((base + pd.DateOffset(months=k)).year, (base + pd.DateOffset(months=k)).month) for k in range(3)]

    ks.cell(row=row, column=1, value="Rooms by segment - Group vs Transient (room-nights)").font = Font(bold=True, size=12)
    row += 1
    seg_heads = ["Month", "Seg", "OTB Rooms", "Fcst Rooms", "Budget Rooms", "Fcst vs Bud", "OTB vs Bud"]
    for j, h in enumerate(seg_heads, start=1):
        c = ks.cell(row=row, column=j, value=h)
        c.font = white; c.alignment = center; c.fill = PatternFill("solid", fgColor=NAVY)
    row += 1
    for i, ym in enumerate(yms):
        s = d[d["_ym"] == ym]
        for seg, otb, fc, bud, fill in [
            ("Group", s["_otb_g"].sum(), s["_fc_g"].sum(), s["_bud_g"].sum(), GREEN),
            ("Transient", s["_otb_t"].sum(), s["_fc_t"].sum(), s["_bud_t"].sum(), PURPLE)]:
            vals = [f"{calendar.month_abbr[ym[1]]} {ym[0]}", seg, otb, fc, bud, fc - bud, otb - bud]
            for j, v in enumerate(vals, start=1):
                c = ks.cell(row=row, column=j, value=v)
                c.border = thin; c.alignment = center; c.font = Font(size=10)
                if j == 1:
                    c.font = Font(size=10, bold=True, color="FFFFFF")
                    c.fill = PatternFill("solid", fgColor=monthcolors[i % 3])
                elif j == 2:
                    c.font = Font(size=10, bold=True, color="FFFFFF")
                    c.fill = PatternFill("solid", fgColor=fill)
                else:
                    c.number_format = "#,##0"
                    if j in (6, 7) and isinstance(v, (int, float)) and pd.notna(v):
                        c.font = Font(size=10, bold=True, color=("1a7f37" if v >= 0 else "C00000"))
            row += 1

    for cc, w in zip("ABCDEFG", [15, 14, 14, 14, 13, 15, 15]):
        ks.column_dimensions[cc].width = w
    ks.sheet_view.showGridLines = False
    return ks


st.title("The Compton — Daily Pick-Up Report Builder")
st.caption("Upload your four SynXis / rate-shop exports and get the full pick-up "
           "report instantly. No copy/paste required.")

with st.sidebar:
    st.header("1 - Upload your files")
    ups = st.file_uploader(
        "Drop in PCDC, Data Extract, Market Seg and the Rate-Shop file "
        "(all four at once is fine)",
        type=["xlsx"], accept_multiple_files=True,
    )

if not ups:
    st.info("Upload your files in the sidebar to build the report.")
    st.stop()

parts = detect_and_load(ups)

# The date range actually covered by the PCDC (pace/pickup) report, if uploaded.
# Data Extract typically spans the WHOLE YEAR, but PCDC only covers its shorter
# pace window (e.g. ~60-90 days). We use this as the default view so the report
# isn't cluttered with the full year of Data-Extract-only dates.
_pcdc_min = _pcdc_max = None
if parts.get("pcdc") is not None and "Occupancy Date" in parts["pcdc"].columns:
    _pcdc_dates = pd.to_datetime(parts["pcdc"]["Occupancy Date"], errors="coerce").dropna()
    if len(_pcdc_dates):
        _pcdc_min, _pcdc_max = _pcdc_dates.min().date(), _pcdc_dates.max().date()

# 'Days Left' should always default to the REAL current date — not a date read
# out of the PCDC file — since the file could have been exported days ago and a
# stale reference date silently throws off every Days Left calculation.
#
# IMPORTANT Streamlit rule: once a widget with a given `key` has been created,
# you can NOT overwrite st.session_state[key] later in that same script run —
# doing so raises a StreamlitAPIException. The only safe ways to set a widget's
# value from code are (a) BEFORE the widget is instantiated in this run, or
# (b) inside a button's `on_click=` callback (callbacks run before the widget
# is re-created on the next rerun). Both are used below.
real_today = pd.Timestamp.today().date()

# (a) Seed/refresh the value BEFORE the widget exists this run:
#     - first time this session has ever seen this key, OR
#     - a new/different set of files was just uploaded (by name+size) —
#       this is what guarantees the date snaps back to today even if an
#       earlier session left a stale value sitting in session_state.
_file_sig = tuple(sorted((f.name, getattr(f, "size", None)) for f in ups))
if ("as_of_input" not in st.session_state
        or st.session_state.get("_uploaded_file_sig") != _file_sig):
    st.session_state["_uploaded_file_sig"] = _file_sig
    st.session_state["as_of_input"] = real_today

# (b) Callback for the manual reset button — safe because callbacks run
#     before the widget is redrawn, never in the same pass as the widget.
def _reset_as_of_to_today():
    st.session_state["as_of_input"] = pd.Timestamp.today().date()

with st.sidebar:
    st.header("2 - Settings")
    st.caption(f"🕒 Server's current date right now: **{real_today:%b %d, %Y}**")
    as_of = st.date_input(
        "Report as-of date (drives 'Days Left')",
        key="as_of_input",
        help="Defaults to today's real date every time you upload files. Change "
             "it here only if you want to look at pace as of a different date.",
    )
    st.button("🔄 Reset as-of date to today", use_container_width=True,
              on_click=_reset_as_of_to_today)
    if parts.get("pcdc_asof"):
        st.caption(f"ℹ️ For reference, this PCDC report was generated on "
                   f"**{pd.Timestamp(parts['pcdc_asof']):%b %d, %Y}**.")
    comp_method = st.radio("Comp Set Avg method", ["Average", "Median"], horizontal=True)
    show_change = st.radio(
        "Show competitor rate change vs.",
        ["None", "Yesterday", "3 days ago", "7 days ago"],
        help="Adds a 'Change' column beside each competitor, pulled from the "
             "matching 'vs.' tab in the rate-shop file.",
    )
    show_own_change = st.checkbox(
        "Show our BAR change (7 days)", value=True,
        help="Adds a 'BAR Chg 7d' column showing how our own rate moved over the "
             "last 7 days (from the rate-shop file).",
    )

    st.markdown("---")
    st.header("3 - Signal sensitivity")
    st.caption("The Compton prices as the market leader (#1 in the comp set). "
               "The ⚡ Action column is purely about rate-leadership position.")
    with st.expander("🔴 Rate leadership (#1 vs #2)", expanded=False):
        st.caption("🔴 Not #1 = a competitor is at or above our BAR.  "
                   "🟠 Too close to #2 = we're #1, but by less than the cushion below.")
        lead_buffer = st.slider("Minimum $ cushion we want over the #2 (top competitor)",
                                5, 150, 20, 5,
                                help="If we're #1 but ahead by less than this, flag "
                                     "'Too close to #2'. If a competitor is at/above our "
                                     "BAR, that's always flagged as 'Not #1' regardless "
                                     "of this slider.")
    with st.expander("🔥 Spike alert", expanded=False):
        st.caption("This self-calibrates to this property's own pickup pattern instead of "
                   "a fixed number — it flags days with unusually high pickup relative to "
                   "everything else in the current report.")
        pickup_pctile = st.slider("Flag pickup at or above this percentile of all pickup days",
                                  50, 99, 75, 1,
                                  help="Lower = flags MORE days (e.g. 75 flags the busiest "
                                       "quarter of pickup days). Higher = flags fewer, more "
                                       "extreme outliers.")
        pickup_min = st.slider("...but never flag below this many rooms (floor)",
                               1, 30, 3, 1,
                               help="Keeps quiet reports from flagging tiny 1-room pickups "
                                    "just because they're 'relatively' high.")

    thr = {"lead_buffer": lead_buffer,
           "pickup_pctile": pickup_pctile, "pickup_min": pickup_min}

    st.markdown("---")
    st.caption("Budget & forecast are pulled automatically from the Data Extract "
               "(Budget / Forecast columns) into the KPI Summary tab.")
    st.caption("Tip: leave the file names as SynXis exports them — the app "
               "auto-detects each report by its sheets.")

c1, c2, c3, c4 = st.columns(4)
c1.metric("PCDC",         "OK" if parts["pcdc"] is not None else "-")
c2.metric("Data Extract", "OK" if parts["data"] is not None else "-")
c3.metric("Market Seg",   "OK" if parts["mseg"] is not None else "-")
c4.metric("Rate Shop",    "OK" if parts["shop"] is not None else "-")

st.caption(f"📅 'Days Left' counts from **{pd.Timestamp(as_of):%b %d, %Y}** "
           f"(sidebar 'Report as-of date'). Defaults to today; change it in the "
           f"sidebar if you want a different reference date.")

report = build_report(parts, as_of, comp_method, show_change, thr, show_own_change)

# =============================================================== #
#  Monthly recap (owner view) — current / next / following month
# =============================================================== #
st.subheader("Monthly Recap — On the Books vs Forecast vs Budget vs Last Year")
_raw = parts.get("data")
if _raw is not None and "Occupancy Date" in getattr(_raw, "columns", []):
    labels, metrics, fc_kind = compute_monthly_recap(_raw, as_of)

    _pct_metrics = {"Occupancy %"}
    _money_metrics = {"ADR", "RevPAR", "Room Revenue"}

    def _fmt(metric, v):
        if not isinstance(v, (int, float)) or pd.isna(v):
            return "–"
        if metric in _pct_metrics:
            return f"{v*100:,.1f}%"
        if metric in _money_metrics:
            return f"${v:,.0f}"
        return f"{v:,.0f}"

    def _fmt_var(metric, v):
        if not isinstance(v, (int, float)) or pd.isna(v):
            return "–"
        if metric in _pct_metrics:
            return f"{v*100:+,.1f} pts"
        if metric in _money_metrics:
            return f"${v:+,.0f}"
        return f"{v:+,.0f}"

    # one table per month, metrics as rows, scenarios + variances as columns
    cols_order = ["On the Books", "Forecast", "Budget", "Last Year",
                  "Fcst vs Budget", "OTB vs Budget"]
    for mi, label in enumerate(labels):
        rows = {}
        for metric, per_month in metrics.items():
            m = per_month[mi]
            fvb = (m["fcst"] - m["budget"]) if pd.notna(m["fcst"]) and pd.notna(m["budget"]) else np.nan
            ovb = (m["otb"] - m["budget"]) if pd.notna(m["otb"]) and pd.notna(m["budget"]) else np.nan
            rows[metric] = {
                "On the Books": _fmt(metric, m["otb"]),
                "Forecast": _fmt(metric, m["fcst"]),
                "Budget": _fmt(metric, m["budget"]),
                "Last Year": _fmt(metric, m["stly"]),
                "Fcst vs Budget": _fmt_var(metric, fvb),
                "OTB vs Budget": _fmt_var(metric, ovb),
                "_fvb": fvb, "_ovb": ovb,
            }
        tdf = pd.DataFrame(rows).T[cols_order]

        def _color_var(col):
            key = "_fvb" if col.name == "Fcst vs Budget" else "_ovb"
            styles = []
            for metric in col.index:
                raw = rows[metric][key]
                if isinstance(raw, (int, float)) and pd.notna(raw):
                    styles.append("color:#1a7f37;font-weight:600" if raw >= 0 else "color:#cf222e;font-weight:600")
                else:
                    styles.append("")
            return styles

        st.markdown(f"**{label}**")
        sty = tdf.style.apply(_color_var, subset=["Fcst vs Budget"]) \
                       .apply(_color_var, subset=["OTB vs Budget"])
        st.dataframe(sty, use_container_width=True)

    st.caption(f"Forecast source: {fc_kind}. RevPAR & Occupancy use physical capacity "
               "(available room-nights). Budget, forecast, and last-year all pulled from the "
               "Data Extract. The Excel export's KPI Summary tab adds the Group vs Transient split.")
else:
    st.info("Upload the Data Extract with Budget/Forecast columns to see the monthly recap.")

# =============================================================== #
#  Filters
# =============================================================== #
st.subheader("Filters")
min_d, max_d = report["Date"].min(), report["Date"].max()

_window_options = ["Custom", "Next 7 days", "Next 14 days", "Next 30 days",
                  "Next 60 days", "Next 90 days", "All future dates", "All dates (incl. past)"]
_default_window_idx = 7
if _pcdc_min is not None:
    _window_options = ["PCDC report range"] + _window_options
    _default_window_idx = 0     # default to exactly the dates covered by PCDC

fa, fb, fc = st.columns([1.2, 1, 1])
window = fa.selectbox("Booking window", _window_options, index=_default_window_idx,
                      help="'PCDC report range' (default, when a PCDC file is uploaded) "
                           "shows exactly the dates covered by your PCDC pace report — "
                           f"{_pcdc_min:%b %d} to {_pcdc_max:%b %d, %Y}"
                           if _pcdc_min is not None else
                           "'All future dates' or any 'Next X days' option will HIDE "
                           "dates before your as-of date (negative Days Left) — use "
                           "the look-back option below to also see recent past dates.")
dow_pick = fb.multiselect("Day of week", ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"],
                          default=["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"])
action_pick = fc.multiselect("Action", ["🔴 Not #1", "🟠 Too close to #2", "🟠 At ceiling", "✅ Leading"],
                             default=["🔴 Not #1", "🟠 Too close to #2", "🟠 At ceiling", "✅ Leading"])

lookback = st.number_input(
    "Also include the last N days before the as-of date (negative Days Left)",
    min_value=0, max_value=90, value=0, step=1,
    help="0 = forward-looking only (default). Set this above 0 to also pull in recent "
         "past dates alongside 'Next X days' / 'All future dates', e.g. to review last "
         "week's pickup. Ignored when 'Custom' or 'All dates (incl. past)' is selected.")

fd, fe, ff, fg = st.columns([1, 1, 1, 1])
events_only = fd.checkbox("Events only", value=False)
occ_band = fe.select_slider("Occ Forecast % band",
                            options=[0, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100],
                            value=(0, 100))
var_thresh = ff.number_input("Min |variance vs STLY| (rooms)", min_value=0, value=0, step=1)
pickup_only = fg.checkbox("🔥 Spike days only", value=False)

# custom date range (used when window == Custom)
start, end = min_d, max_d          # defaults so the filename always has a range
if window == "Custom":
    d1, d2 = st.columns(2)
    default_start = max(min_d, as_of)
    start = d1.date_input("From", value=default_start, min_value=min_d, max_value=max_d)
    end   = d2.date_input("To",   value=max_d,         min_value=min_d, max_value=max_d)

# ---- apply filters ---- #
v = report.copy()
win_map = {"Next 7 days": 7, "Next 14 days": 14, "Next 30 days": 30,
           "Next 60 days": 60, "Next 90 days": 90}
floor_days = -int(lookback)          # e.g. lookback=7 -> also show Days Left down to -7
if window == "PCDC report range" and _pcdc_min is not None:
    v = v[(v["Date"] >= _pcdc_min) & (v["Date"] <= _pcdc_max)]
elif window in win_map:
    v = v[(v["Days Left"] >= floor_days) & (v["Days Left"] <= win_map[window])]
elif window == "All future dates":
    v = v[v["Days Left"] >= floor_days]
elif window == "All dates (incl. past)":
    pass  # no Days Left filter at all — show everything in the uploaded range
else:  # Custom
    v = v[(v["Date"] >= start) & (v["Date"] <= end)]

dow_full = {"Mon": "Monday", "Tue": "Tuesday", "Wed": "Wednesday", "Thu": "Thursday",
            "Fri": "Friday", "Sat": "Saturday", "Sun": "Sunday"}
keep_dow = [dow_full[d] for d in dow_pick]
v = v[v["DOW"].isin(keep_dow)]

if events_only:
    v = v[v["Events"].astype(str).str.strip() != ""]

if "⚡ Action" in v:
    a = v["⚡ Action"].fillna("")
    sel = pd.Series(False, index=v.index)
    if "🔴 Not #1" in action_pick:          sel |= a.str.contains("Not #1")
    if "🟠 Too close to #2" in action_pick: sel |= a.str.contains("Too close")
    if "🟠 At ceiling" in action_pick:      sel |= a.str.contains("At ceiling")
    if "✅ Leading" in action_pick:         sel |= a.str.contains("Leading")
    sel |= (a.str.strip() == "")      # keep past/blank rows
    v = v[sel]

ofp_col = pd.to_numeric(v.get("Occ Forecast % | Total Hotel"), errors="coerce") * 100
lo, hi = occ_band
v = v[(ofp_col.isna()) | ((ofp_col >= lo) & (ofp_col <= hi))]

if var_thresh > 0 and "Rooms OTB STLY | Total Hotel | Change" in v:
    var_col = pd.to_numeric(v["Rooms OTB STLY | Total Hotel | Change"], errors="coerce").abs()
    v = v[var_col >= var_thresh]

if pickup_only and "🔥 Spike" in v:
    v = v[v["🔥 Spike"].astype(str).str.strip() != ""]

view = v.reset_index(drop=True)

st.subheader(f"Pick-Up Report — {len(view)} dates")

pct_cols   = [c for c in view.columns if c.startswith("Occ Forecast %")]
shop_all   = [c for c in view.columns if c.startswith("Competitor Shops")]
shop_chg   = [c for c in shop_all if c.endswith("Change")]
shop_cur   = [c for c in shop_all if c not in shop_chg]
money_cols = (["BAR", "Comp Set Avg", "Last Room Value", "Estimated ADR"]
              + [c for c in view.columns if c.startswith("Booked ADR")])
rooms_chg  = [c for c in view.columns
              if (c.startswith("Rooms Sold") or c.startswith("Rooms OTB STLY"))
              and c.endswith("Change")]
var_cols   = ([c for c in view.columns if "Variance" in c] + shop_chg + rooms_chg
              + [c for c in ("Lead Gap vs Comp", "BAR Chg 7d") if c in view.columns])
otb_sold   = "Rooms Left to Sell"      # heat-map column

def _money(v):
    if isinstance(v, (int, float)) and not pd.isna(v):
        return f"${v:,.0f}"
    return "" if v is None or (isinstance(v, float) and pd.isna(v)) else str(v)

def _signed(v):
    if isinstance(v, (int, float)) and not pd.isna(v) and v != 0:
        return f"{v:+,.0f}"
    return ""

fmt = {}
for c in pct_cols:
    fmt[c] = "{:.1%}"
for c in money_cols:
    fmt[c] = "${:,.0f}"
for c in shop_cur:
    fmt[c] = _money
for c in shop_chg:
    fmt[c] = _signed
for c in ("Lead Gap vs Comp", "BAR Chg 7d"):
    if c in view.columns:
        fmt[c] = _signed
for c in view.columns:
    if c in ("Days Left", "OOO", "Ovrbk") or c.startswith(("Rooms Sold", "Rooms OTB STLY",
             "Remaining Demand", "Occ Forecast |")) and c not in pct_cols:
        fmt.setdefault(c, "{:,.0f}")

styler = (view.style
          .format(fmt, na_rep="")
          .map(lambda v: "color:#1a7f37;font-weight:600" if isinstance(v, (int, float)) and v > 0
               else ("color:#cf222e;font-weight:600" if isinstance(v, (int, float)) and v < 0 else ""),
               subset=var_cols))


# Heat map on the OTB rooms sold column — pure-Python red->yellow->green
# gradient (no matplotlib needed, so it works on Streamlit Cloud out of the box).
def _heat_bg(series):
    # For "Rooms Left to Sell": GREEN = lots of rooms left (wide open),
    # RED = few rooms left (near sell-out).  Low value -> red, high value -> green.
    vals = pd.to_numeric(series, errors="coerce")
    lo, hi = vals.min(), vals.max()
    span = (hi - lo) or 1
    out = []
    for v in vals:
        if pd.isna(v):
            out.append("")
            continue
        t = (v - lo) / span                    # low rooms left -> red, high -> green
        if t < 0.5:
            f = t / 0.5
            red, grn, blu = 248, int(105 + f * (235 - 105)), int(107 + f * (132 - 107))
        else:
            f = (t - 0.5) / 0.5
            red, grn, blu = int(255 - f * (255 - 99)), int(235 - f * (235 - 190)), int(132 - f * (132 - 123))
        out.append(f"background-color: #{red:02X}{grn:02X}{blu:02X}")
    return out

# Heat map is applied directly inside the pick-up report on the OTB column
if otb_sold in view.columns and view[otb_sold].notna().any():
    styler = styler.apply(_heat_bg, subset=[otb_sold])

# Colour the ⚡ Action column so recommendations pop
def _action_bg(series):
    out = []
    for v in series.astype(str):
        if "Not #1" in v:
            out.append("background-color:#F8D7DA;color:#842029;font-weight:600")
        elif "Too close" in v:
            out.append("background-color:#FFF3CD;color:#664D03;font-weight:600")
        elif "At ceiling" in v:
            out.append("background-color:#FFE5B4;color:#7A4A00;font-weight:600")
        elif "Leading" in v:
            out.append("background-color:#D1E7DD;color:#0F5132;font-weight:600")
        else:
            out.append("")
    return out

if "⚡ Action" in view.columns:
    styler = styler.apply(_action_bg, subset=["⚡ Action"])

# Colour the 🔥 Spike column so big pickup days pop
def _pickup_bg(series):
    return ["background-color:#FDE7CE;color:#B45309;font-weight:600" if "Spike" in str(v)
            else "" for v in series]

if "🔥 Spike" in view.columns:
    styler = styler.apply(_pickup_bg, subset=["🔥 Spike"])

st.dataframe(styler, use_container_width=True, height=560)

# filename reflects the actual filtered date range (falls back to the full range)
if len(view) > 0:
    _fn_start, _fn_end = view["Date"].min(), view["Date"].max()
else:
    _fn_start, _fn_end = start, end

st.download_button(
    "Download report as Excel",
    data=to_excel(view, {"data_raw": parts.get("data"), "as_of": as_of}),
    file_name=f"Compton_PickUp_{_fn_start:%Y%m%d}_{_fn_end:%Y%m%d}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

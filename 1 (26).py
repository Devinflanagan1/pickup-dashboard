"""
The Compton — Daily Pick-Up Report Builder
Run it with:   streamlit run app.py

Two views (switch with the "View" radio at the top of the sidebar):
  1. Pick-Up Report (DD) — the original daily pick-up report (unchanged).
  2. Dashboard          — reproduces the workbook 'Dashboard' tab: Total /
                          Transient / Group on-the-books by month vs Last
                          Report, Budget and Freeze Forecast.
"""

import io
import os
import json
import calendar
import numpy as np
import pandas as pd
import streamlit as st

st.set_page_config(page_title="Compton Daily Pick-Up Report", layout="wide")


def _num(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return np.nan


def _read_all_sheets(uploaded_file):
    xls = pd.ExcelFile(uploaded_file, engine="openpyxl")
    return {s: xls.parse(s, header=None) for s in xls.sheet_names}


def detect_and_load(uploaded_files):
    found = {"pcdc": None, "data": None, "mseg": None, "shop": None,
             "shop_chg_1": None, "shop_chg_3": None, "shop_chg_7": None,
             "pcdc_asof": None, "mseg_raw": None, "mseg_file": None}

    for f in uploaded_files:
        try:
            sheets = _read_all_sheets(f)
        except Exception as e:
            st.warning(f"Could not open **{f.name}** ({e}).")
            continue
        names = [s.lower() for s in sheets.keys()]

        if any("rate" in n for n in names) and any("overview" in n for n in names):
            raw = sheets[[s for s in sheets if s.lower() == "rates"][0]]
            found["shop"] = _parse_rateshop(raw)
            for tab, key in [("vs. yesterday", "shop_chg_1"),
                             ("vs. 3 days ago", "shop_chg_3"),
                             ("vs. 7 days ago", "shop_chg_7")]:
                match = [s for s in sheets if s.lower() == tab]
                if match:
                    found[key] = _parse_shop_change(sheets[match[0]])

        elif any("changereport" in n for n in names):
            raw = sheets[[s for s in sheets if "changereport" in s.lower()][0]]
            found["pcdc"] = _parse_pcdc(raw)
            crit = [s for s in sheets if s.lower() == "report criteria"]
            if crit:
                found["pcdc_asof"] = _parse_pcdc_asof(sheets[crit[0]])

        elif any("market segment" in n for n in names):
            raw = sheets[[s for s in sheets if s.lower() == "market segment"][0]]
            found["mseg"] = _parse_mseg(raw)
            found["mseg_raw"] = raw          # keep raw sheet for the Dashboard page
            found["mseg_file"] = getattr(f, "name", "")  # remember source filename

        elif any(n == "property" for n in names):
            raw = sheets[[s for s in sheets if s.lower() == "property"][0]]
            found["data"] = _parse_data(raw)

        else:
            st.warning(f"Didn't recognise **{f.name}** — skipping. "
                       f"(sheets: {', '.join(sheets.keys())})")
    return found


def _parse_data(raw):
    df = raw.copy()
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)
    df["Occupancy Date"] = pd.to_datetime(df["Occupancy Date"], errors="coerce")
    df = df.dropna(subset=["Occupancy Date"])
    return df


def _parse_pcdc(raw):
    df = raw.iloc[3:].copy()
    df[0] = pd.to_datetime(df[0], errors="coerce")
    df = df.dropna(subset=[0]).reset_index(drop=True)
    out = pd.DataFrame({"Occupancy Date": df[0]})
    out["pc_event"]      = df[2]
    out["trans_cur"]     = pd.to_numeric(df[3], errors="coerce")
    out["trans_chg"]     = pd.to_numeric(df[4], errors="coerce")
    out["grp_cur"]       = pd.to_numeric(df[5], errors="coerce")
    out["grp_chg"]       = pd.to_numeric(df[6], errors="coerce")
    out["grp_blocked"]   = pd.to_numeric(df[7], errors="coerce")
    out["grp_avail"]     = pd.to_numeric(df[8], errors="coerce")
    out["grp_pickup"]    = pd.to_numeric(df[9], errors="coerce")
    out["fcst_trans"]    = pd.to_numeric(df[10], errors="coerce")
    out["fcst_grp"]      = pd.to_numeric(df[12], errors="coerce")
    out["rev_trans"]     = pd.to_numeric(df[14], errors="coerce")
    out["rev_grp"]       = pd.to_numeric(df[16], errors="coerce")
    out["adr_trans"]     = pd.to_numeric(df[22], errors="coerce")
    out["adr_grp"]       = pd.to_numeric(df[24], errors="coerce")
    return out


def _parse_pcdc_asof(raw):
    """
    Read the PCDC 'Report Criteria' sheet and return the reference date for
    'Days Left' (i.e. "today" for pace purposes).

    'Generated On' is the actual timestamp the report was run/exported, so it's
    the most reliable proxy for "today". 'Analysis Start Date' and 'Activity
    Start Date' are report-WINDOW parameters the revenue manager can set to any
    date (e.g. to look at pace from a few days back) — they are NOT guaranteed
    to equal today, so they're only used as a fallback if 'Generated On' is
    missing. Returns a datetime.date or None.
    """
    try:
        # find the header row that contains the labels, then the value row below it
        for i in range(len(raw) - 1):
            rowvals = [str(x).strip().lower() for x in raw.iloc[i].tolist()]
            if "analysis start date" in rowvals or "generated on" in rowvals:
                header = raw.iloc[i].tolist()
                values = raw.iloc[i + 1].tolist()
                lut = {str(h).strip().lower(): v for h, v in zip(header, values)}
                for key in ("generated on", "activity start date", "analysis start date"):
                    if key in lut and lut[key] not in (None, ""):
                        d = pd.to_datetime(lut[key], errors="coerce")
                        if pd.notna(d):
                            return d.date()
    except Exception:
        pass
    return None


def _parse_mseg(raw):
    df = raw.copy()
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)
    df["Occupancy Date"] = pd.to_datetime(df["Occupancy Date"], errors="coerce")
    df = df.dropna(subset=["Occupancy Date"])
    df["Occupancy On Books This Year"] = pd.to_numeric(
        df["Occupancy On Books This Year"], errors="coerce").fillna(0)
    df["is_trans"] = df["Market Segment"].astype(str).str.startswith("Transient")
    agg = df.groupby("Occupancy Date").apply(
        lambda g: pd.Series({
            "ms_trans": g.loc[g["is_trans"], "Occupancy On Books This Year"].sum(),
            "ms_group": g.loc[~g["is_trans"], "Occupancy On Books This Year"].sum(),
        })
    ).reset_index()
    return agg


def _clean_shop(x):
    if x is None:
        return None
    v = _num(x)
    if not np.isnan(v):
        return int(v) if float(v).is_integer() else v
    s = str(x).strip()
    return s if s else None


def _parse_rateshop(raw):
    df = raw.iloc[5:].copy()
    df[2] = pd.to_datetime(df[2], errors="coerce")
    df = df.dropna(subset=[2]).reset_index(drop=True)
    out = pd.DataFrame({"Occupancy Date": df[2]})
    out["own_bar"] = df[4].map(_num)
    for key, col in [("c_21c", 5), ("c_motto", 6), ("c_ac", 7), ("c_dt", 8)]:
        out[key] = df[col].map(_clean_shop)
        out[key + "_n"] = df[col].map(_num)
    return out


def _parse_shop_change(raw):
    df = raw.iloc[5:].copy()
    df[2] = pd.to_datetime(df[2], errors="coerce")
    df = df.dropna(subset=[2]).reset_index(drop=True)
    out = pd.DataFrame({"Occupancy Date": df[2]})
    # our own BAR change sits in column 6 (right of our rate in column 5)
    out["own_chg"] = df[6].map(_num) if 6 in df.columns else np.nan
    for key, chg_col in [("c_21c", 8), ("c_motto", 10), ("c_ac", 12), ("c_dt", 14)]:
        out[key + "_chg"] = df[chg_col].map(_num) if chg_col in df.columns else np.nan
    return out


def compute_monthly_recap(raw, as_of):
    """Owner-grade monthly recap. Returns (months, metrics) where months is a list
    of 'Mon YYYY' labels and metrics is a dict keyed by metric name -> list of
    per-month dicts {otb, fcst, budget, stly}. All pulled from the Data Extract."""
    import calendar as _cal
    d = raw.copy()
    d["Occupancy Date"] = pd.to_datetime(d["Occupancy Date"], errors="coerce")
    d = d.dropna(subset=["Occupancy Date"])

    def col(name):
        return pd.to_numeric(d[name], errors="coerce") if name in d.columns else pd.Series(0.0, index=d.index)

    myf = col("My Forecast Revenue This Year")
    fc_rev_src = "My Forecast Revenue This Year" if myf.fillna(0).abs().sum() > 0 else "Forecasted Room Revenue This Year"
    myfo = col("My Forecast Occupancy - Total This Year")
    fc_occ_src = "My Forecast Occupancy - Total This Year" if myfo.fillna(0).abs().sum() > 0 else "Occupancy Forecast - Total This Year"
    fc_kind = "Your forecast" if fc_rev_src.startswith("My") else "System forecast"

    d["_avail"]   = col("Physical Capacity This Year")
    d["_otb_rev"] = col("Booked Room Revenue This Year")
    d["_otb_rms"] = col("Occupancy On Books This Year")
    d["_fc_rev"]  = col(fc_rev_src)
    d["_fc_rms"]  = col(fc_occ_src)
    d["_bud_rev"] = col("Budget Room Revenue This Year")
    d["_bud_rms"] = col("Budget Occupancy - Total This Year")
    d["_sy_rev"]  = col("Booked Room Revenue STLY")
    d["_sy_rms"]  = col("Occupancy On Books STLY")
    d["_ym"] = list(zip(d["Occupancy Date"].dt.year, d["Occupancy Date"].dt.month))

    base = pd.Timestamp(as_of)
    yms = [((base + pd.DateOffset(months=k)).year, (base + pd.DateOffset(months=k)).month) for k in range(3)]
    labels = [f"{_cal.month_abbr[m]} {y}" for y, m in yms]

    def scen(sub, rev_c, rms_c):
        # cast to plain Python float so downstream isinstance()/formatting is reliable
        rev = float(sub[rev_c].sum()); rms = float(sub[rms_c].sum()); avail = float(sub["_avail"].sum())
        return {
            "Rooms Sold": rms,
            "Occupancy %": (rms / avail) if avail else np.nan,
            "ADR": (rev / rms) if rms else np.nan,
            "RevPAR": (rev / avail) if avail else np.nan,
            "Room Revenue": rev,
        }

    metric_names = ["Rooms Sold", "Occupancy %", "ADR", "RevPAR", "Room Revenue"]
    metrics = {mn: [] for mn in metric_names}
    for ym in yms:
        sub = d[d["_ym"] == ym]
        otb = scen(sub, "_otb_rev", "_otb_rms")
        fcst = scen(sub, "_fc_rev", "_fc_rms")
        bud = scen(sub, "_bud_rev", "_bud_rms")
        stly = scen(sub, "_sy_rev", "_sy_rms")
        for mn in metric_names:
            metrics[mn].append({"otb": otb[mn], "fcst": fcst[mn], "budget": bud[mn], "stly": stly[mn]})
    return labels, metrics, fc_kind


def build_report(parts, as_of, comp_method="Average", show_change="None", thr=None,
                 show_own_change=True):
    data, pcdc, mseg, shop = parts["data"], parts["pcdc"], parts["mseg"], parts["shop"]
    if data is None:
        st.error("The **Data Extract** file is required (couldn't find a 'Property' sheet).")
        st.stop()

    df = data.copy()
    for extra in (pcdc, mseg, shop):
        if extra is not None:
            df = df.merge(extra, on="Occupancy Date", how="left")

    chg_key = {"Yesterday": "shop_chg_1", "3 days ago": "shop_chg_3",
               "7 days ago": "shop_chg_7"}.get(show_change)
    chg_df = parts.get(chg_key) if chg_key else None
    if chg_df is not None:
        # avoid duplicate own_chg if we also pull the 7-day own change below
        df = df.merge(chg_df.drop(columns=[c for c in ["own_chg"] if c in chg_df], errors="ignore"),
                      on="Occupancy Date", how="left")

    # Always pull OUR OWN 7-day BAR change (independent of the competitor toggle)
    own7 = parts.get("shop_chg_7")
    if own7 is not None and "own_chg" in own7:
        df = df.merge(own7[["Occupancy Date", "own_chg"]].rename(columns={"own_chg": "bar_chg_7d"}),
                      on="Occupancy Date", how="left")

    # Optional 7-day PCDC (a change report run over a 7-day window): its per-date
    # transient/group room change IS the day-by-day 7-day pickup. Merge it in with
    # distinct names so it doesn't collide with the daily PCDC change columns.
    pcdc7 = parts.get("pcdc7")
    has_wk_pu = pcdc7 is not None and "Occupancy Date" in getattr(pcdc7, "columns", [])
    if has_wk_pu:
        w = pcdc7[["Occupancy Date", "trans_chg", "grp_chg"]].rename(
            columns={"trans_chg": "w_trans_chg", "grp_chg": "w_grp_chg"})
        df = df.merge(w, on="Occupancy Date", how="left")

    def dnum(col):
        return pd.to_numeric(df.get(col), errors="coerce")

    cap        = dnum("Physical Capacity This Year")
    occ_ty     = dnum("Occupancy On Books This Year")
    occ_stly   = dnum("Occupancy On Books STLY")
    tr_ty      = dnum("Rooms Sold - Transient This Year")
    tr_stly    = dnum("Rooms Sold - Transient STLY")
    gr_ty      = dnum("Rooms Sold - Group This Year")
    gr_stly    = dnum("Rooms Sold - Group STLY")
    dem_tot    = dnum("User Total Demand - Total This Year")
    dem_grp    = dnum("User Constrained Total Demand - Group This Year")
    dem_tr     = dnum("User Unconstrained Total Demand - Transient This Year")

    r = pd.DataFrame()
    r["DOW"]  = df["Day of Week"]
    r["Date"] = df["Occupancy Date"].dt.date
    r["Days Left"] = (df["Occupancy Date"] - pd.Timestamp(as_of)).dt.days

    ev = df.get("Special Event This Year")
    ev = ev.fillna("") if ev is not None else ""
    if "pc_event" in df:
        ev = ev.where(ev.astype(str).str.strip() != "", df["pc_event"].fillna(""))
    r["Events"] = pd.Series(ev, index=df.index).replace("nan", "")

    r["Rooms Left to Sell"] = dnum("Remaining Capacity This Year")
    r["BAR"]                = dnum("BAR")
    if show_own_change:
        r["BAR Chg 7d"]     = dnum("bar_chg_7d")     # our own rate move over last 7 days

    if "c_21c_n" in df:
        comps_n = df[["c_21c_n", "c_motto_n", "c_ac_n", "c_dt_n"]]
        r["Comp Set Avg"] = (comps_n.mean(axis=1) if comp_method == "Average"
                             else comps_n.median(axis=1)).round(0)
    else:
        r["Comp Set Avg"] = np.nan

    r["Last Room Value"] = dnum("Last Room Value This Year")

    add_chg = chg_df is not None
    for name, key in [("21c Museum Hotel Bentonville - MGallery", "c_21c"),
                      ("Motto By Hilton Bentonville Downtown",    "c_motto"),
                      ("AC Hotel by Marriott Bentonville",        "c_ac"),
                      ("DoubleTree Suites by Hilton Bentonville", "c_dt")]:
        base = "Competitor Shops | " + name
        if add_chg:
            r[base + " | Current"] = df.get(key)
            r[base + " | Change"]  = df.get(key + "_chg")
        else:
            r[base] = df.get(key)

    r["OOO"]   = dnum("Rooms N/A - Out of Order This Year")
    r["Ovrbk"] = dnum("Overbooking This Year")

    tc, tch = df.get("trans_cur"), df.get("trans_chg")
    gc, gch = df.get("grp_cur"),   df.get("grp_chg")
    r["Rooms Sold | Total Hotel | Current"]   = (tc.fillna(0) + gc.fillna(0)) if tc is not None else occ_ty
    r["Rooms Sold | Total Hotel | Change"]    = (tch.fillna(0) + gch.fillna(0)) if tch is not None else np.nan
    r["Rooms Sold | Total Transient | Current"] = tc
    r["Rooms Sold | Total Transient | Change"]  = tch
    r["Rooms Sold | Total Group | Current"]     = gc
    r["Rooms Sold | Total Group | Change"]      = gch
    r["Rooms Sold | Total Group | Blocked"]     = df.get("grp_blocked")
    r["Rooms Sold | Total Group | P/U"]         = df.get("grp_pickup")
    r["Rooms Sold | Total Group | Remaining"]   = df.get("grp_avail")

    # 7-day pickup per occupancy date (from the optional 7-day PCDC)
    if has_wk_pu:
        wtc = pd.to_numeric(df.get("w_trans_chg"), errors="coerce")
        wgc = pd.to_numeric(df.get("w_grp_chg"), errors="coerce")
        r["7-Day P/U | Total Hotel"]     = wtc.fillna(0) + wgc.fillna(0)
        r["7-Day P/U | Total Transient"] = wtc
        r["7-Day P/U | Total Group"]     = wgc

    r["Rooms OTB STLY | Total Hotel | Current"]     = occ_stly
    r["Rooms OTB STLY | Total Hotel | Change"]      = occ_ty - occ_stly
    r["Rooms OTB STLY | Total Transient | Current"] = tr_stly
    r["Rooms OTB STLY | Total Transient | Change"]  = tr_ty - tr_stly
    r["Rooms OTB STLY | Total Group | Current"]     = gr_stly
    r["Rooms OTB STLY | Total Group | Change"]      = gr_ty - gr_stly

    r["Remaining Demand | Total Hotel"]     = (dem_tot - occ_ty).clip(lower=0)
    r["Remaining Demand | Total Transient"] = (dem_tr - tr_ty).clip(lower=0)
    r["Remaining Demand | Total Group"]     = (dem_grp - gr_ty).clip(lower=0)

    ft, fg = df.get("fcst_trans"), df.get("fcst_grp")
    fcst_tot = (ft.fillna(0) + fg.fillna(0)) if ft is not None else dnum("Forecasted Room Revenue This Year")*np.nan
    r["Occ Forecast | Total Hotel"]     = fcst_tot.round(1) if ft is not None else np.nan
    r["Occ Forecast | Total Transient"] = ft.round(1) if ft is not None else np.nan
    r["Occ Forecast | Total Group"]     = fg.round(1) if fg is not None else np.nan

    r["Occ Forecast % | Total Hotel"]     = (fcst_tot / cap) if ft is not None else np.nan
    r["Occ Forecast % | Total Transient"] = (ft / cap) if ft is not None else np.nan
    r["Occ Forecast % | Total Group"]     = (fg / cap) if fg is not None else np.nan

    rev_t, rev_g = df.get("rev_trans"), df.get("rev_grp")
    if tc is not None:
        total_rooms = (tc.fillna(0) + gc.fillna(0)).replace(0, np.nan)
        r["Booked ADR | Total Hotel"] = ((rev_t.fillna(0) + rev_g.fillna(0)) / total_rooms).round(2)
    else:
        r["Booked ADR | Total Hotel"] = dnum("ADR On Books This Year")
    r["Booked ADR | Total Transient"] = df.get("adr_trans")
    r["Booked ADR | Total Group"]     = df.get("adr_grp")

    r["Estimated ADR"] = dnum("ADR Forecast This Year")

    # ================================================================= #
    #  PRICING DECISION SIGNALS  (VP-of-Revenue toolkit)
    # ================================================================= #
    bar   = r["BAR"]
    comp  = r["Comp Set Avg"]
    rl    = r["Rooms Left to Sell"]
    rd    = r["Remaining Demand | Total Hotel"]
    ofp   = r["Occ Forecast % | Total Hotel"]
    pace  = r["Rooms OTB STLY | Total Hotel | Change"]      # TY − STLY (+ = ahead)
    dleft = r["Days Left"]
    floor   = dnum("Floor")
    ceiling = dnum("Ceiling")

    # --- Rate Rank badge:  where our BAR sits in the comp set (1 = highest) --- #
    rank_cols = [c for c in ("c_21c_n", "c_motto_n", "c_ac_n", "c_dt_n") if c in df]
    if rank_cols:
        price_df = pd.DataFrame({"own": pd.to_numeric(r["BAR"], errors="coerce").values})
        for cc in rank_cols:
            price_df[cc] = pd.to_numeric(df[cc], errors="coerce").values

        def _rank_row(row):
            own = row["own"]
            allv = [v for v in row.tolist() if pd.notna(v)]
            if pd.isna(own) or not allv:
                return ""
            higher = sum(1 for v in allv if v > own)
            return f"{higher + 1} of {len(allv)}"

        r["Rate Rank"] = [_rank_row(price_df.iloc[i]) for i in range(len(price_df))]
    else:
        r["Rate Rank"] = ""

    # --- Signal thresholds (tunable from the sidebar) --- #
    thr = thr or {}
    lead_buffer   = thr.get("lead_buffer", 20)    # min $ cushion over the #2 to be "safely" #1
    pickup_pctile = thr.get("pickup_pctile", 75)  # percentile of positive pickups = "big"
    pickup_min    = thr.get("pickup_min", 3)      # floor so quiet reports don't flag tiny pickup

    # Strategy: The Compton prices as the market LEADER. The Action column is
    # purely about rate-leadership position — two flags only:
    #   🔴 Not #1        — a competitor is at or above our BAR
    #   🟠 Too close #2  — we're #1, but our cushion over the top competitor
    #                      is thinner than we want (within "lead_buffer" $)

    # top competitor rate per date (max of the 4 shops) — effectively "the #2"
    if rank_cols:
        comp_matrix = pd.DataFrame({cc: pd.to_numeric(df[cc], errors="coerce").values
                                    for cc in rank_cols})
        max_comp = comp_matrix.max(axis=1)
        max_comp.index = r.index
    else:
        max_comp = pd.Series(np.nan, index=r.index)

    # gap to the #2 (top competitor): + = we're above them, 0/− = not #1
    lead_gap = (bar - max_comp)
    r["Lead Gap vs Comp"] = lead_gap.round(0)

    not_leader = pd.notna(max_comp) & (lead_gap <= 0)             # someone matches/beats us
    too_close  = pd.notna(max_comp) & (lead_gap > 0) & (lead_gap < lead_buffer)
    at_ceiling = bar >= ceiling

    def _label(i):
        if pd.isna(dleft.iloc[i]) or dleft.iloc[i] < 0:
            return ""                                     # past date
        if pd.isna(max_comp.iloc[i]):
            return ""                                     # no comp-set data for this date
        if bool(not_leader.iloc[i]):
            return "🔴 Not #1" if not bool(at_ceiling.iloc[i]) else "🟠 Not #1 · at ceiling"
        if bool(too_close.iloc[i]):
            return "🟠 Too close to #2" if not bool(at_ceiling.iloc[i]) else "🟠 At ceiling"
        return "✅ Leading"

    r["⚡ Action"] = [_label(i) for i in range(len(r))]

    # --- 🔥 Spike: flags an unusually large day's pickup ------------- #
    # "Big" is relative, not a fixed number, so this self-calibrates off this
    # property's own pickup pattern: it flags a date when its pickup sits at or
    # above the Nth percentile of all POSITIVE daily pickups in this report
    # (default 90th = top 10%), with a small floor so a quiet report full of
    # 1-2 room days doesn't get flagged just because it's "relatively" high.
    pickup = r["Rooms Sold | Total Hotel | Change"]
    pos_pickup = pickup[pickup > 0]
    if len(pos_pickup) >= 5:
        pctile_val = pos_pickup.quantile(pickup_pctile / 100)
        pickup_threshold = max(pctile_val, pickup_min)
    else:
        pickup_threshold = pickup_min          # not enough data to calibrate — use the floor

    is_big_pickup = (pickup >= pickup_threshold) & (dleft >= 0)

    def _pickup_label(i):
        if bool(is_big_pickup.iloc[i]):
            v = pickup.iloc[i]
            return f"🔥 Spike (+{v:,.0f})"
        return ""

    r["🔥 Spike"] = [_pickup_label(i) for i in range(len(r))]

    # ---- Enforce the exact original report column order ---- #
    lead = ["⚡ Action", "🔥 Spike", "DOW", "Date", "Days Left", "Events",
            "Rooms Left to Sell", "BAR", "BAR Chg 7d", "Comp Set Avg", "Rate Rank",
            "Lead Gap vs Comp", "Last Room Value"]
    comp = [c for c in r.columns if c.startswith("Competitor Shops")]
    mid  = ["OOO", "Ovrbk"]
    groups = ["Rooms Sold", "7-Day P/U", "Rooms OTB STLY", "Remaining Demand",
              "Occ Forecast |", "Occ Forecast %", "Booked ADR"]
    grouped = []
    for g in groups:
        grouped += [c for c in r.columns if c.startswith(g)]
    tail = ["Estimated ADR"]
    order = lead + comp + mid + grouped + tail
    order = [c for c in order if c in r.columns]          # keep only existing
    order += [c for c in r.columns if c not in order]     # safety: append any strays
    r = r[order]

    return r


def _levels(col):
    parts = [p.strip() for p in str(col).split("|")]
    while len(parts) < 3:
        parts.append("")
    return tuple(parts[:3])


def to_excel(r, kpi_ctx=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    lv = [_levels(c) for c in r.columns]
    n = len(lv)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pick-Up Report"

    NAVY, TEAL, ORANGE, PURPLE, GREEN = "16365C", "215967", "E26B0A", "60497A", "76933C"
    white = Font(color="FFFFFF", bold=True, size=9)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

    def PF(hexcol):
        return PatternFill("solid", fgColor=hexcol)

    def seg_color(txt):
        if "Total Hotel" in txt:
            return ORANGE
        if "Total Transient" in txt:
            return PURPLE
        if "Total Group" in txt:
            return GREEN
        return None

    def band_fill(l1, l2):
        return PF(NAVY) if l2 == "" else PF(TEAL)

    def lvl2_fill(l1, l2):
        sc = seg_color(l2)
        if sc:
            return PF(sc)
        if l1 == "Competitor Shops":
            return PF(NAVY)
        return PF(TEAL)

    def lvl3_fill(l1, l2):
        sc = seg_color(l2)
        return PF(sc) if sc else PF(NAVY)

    j = 0
    while j < n:
        l1 = lv[j][0]
        k = j
        while k + 1 < n and lv[k + 1][0] == l1 and lv[j][1]:
            k += 1
        fill = band_fill(l1, lv[j][1])
        if lv[j][1] == "" and l1 in ("Estimated ADR",):
            fill = PF(ORANGE)
        cell = ws.cell(row=1, column=j + 1, value=l1)
        cell.font = white
        cell.alignment = center
        cell.fill = fill
        if lv[j][1] == "":
            ws.merge_cells(start_row=1, start_column=j + 1, end_row=3, end_column=j + 1)
        elif k > j:
            ws.merge_cells(start_row=1, start_column=j + 1, end_row=1, end_column=k + 1)
        j = k + 1

    j = 0
    while j < n:
        l1, l2, l3 = lv[j]
        if l2:
            k = j
            while k + 1 < n and lv[k + 1][0] == l1 and lv[k + 1][1] == l2:
                k += 1
            cell = ws.cell(row=2, column=j + 1, value=l2)
            cell.font = white
            cell.alignment = center
            cell.fill = lvl2_fill(l1, l2)
            if l3 == "":
                ws.merge_cells(start_row=2, start_column=j + 1, end_row=3, end_column=j + 1)
            elif k > j:
                ws.merge_cells(start_row=2, start_column=j + 1, end_row=2, end_column=k + 1)
            j = k + 1
        else:
            j += 1

    for j, (l1, l2, l3) in enumerate(lv, start=1):
        if l3:
            cell = ws.cell(row=3, column=j, value=l3)
            cell.font = white
            cell.alignment = center
            cell.fill = lvl3_fill(l1, l2)

    _hotelnames = ("Bentonville", "MGallery")
    DATA0 = 4
    pct = {i for i, c0 in enumerate(r.columns) if c0.startswith("Occ Forecast %")}
    signed_cols = ("BAR Chg 7d", "Lead Gap vs Comp")
    chg = {i for i, c0 in enumerate(r.columns)
           if (any(h in c0 for h in _hotelnames) and c0.endswith("Change"))
           or c0 in signed_cols or c0.startswith("7-Day P/U")}
    money = {i for i, c0 in enumerate(r.columns)
             if (c0 in ("BAR", "Comp Set Avg", "Last Room Value", "Estimated ADR")
                 or "ADR" in c0 or any(h in c0 for h in _hotelnames))
             and i not in chg}
    otb_sold_col = None
    for i, c0 in enumerate(r.columns):
        if c0 == "Rooms Left to Sell":
            otb_sold_col = i

    def _action_fill(txt):
        t = str(txt)
        if "Not #1" in t:          return PatternFill("solid", fgColor="F8D7DA")
        if "Too close" in t:       return PatternFill("solid", fgColor="FFF3CD")
        if "At ceiling" in t:      return PatternFill("solid", fgColor="FFE5B4")
        if "Leading" in t:         return PatternFill("solid", fgColor="D1E7DD")
        return None

    def _pickup_fill(txt):
        return PatternFill("solid", fgColor="FDE7CE") if "Spike" in str(txt) else None

    for ri, (_, row) in enumerate(r.iterrows(), start=DATA0):
        for ci2, col in enumerate(r.columns, start=1):
            v = row[col]
            cell = ws.cell(row=ri, column=ci2, value=(None if pd.isna(v) else v))
            cell.border = thin
            cell.font = Font(size=9)
            # center everything; left-align the long-text "Events" column
            halign = "left" if col == "Events" else "center"
            cell.alignment = Alignment(horizontal=halign, vertical="center")
            if col == "⚡ Action":
                af = _action_fill(v)
                if af is not None:
                    cell.fill = af
            elif col == "🔥 Spike":
                pf = _pickup_fill(v)
                if pf is not None:
                    cell.fill = pf
                    cell.font = Font(size=9, bold=True, color="B45309")
            idx = ci2 - 1
            if idx in pct:
                cell.number_format = "0.0%"
            elif idx in chg:
                cell.number_format = "+#,##0;-#,##0;\"\""
            elif idx in money:
                cell.number_format = "$#,##0"
            elif isinstance(v, float) and float(v).is_integer():
                cell.number_format = "#,##0"

    if otb_sold_col is not None and len(r) > 0:
        L = get_column_letter(otb_sold_col + 1)
        rng = f"{L}{DATA0}:{L}{DATA0 + len(r) - 1}"
        # Rooms Left to Sell: RED = few left (near sell-out), GREEN = lots left
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(start_type="min", start_color="F8696B",         # red (few left)
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",  # yellow
                           end_type="max", end_color="63BE7B"))            # green (lots left)

    ws.freeze_panes = "C4"
    for j in range(1, n + 1):
        ws.column_dimensions[get_column_letter(j)].width = 12
    ws.column_dimensions["A"].width = 16   # ⚡ Action
    ws.column_dimensions["E"].width = 22   # Events

    # Excel AutoFilter so the team can filter/sort right in Excel.
    # Anchored on the bottom header row (row 3) through the last data row.
    last_row = DATA0 + len(r) - 1
    ws.auto_filter.ref = f"A3:{get_column_letter(n)}{last_row}"

    # ================================================================= #
    #  KPI Summary tab (self-contained snapshot for sharing)
    # ================================================================= #
    _build_kpi_sheet(wb, r, NAVY, TEAL, ORANGE, PURPLE, GREEN, kpi_ctx)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _month_key(d):
    return (d.year, d.month)


def _month_label(y, m):
    import calendar
    return f"{calendar.month_abbr[m]} {y}"


def _build_kpi_sheet(wb, r, NAVY, TEAL, ORANGE, PURPLE, GREEN, kpi_ctx=None):
    """Owner-grade monthly recap tab: Rooms / Occupancy% / ADR / RevPAR / Revenue,
    compared across On-the-Books, Forecast, Budget and Last Year, for the current /
    next / following month. Plus a Group vs Transient rooms split."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import calendar
    kpi_ctx = kpi_ctx or {}
    as_of = kpi_ctx.get("as_of", pd.Timestamp.today().date())
    raw = kpi_ctx.get("data_raw")

    ks = wb.create_sheet("KPI Summary", 0)
    white = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    thin = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

    ks.merge_cells("A1:G1")
    t = ks["A1"]; t.value = "The Compton - Monthly Owner Recap"
    t.font = Font(color="FFFFFF", bold=True, size=14); t.alignment = center
    t.fill = PatternFill("solid", fgColor=NAVY); ks.row_dimensions[1].height = 30

    if raw is None or "Occupancy Date" not in getattr(raw, "columns", []):
        ks["A3"] = "Upload the Data Extract with Budget/Forecast columns to populate this recap."
        ks["A3"].font = Font(size=11, italic=True, color="884400")
        return ks

    labels, metrics, fc_kind = compute_monthly_recap(raw, as_of)

    ks.merge_cells("A2:G2")
    s2 = ks["A2"]
    s2.value = f"As of {pd.Timestamp(as_of):%b %d, %Y}   -   Forecast source: {fc_kind}   -   Occupancy & RevPAR use physical capacity"
    s2.font = Font(size=10, color="555555"); s2.alignment = center

    pct_metrics = {"Occupancy %"}
    money_metrics = {"ADR", "RevPAR", "Room Revenue"}

    def numfmt(metric):
        if metric in pct_metrics: return "0.0%"
        if metric in money_metrics: return "$#,##0"
        return "#,##0"

    def var_numfmt(metric):
        if metric in pct_metrics: return "+0.0%;-0.0%"
        if metric in money_metrics: return "+$#,##0;-$#,##0"
        return "+#,##0;-#,##0"

    metric_order = ["Rooms Sold", "Occupancy %", "ADR", "RevPAR", "Room Revenue"]
    heads = ["Metric", "On the Books", "Forecast", "Budget", "Last Year",
             "Fcst vs Budget", "OTB vs Budget"]
    monthcolors = [ORANGE, PURPLE, "3A6EA5"]

    row = 4
    for mi, label in enumerate(labels):
        # month banner
        ks.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        b = ks.cell(row=row, column=1, value=label)
        b.font = Font(color="FFFFFF", bold=True, size=12); b.alignment = center
        b.fill = PatternFill("solid", fgColor=monthcolors[mi % 3])
        ks.row_dimensions[row].height = 22
        row += 1
        # header
        for j, h in enumerate(heads, start=1):
            c = ks.cell(row=row, column=j, value=h)
            c.font = white; c.alignment = center; c.fill = PatternFill("solid", fgColor=TEAL)
        row += 1
        # metric rows
        for metric in metric_order:
            m = metrics[metric][mi]
            fvb = (m["fcst"] - m["budget"]) if pd.notna(m["fcst"]) and pd.notna(m["budget"]) else None
            ovb = (m["otb"] - m["budget"]) if pd.notna(m["otb"]) and pd.notna(m["budget"]) else None
            cells = [metric, m["otb"], m["fcst"], m["budget"], m["stly"], fvb, ovb]
            for j, v in enumerate(cells, start=1):
                c = ks.cell(row=row, column=j,
                            value=(None if (isinstance(v, float) and pd.isna(v)) else v))
                c.border = thin
                if j == 1:
                    c.font = Font(size=10, bold=True); c.alignment = left
                else:
                    c.alignment = center; c.font = Font(size=10)
                    if j in (6, 7):
                        c.number_format = var_numfmt(metric)
                        if isinstance(v, (int, float)) and pd.notna(v):
                            c.font = Font(size=10, bold=True, color=("1a7f37" if v >= 0 else "C00000"))
                    else:
                        c.number_format = numfmt(metric)
            row += 1
        row += 1  # spacer between months

    # ---- Group vs Transient split (rooms) ----
    d = raw.copy()
    d["Occupancy Date"] = pd.to_datetime(d["Occupancy Date"], errors="coerce")
    d = d.dropna(subset=["Occupancy Date"])
    def col(name):
        return pd.to_numeric(d[name], errors="coerce") if name in d.columns else pd.Series(0.0, index=d.index)
    myf_g = col("My Forecast Occupancy - Group This Year")
    fcg = "My Forecast Occupancy - Group This Year" if myf_g.fillna(0).abs().sum() > 0 else "Occupancy Forecast - Group This Year"
    myf_t = col("My Forecast Occupancy - Transient This Year")
    fct = "My Forecast Occupancy - Transient This Year" if myf_t.fillna(0).abs().sum() > 0 else "Occupancy Forecast - Transient This Year"
    d["_otb_g"] = col("Rooms Sold - Group This Year"); d["_otb_t"] = col("Rooms Sold - Transient This Year")
    d["_fc_g"] = col(fcg); d["_fc_t"] = col(fct)
    d["_bud_g"] = col("Budget Occupancy - Group This Year"); d["_bud_t"] = col("Budget Occupancy - Transient This Year")
    d["_ym"] = list(zip(d["Occupancy Date"].dt.year, d["Occupancy Date"].dt.month))
    base = pd.Timestamp(as_of)
    yms = [((base + pd.DateOffset(months=k)).year, (base + pd.DateOffset(months=k)).month) for k in range(3)]

    ks.cell(row=row, column=1, value="Rooms by segment - Group vs Transient (room-nights)").font = Font(bold=True, size=12)
    row += 1
    seg_heads = ["Month", "Seg", "OTB Rooms", "Fcst Rooms", "Budget Rooms", "Fcst vs Bud", "OTB vs Bud"]
    for j, h in enumerate(seg_heads, start=1):
        c = ks.cell(row=row, column=j, value=h)
        c.font = white; c.alignment = center; c.fill = PatternFill("solid", fgColor=NAVY)
    row += 1
    for i, ym in enumerate(yms):
        s = d[d["_ym"] == ym]
        for seg, otb, fc, bud, fill in [
            ("Group", s["_otb_g"].sum(), s["_fc_g"].sum(), s["_bud_g"].sum(), GREEN),
            ("Transient", s["_otb_t"].sum(), s["_fc_t"].sum(), s["_bud_t"].sum(), PURPLE)]:
            vals = [f"{calendar.month_abbr[ym[1]]} {ym[0]}", seg, otb, fc, bud, fc - bud, otb - bud]
            for j, v in enumerate(vals, start=1):
                c = ks.cell(row=row, column=j, value=v)
                c.border = thin; c.alignment = center; c.font = Font(size=10)
                if j == 1:
                    c.font = Font(size=10, bold=True, color="FFFFFF")
                    c.fill = PatternFill("solid", fgColor=monthcolors[i % 3])
                elif j == 2:
                    c.font = Font(size=10, bold=True, color="FFFFFF")
                    c.fill = PatternFill("solid", fgColor=fill)
                else:
                    c.number_format = "#,##0"
                    if j in (6, 7) and isinstance(v, (int, float)) and pd.notna(v):
                        c.font = Font(size=10, bold=True, color=("1a7f37" if v >= 0 else "C00000"))
            row += 1

    for cc, w in zip("ABCDEFG", [15, 14, 14, 14, 13, 15, 15]):
        ks.column_dimensions[cc].width = w
    ks.sheet_view.showGridLines = False
    return ks


# =========================================================================== #
#  DASHBOARD PAGE  — reproduces the workbook 'Dashboard' tab
# =========================================================================== #
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# --------------------------------------------------------------------------- #
#  Static reference tables (pulled straight from the workbook Dashboard tab).
#  Tuple = (Budget RN, Budget ADR, Freeze RN, Freeze ADR).
# --------------------------------------------------------------------------- #
_REF_TRANSIENT = {
    "Jan": (900, 329.85, 699, 304.42),
    "Feb": (1510, 343.97, 1041, 330.88),
    "Mar": (2006, 369.38, 1670, 306.26),
    "Apr": (1982, 376.04, 2007, 340.52),
    "May": (2254, 385.27, 1279, 415.08),
    "Jun": (2441, 382.19, 1532, 429.48),
    "Jul": (2182, 376.15, 1257, 369.70),
    "Aug": (1989, 361.19, 1579, 381.60),
    "Sep": (2190, 390.36, 1849, 477.54),
    "Oct": (2233, 407.01, 2233, 407.01),
    "Nov": (2003, 352.54, 2003, 352.54),
    "Dec": (1990, 344.98, 1990, 344.98),
}
_REF_GROUP = {
    "Jan": (212, 206.81, 232, 208.85),
    "Feb": (275, 249.48, 275, 249.48),
    "Mar": (600, 278.00, 580, 316.77),
    "Apr": (500, 279.00, 764, 269.00),
    "May": (650, 289.00, 650, 289.00),
    "Jun": (700, 297.57, 694, 272.25),
    "Jul": (500, 224.00, 250, 294.54),
    "Aug": (525, 239.00, 333, 381.70),
    "Sep": (650, 275.00, 950, 300.42),
    "Oct": (600, 294.45, 600, 294.45),
    "Nov": (420, 229.00, 420, 229.00),
    "Dec": (350, 243.00, 350, 243.00),
}


def default_reference_frames():
    """Return (budget_df, freeze_df) editable tables indexed like the sidebar
    data-editors expect: columns [Month, Trans RN, Trans ADR, Group RN, Group ADR]."""
    brows, frows = [], []
    for m in MONTHS:
        tb_rn, tb_adr, tf_rn, tf_adr = _REF_TRANSIENT[m]
        gb_rn, gb_adr, gf_rn, gf_adr = _REF_GROUP[m]
        brows.append({"Month": m, "Trans RN": tb_rn, "Trans ADR": tb_adr,
                      "Group RN": gb_rn, "Group ADR": gb_adr})
        frows.append({"Month": m, "Trans RN": tf_rn, "Trans ADR": tf_adr,
                      "Group RN": gf_rn, "Group ADR": gf_adr})
    return pd.DataFrame(brows), pd.DataFrame(frows)


def classify_segment(market_segment):
    """Map a detailed SynXis market-segment name to Transient / Group / Comp / Other,
    exactly matching how the workbook's SEGMENT helper column rolls things up."""
    m = str(market_segment).strip()
    if m.startswith("Transient") or m == "Contract":
        return "Transient"
    if m.startswith("Group"):
        return "Group"
    if m.startswith("Comp"):
        return "Comp"
    return "Other"


def parse_market_segment(raw):
    """
    Aggregate a Market Segment export (raw sheet read with header=None) into a
    dict keyed by (month_label, segment) -> {'rooms': x, 'rev': y}, where
    month_label is like 'Jan-2026' and segment is 'Transient' or 'Group'.
    Returns (aggregation_dict, set_of_years).
    """
    df = raw.copy()
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)

    def find(*names):
        for n in names:
            for c in df.columns:
                if str(c).strip().lower() == n.lower():
                    return c
        return None

    c_date = find("Occupancy Date")
    c_seg = find("Market Segment")
    c_rms = find("Occupancy On Books This Year", "Occupancy on Books This Year")
    c_rev = find("Booked Room Revenue This Year")
    if c_date is None or c_seg is None or c_rms is None:
        return {}, set()

    dates = pd.to_datetime(df[c_date], errors="coerce")
    rooms = pd.to_numeric(df[c_rms], errors="coerce").fillna(0)
    rev = (pd.to_numeric(df[c_rev], errors="coerce").fillna(0)
           if c_rev is not None else pd.Series(0.0, index=df.index))
    seg = df[c_seg].map(classify_segment)

    agg = {}
    years = set()
    for d, s, rm, rv in zip(dates, seg, rooms, rev):
        if pd.isna(d) or s not in ("Transient", "Group"):
            continue
        label = f"{calendar.month_abbr[d.month]}-{d.year}"
        years.add(d.year)
        key = (label, s)
        cur = agg.setdefault(key, {"rooms": 0.0, "rev": 0.0})
        cur["rooms"] += float(rm)
        cur["rev"] += float(rv)
    return agg, years


def _safe_div(a, b):
    return (a / b) if b else 0.0


def _month_days(month_abbr, year):
    m = MONTHS.index(month_abbr) + 1
    return calendar.monthrange(year, m)[1]


def budget_freeze_from_extract(data_extract, year):
    """
    Derive per-month, per-segment Budget and Freeze Forecast figures straight
    from a Data Extraction export (the 'Property' sheet, already parsed so its
    first row is the header).

      * Budget    -> 'Budget Occupancy - Group/Transient This Year' (rooms) and
                     'Budget Room Revenue This Year' (total revenue).
      * Freeze    -> the USER forecast when it is populated, otherwise the SYSTEM
                     forecast:
                        user   = 'My Forecast Occupancy - …' + 'My Forecast Revenue This Year'
                        system = 'Occupancy Forecast - …'   + 'Forecasted Room Revenue This Year'

    The export only carries revenue at the TOTAL level (no segment revenue), and
    Budget/Forecast occupancy Total == Group + Transient exactly, so segment
    revenue is split using the blended monthly ADR (total revenue / total rooms).
    Each segment therefore shares that scenario's blended ADR while Group +
    Transient revenue still reconciles to the total.

    Returns (bf, freeze_source) where:
        bf[(month_abbr, seg)] = (bud_rn, bud_adr, frz_rn, frz_adr)   seg in {Transient, Group}
        freeze_source         = 'Your forecast' or 'System forecast'
    Returns (None, None) if the required columns are not present.
    """
    if data_extract is None:
        return None, None
    d = data_extract.copy()
    if "Occupancy Date" not in d.columns:
        return None, None
    d["Occupancy Date"] = pd.to_datetime(d["Occupancy Date"], errors="coerce")
    d = d.dropna(subset=["Occupancy Date"])
    d = d[d["Occupancy Date"].dt.year == year]
    if d.empty:
        return None, None

    def col(name):
        return (pd.to_numeric(d[name], errors="coerce").fillna(0)
                if name in d.columns else None)

    bud_rev = col("Budget Room Revenue This Year")
    bud_g = col("Budget Occupancy - Group This Year")
    bud_t = col("Budget Occupancy - Transient This Year")
    if bud_rev is None or bud_g is None or bud_t is None:
        return None, None

    # Freeze source: user forecast if it carries any value, else system forecast.
    my_rev = col("My Forecast Revenue This Year")
    my_g = col("My Forecast Occupancy - Group This Year")
    my_t = col("My Forecast Occupancy - Transient This Year")
    use_user = (my_rev is not None and float(my_rev.abs().sum()) > 0
                and my_g is not None and my_t is not None
                and float((my_g.abs() + my_t.abs()).sum()) > 0)
    if use_user:
        frz_rev, frz_g, frz_t = my_rev, my_g, my_t
        freeze_source = "Your forecast"
    else:
        frz_rev = col("Forecasted Room Revenue This Year")
        frz_g = col("Occupancy Forecast - Group This Year")
        frz_t = col("Occupancy Forecast - Transient This Year")
        freeze_source = "System forecast"
        if frz_rev is None or frz_g is None or frz_t is None:
            frz_rev = bud_rev * 0.0
            frz_g = bud_g * 0.0
            frz_t = bud_t * 0.0

    work = pd.DataFrame({
        "_mo": d["Occupancy Date"].dt.month.values,
        "bud_rev": bud_rev.values, "bud_g": bud_g.values, "bud_t": bud_t.values,
        "frz_rev": frz_rev.values, "frz_g": frz_g.values, "frz_t": frz_t.values,
    })
    grp = work.groupby("_mo").sum()

    bf = {}
    for mo, row in grp.iterrows():
        mabbr = MONTHS[int(mo) - 1]
        bud_rn_tot = row["bud_g"] + row["bud_t"]
        bud_adr = _safe_div(row["bud_rev"], bud_rn_tot)   # blended budget ADR
        frz_rn_tot = row["frz_g"] + row["frz_t"]
        frz_adr = _safe_div(row["frz_rev"], frz_rn_tot)   # blended freeze ADR
        bf[(mabbr, "Transient")] = (float(row["bud_t"]), bud_adr,
                                    float(row["frz_t"]), frz_adr)
        bf[(mabbr, "Group")] = (float(row["bud_g"]), bud_adr,
                                float(row["frz_g"]), frz_adr)
    return bf, freeze_source


def compute_dashboard(cur_agg, last_agg, capacity, year,
                      data_extract=None, budget_df=None, freeze_df=None,
                      week_agg=None):
    """
    Build the three dashboard tables. Returns a dict:
        {'Total': rows, 'Transient': rows, 'Group': rows, '_freeze_source': str}
    where rows is a list of 13 dicts (12 months + Total), each holding every
    metric cell needed to render the table.

    Budget and Freeze Forecast are pulled from `data_extract` (the Data
    Extraction export) when it carries the Budget/Forecast columns; otherwise
    they fall back to the static `budget_df` / `freeze_df` reference tables.

    When `week_agg` (the monthly/segment aggregation from ~7 days ago) is given,
    each row also carries a 7-days-ago baseline (wk_rn/wk_adr/wk_rev) and the
    7-day pickup (wpu_rn/wpu_adr/wpu_rev = current minus 7-days-ago).
    """
    has_week = bool(week_agg)
    bf, freeze_source = budget_freeze_from_extract(data_extract, year)
    use_extract = bf is not None

    if budget_df is None or freeze_df is None:
        budget_df, freeze_df = default_reference_frames()
    bud = {r["Month"]: r for _, r in budget_df.iterrows()}
    frz = {r["Month"]: r for _, r in freeze_df.iterrows()}

    def ref(month, seg, table, kind):
        if use_extract:
            b_rn, b_adr, f_rn, f_adr = bf.get((month, seg), (0.0, 0.0, 0.0, 0.0))
            return (b_rn, b_adr) if table == "budget" else (f_rn, f_adr)
        row = (bud if table == "budget" else frz)[month]
        if seg == "Transient":
            rn = row["Trans RN"]; adr = row["Trans ADR"]
        else:
            rn = row["Group RN"]; adr = row["Group ADR"]
        return (float(rn or 0), float(adr or 0))

    def seg_month(agg, month, seg):
        key = (f"{month}-{year}", seg)
        d = agg.get(key, {"rooms": 0.0, "rev": 0.0})
        return float(d["rooms"]), float(d["rev"])

    def week_month(m, seg):
        """7-days-ago rooms & rev for a month/segment (Total = Trans + Group)."""
        if not has_week:
            return 0.0, 0.0
        if seg in ("Transient", "Group"):
            return seg_month(week_agg, m, seg)
        t = seg_month(week_agg, m, "Transient"); g = seg_month(week_agg, m, "Group")
        return t[0] + g[0], t[1] + g[1]

    def build_rows(seg):
        rows = []
        tot = {k: 0.0 for k in ("otb_rn", "otb_rev", "lr_rn", "lr_rev",
                                "bud_rn", "bud_rev", "frz_rn", "frz_rev",
                                "wk_rn", "wk_rev")}
        for m in MONTHS:
            if seg in ("Transient", "Group"):
                otb_rn, otb_rev = seg_month(cur_agg, m, seg)
                lr_rn, lr_rev = seg_month(last_agg, m, seg)
                b_rn, b_adr = ref(m, seg, "budget", None)
                f_rn, f_adr = ref(m, seg, "freeze", None)
            else:  # Total = Transient + Group
                t = seg_month(cur_agg, m, "Transient"); g = seg_month(cur_agg, m, "Group")
                otb_rn, otb_rev = t[0] + g[0], t[1] + g[1]
                tl = seg_month(last_agg, m, "Transient"); gl = seg_month(last_agg, m, "Group")
                lr_rn, lr_rev = tl[0] + gl[0], tl[1] + gl[1]
                tb = ref(m, "Transient", "budget", None); gb = ref(m, "Group", "budget", None)
                b_rn = tb[0] + gb[0]
                b_rev_tmp = tb[0] * tb[1] + gb[0] * gb[1]
                b_adr = _safe_div(b_rev_tmp, b_rn)
                tf = ref(m, "Transient", "freeze", None); gf = ref(m, "Group", "freeze", None)
                f_rn = tf[0] + gf[0]
                f_rev_tmp = tf[0] * tf[1] + gf[0] * gf[1]
                f_adr = _safe_div(f_rev_tmp, f_rn)

            bud_rev = b_rn * b_adr
            frz_rev = f_rn * f_adr
            otb_adr = _safe_div(otb_rev, otb_rn)
            lr_adr = _safe_div(lr_rev, lr_rn)
            days = _month_days(m, year)
            occ = _safe_div(otb_rn, capacity * days)

            row = {
                "month": f"{m}-{year}",
                "occ": occ,
                "otb_rn": otb_rn, "otb_adr": otb_adr, "otb_rev": otb_rev,
                "lr_rn": lr_rn, "lr_adr": lr_adr, "lr_rev": lr_rev,
                "pu_rn": otb_rn - lr_rn, "pu_adr": otb_adr - lr_adr, "pu_rev": otb_rev - lr_rev,
                "bud_rn": b_rn, "bud_adr": b_adr, "bud_rev": bud_rev,
                "bv_rn": otb_rn - b_rn, "bv_adr": otb_adr - b_adr, "bv_rev": otb_rev - bud_rev,
                "frz_rn": f_rn, "frz_adr": f_adr, "frz_rev": frz_rev,
                "fv_rn": otb_rn - f_rn, "fv_adr": otb_adr - f_adr, "fv_rev": otb_rev - frz_rev,
            }
            if has_week:
                wk_rn, wk_rev = week_month(m, seg)
                wk_adr = _safe_div(wk_rev, wk_rn)
                row.update({
                    "wk_rn": wk_rn, "wk_adr": wk_adr, "wk_rev": wk_rev,
                    "wpu_rn": otb_rn - wk_rn, "wpu_adr": otb_adr - wk_adr,
                    "wpu_rev": otb_rev - wk_rev,
                })
                tot["wk_rn"] += wk_rn; tot["wk_rev"] += wk_rev
            rows.append(row)
            tot["otb_rn"] += otb_rn; tot["otb_rev"] += otb_rev
            tot["lr_rn"] += lr_rn; tot["lr_rev"] += lr_rev
            tot["bud_rn"] += b_rn; tot["bud_rev"] += bud_rev
            tot["frz_rn"] += f_rn; tot["frz_rev"] += frz_rev

        # ---- Total row (bottom) ----
        year_days = 366 if calendar.isleap(year) else 365
        otb_adr = _safe_div(tot["otb_rev"], tot["otb_rn"])
        lr_adr = _safe_div(tot["lr_rev"], tot["lr_rn"])
        bud_adr = _safe_div(tot["bud_rev"], tot["bud_rn"])
        frz_adr = _safe_div(tot["frz_rev"], tot["frz_rn"])
        trow = {
            "month": "Total",
            "occ": _safe_div(tot["otb_rn"], capacity * year_days),
            "otb_rn": tot["otb_rn"], "otb_adr": otb_adr, "otb_rev": tot["otb_rev"],
            "lr_rn": tot["lr_rn"], "lr_adr": lr_adr, "lr_rev": tot["lr_rev"],
            "pu_rn": tot["otb_rn"] - tot["lr_rn"], "pu_adr": otb_adr - lr_adr,
            "pu_rev": tot["otb_rev"] - tot["lr_rev"],
            "bud_rn": tot["bud_rn"], "bud_adr": bud_adr, "bud_rev": tot["bud_rev"],
            "bv_rn": tot["otb_rn"] - tot["bud_rn"], "bv_adr": otb_adr - bud_adr,
            "bv_rev": tot["otb_rev"] - tot["bud_rev"],
            "frz_rn": tot["frz_rn"], "frz_adr": frz_adr, "frz_rev": tot["frz_rev"],
            "fv_rn": tot["otb_rn"] - tot["frz_rn"], "fv_adr": otb_adr - frz_adr,
            "fv_rev": tot["otb_rev"] - tot["frz_rev"],
        }
        if has_week:
            wk_adr_t = _safe_div(tot["wk_rev"], tot["wk_rn"])
            trow.update({
                "wk_rn": tot["wk_rn"], "wk_adr": wk_adr_t, "wk_rev": tot["wk_rev"],
                "wpu_rn": tot["otb_rn"] - tot["wk_rn"], "wpu_adr": otb_adr - wk_adr_t,
                "wpu_rev": tot["otb_rev"] - tot["wk_rev"],
            })
        rows.append(trow)
        return rows

    return {"Total": build_rows("Total"),
            "Transient": build_rows("Transient"),
            "Group": build_rows("Group"),
            "_has_week": has_week,
            "_freeze_source": (freeze_source if use_extract else "Static table"),
            "_budget_source": ("Data Extraction" if use_extract else "Static table")}


# --------------------------------------------------------------------------- #
#  Formatting helpers
# --------------------------------------------------------------------------- #
def _f_pct(v):
    return f"{v*100:,.1f}%"

def _f_int(v):
    return f"{v:,.0f}"

def _f_money0(v):
    return f"${v:,.0f}"

def _f_money2(v):
    return f"${v:,.2f}"

def _f_delta_int(v):
    v = round(v)
    return f"({abs(v):,.0f})" if v < 0 else f"{v:,.0f}"

def _f_delta_money0(v):
    return f"(${abs(v):,.0f})" if v < 0 else f"${v:,.0f}"

def _f_delta_money2(v):
    return f"(${abs(v):,.2f})" if v < 0 else f"${v:,.2f}"


# Excel number formats reused by the layout builder
_NF_INT = "#,##0"
_NF_M2 = '"$"#,##0.00'
_NF_M0 = '"$"#,##0'
_NF_DINT = "#,##0_);[Red](#,##0)"
_NF_D2 = '"$"#,##0.00_);[Red]("$"#,##0.00)'
_NF_D0 = '"$"#,##0_);[Red]("$"#,##0)'


def _dashboard_layout(title, freeze_label="Freeze Forecast",
                      otbvar_label="OTB Var to Freeze Fcst",
                      show_lr=True, show_week=False):
    """Single source of truth for the dashboard's columns. Returns an ordered
    list of column dicts: {key, fmt, delta, group, sub, numfmt, budget}. The
    'Pickup Since Last Report' block is optional (show_lr) and a '7-Day Pickup'
    block (with its own 7-days-ago baseline) is optional (show_week), so the
    same renderer drives every pickup view."""
    cols = []

    def add(key, fmt, delta, group, sub, nf, budget=False):
        cols.append({"key": key, "fmt": fmt, "delta": delta, "group": group,
                     "sub": sub, "numfmt": nf, "budget": budget})

    add("occ", _f_pct, False, title, "Occ% OTB", "0.0%")
    add("otb_rn", _f_int, False, title, "RN OTB", _NF_INT)
    add("otb_adr", _f_money2, False, title, "ADR OTB", _NF_M2)
    add("otb_rev", _f_money0, False, title, "Revenue OTB", _NF_M0)
    if show_lr:
        add("lr_rn", _f_int, False, "Last Report OTB", "RN", _NF_INT)
        add("lr_adr", _f_money2, False, "Last Report OTB", "ADR", _NF_M2)
        add("lr_rev", _f_money0, False, "Last Report OTB", "Revenue", _NF_M0)
        add("pu_rn", _f_delta_int, True, "Pickup Since Last Report", "RN", _NF_DINT)
        add("pu_adr", _f_delta_money2, True, "Pickup Since Last Report", "ADR Chg", _NF_D2)
        add("pu_rev", _f_delta_money0, True, "Pickup Since Last Report", "Revenue", _NF_D0)
    if show_week:
        add("wk_rn", _f_int, False, "7 Days Ago OTB", "RN", _NF_INT)
        add("wk_adr", _f_money2, False, "7 Days Ago OTB", "ADR", _NF_M2)
        add("wk_rev", _f_money0, False, "7 Days Ago OTB", "Revenue", _NF_M0)
        add("wpu_rn", _f_delta_int, True, "7-Day Pickup", "RN", _NF_DINT)
        add("wpu_adr", _f_delta_money2, True, "7-Day Pickup", "ADR Chg", _NF_D2)
        add("wpu_rev", _f_delta_money0, True, "7-Day Pickup", "Revenue", _NF_D0)
    add("bud_rn", _f_int, False, "Budget", "RN", _NF_INT, True)
    add("bud_adr", _f_money2, False, "Budget", "ADR", _NF_M2, True)
    add("bud_rev", _f_money0, False, "Budget", "Rev", _NF_M0, True)
    add("bv_rn", _f_delta_int, True, "Budget Variance", "RN", _NF_DINT)
    add("bv_adr", _f_delta_money2, True, "Budget Variance", "ADR", _NF_D2)
    add("bv_rev", _f_delta_money0, True, "Budget Variance", "Rev", _NF_D0)
    add("frz_rn", _f_int, False, freeze_label, "RN", _NF_INT)
    add("frz_adr", _f_money2, False, freeze_label, "ADR", _NF_M2)
    add("frz_rev", _f_money0, False, freeze_label, "Rev", _NF_M0)
    add("fv_rn", _f_delta_int, True, otbvar_label, "RN", _NF_DINT)
    add("fv_adr", _f_delta_money2, True, otbvar_label, "ADR", _NF_D2)
    add("fv_rev", _f_delta_money0, True, otbvar_label, "Rev", _NF_D0)
    return cols


def _groups_from_cols(cols):
    """Collapse a column list into (group_name, span) banner segments."""
    groups = []
    for c in cols:
        if groups and groups[-1][0] == c["group"]:
            groups[-1] = (c["group"], groups[-1][1] + 1)
        else:
            groups.append((c["group"], 1))
    return groups


def dashboard_css(colors=None):
    c = {
        "band":   "#F4CCCC",   # section + group header band (salmon)
        "sub":    "#FBE4E4",   # sub-header row
        "budget": "#EDE7D3",   # budget/freeze group shading (tan)
        "total":  "#F2F2F2",   # total row
        "neg":    "#C00000",   # negative text
        "grid":   "#D9D9D9",
        "text":   "#1a1a1a",
    }
    if colors:
        c.update(colors)
    return f"""
<style>
.cmp-dash {{ font-family: Segoe UI, Arial, sans-serif; margin-bottom: 26px; }}
.cmp-dash table {{ border-collapse: collapse; width: 100%; font-size: 12px; }}
.cmp-dash th, .cmp-dash td {{ border: 1px solid {c['grid']}; padding: 3px 6px;
    text-align: right; white-space: nowrap; color: {c['text']}; }}
.cmp-dash .title {{ background: {c['band']}; font-weight: 700; text-align: center;
    font-size: 13px; padding: 5px; }}
.cmp-dash .band {{ background: {c['band']}; font-weight: 700; text-align: center; }}
.cmp-dash .sub  {{ background: {c['sub']};  font-weight: 700; text-align: center; }}
.cmp-dash td.month, .cmp-dash th.month {{ text-align: left; font-weight: 600; }}
.cmp-dash tr.total td {{ background: {c['total']}; font-weight: 700;
    border-top: 2px solid #999; }}
.cmp-dash td.budget {{ background: {c['budget']}; }}
.cmp-dash tr.total td.budget {{ background: #E4DEC8; }}
.cmp-dash .neg {{ color: {c['neg']}; }}
.cmp-dash .hdrline {{ font-size: 13px; font-weight: 700; margin: 4px 0 8px 0; }}
</style>
"""


def _freeze_labels(tables):
    """Map the freeze-forecast source to the two group headers that reference it,
    so the sheet says 'My Forecast' vs 'System Forecast' instead of a generic
    'Freeze Forecast'."""
    fs = (tables or {}).get("_freeze_source", "")
    if fs == "Your forecast":
        return "My Forecast", "OTB Var to My Fcst"
    if fs == "System forecast":
        return "System Forecast", "OTB Var to System Fcst"
    return "Freeze Forecast", "OTB Var to Freeze Fcst"


def _table_html(title, rows, cols):
    groups = _groups_from_cols(cols)

    h = ['<div class="cmp-dash"><table>']
    # banner row
    h.append('<tr><th class="title" rowspan="2">Month</th>')
    for name, span in groups:
        cls = "title" if name == title else "band"
        h.append(f'<th class="{cls}" colspan="{span}">{name}</th>')
    h.append('</tr>')
    # sub-header row
    h.append('<tr>')
    for c in cols:
        cls = "sub budget" if c["budget"] else "sub"
        h.append(f'<th class="{cls}">{c["sub"]}</th>')
    h.append('</tr>')
    # data rows
    for r in rows:
        is_total = r["month"] == "Total"
        h.append('<tr class="total">' if is_total else '<tr>')
        h.append(f'<td class="month">{r["month"]}</td>')
        for c in cols:
            v = r.get(c["key"])
            txt = c["fmt"](v) if isinstance(v, (int, float)) else ""
            classes = []
            if c["budget"]:
                classes.append("budget")
            if c["delta"] and isinstance(v, (int, float)) and v < 0:
                classes.append("neg")
            cls = f' class="{" ".join(classes)}"' if classes else ""
            h.append(f"<td{cls}>{txt}</td>")
        h.append('</tr>')
    h.append('</table></div>')
    return "".join(h)


def render_dashboard_html(tables, header_line=None, colors=None,
                          show_lr=True, show_week=None):
    parts = [dashboard_css(colors)]
    if header_line:
        parts.append(f'<div class="cmp-dash"><div class="hdrline">{header_line}</div></div>')
    freeze_label, otbvar_label = _freeze_labels(tables)
    has_week = bool(tables.get("_has_week"))
    show_week = has_week if show_week is None else (show_week and has_week)
    if not show_lr and not show_week:
        show_lr = True                     # never render zero pickup blocks
    for seg, title in [("Total", "Total OTB"),
                       ("Transient", "Transient OTB"),
                       ("Group", "Group OTB")]:
        cols = _dashboard_layout(title, freeze_label, otbvar_label,
                                 show_lr=show_lr, show_week=show_week)
        parts.append(_table_html(title, tables[seg], cols))
    return "\n".join(parts)


# --------------------------------------------------------------------------- #
#  Excel export (mirrors the on-screen layout)
# --------------------------------------------------------------------------- #
def dashboard_to_excel(tables, header_line=None, show_lr=True, show_week=None):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    BAND = "F4CCCC"; SUB = "FBE4E4"; BUDGET = "EDE7D3"; TOTAL = "F2F2F2"
    band_fill = PatternFill("solid", fgColor=BAND)
    sub_fill = PatternFill("solid", fgColor=SUB)
    bud_fill = PatternFill("solid", fgColor=BUDGET)
    tot_fill = PatternFill("solid", fgColor=TOTAL)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    thin = Border(*(Side(style="thin", color="D9D9D9"),) * 4)
    bold = Font(bold=True, size=9)
    reg = Font(size=9)

    freeze_label, otbvar_label = _freeze_labels(tables)
    has_week = bool(tables.get("_has_week"))
    show_week = has_week if show_week is None else (show_week and has_week)
    if not show_lr and not show_week:
        show_lr = True

    ncols = 0
    row = 1
    if header_line:
        ws.cell(row=row, column=1, value=header_line).font = Font(bold=True, size=12)
        row += 2

    for seg, title in [("Total", "Total OTB"), ("Transient", "Transient OTB"),
                       ("Group", "Group OTB")]:
        rows = tables[seg]
        cols = _dashboard_layout(title, freeze_label, otbvar_label,
                                 show_lr=show_lr, show_week=show_week)
        groups = _groups_from_cols(cols)
        ncols = max(ncols, len(cols) + 1)
        # banner row
        ws.cell(row=row, column=1, value="Month")
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=1)
        col = 2
        for name, span in groups:
            ws.cell(row=row, column=col, value=name)
            if span > 1:
                ws.merge_cells(start_row=row, start_column=col,
                               end_row=row, end_column=col + span - 1)
            col += span
        for cc in range(1, col):
            cell = ws.cell(row=row, column=cc)
            cell.fill = band_fill; cell.font = bold; cell.alignment = center; cell.border = thin
        # sub-header row
        sr = row + 1
        for i, c in enumerate(cols):
            cell = ws.cell(row=sr, column=2 + i, value=c["sub"])
            cell.fill = bud_fill if c["budget"] else sub_fill
            cell.font = bold; cell.alignment = center; cell.border = thin
        ws.cell(row=sr, column=1).border = thin
        # data
        dr = sr + 1
        for r in rows:
            is_total = r["month"] == "Total"
            mc = ws.cell(row=dr, column=1, value=r["month"])
            mc.alignment = left; mc.border = thin; mc.font = bold if is_total else reg
            if is_total:
                mc.fill = tot_fill
            for i, c in enumerate(cols):
                v = r.get(c["key"])
                cell = ws.cell(row=dr, column=2 + i,
                               value=(v if isinstance(v, (int, float)) else None))
                cell.number_format = c["numfmt"]
                cell.border = thin; cell.alignment = center
                cell.font = bold if is_total else reg
                if c["budget"]:
                    cell.fill = bud_fill
                elif is_total:
                    cell.fill = tot_fill
            dr += 1
        row = dr + 1

    ws.column_dimensions["A"].width = 10
    for j in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(j)].width = 11
    ws.sheet_view.showGridLines = False

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf



# =========================================================================== #
#  OTB SNAPSHOT HISTORY  — record each upload's on-the-books so the next upload
#  can show pickup automatically (no need to hand-upload the prior file).
#
#  Storage strategy (works on GitHub + Streamlit Community Cloud, no Azure/DB):
#    * primary : st.session_state (survives reruns within a session)
#    * mirror  : a small JSON file next to app.py (survives while the container
#                is warm; Streamlit Cloud disk is EPHEMERAL, so it can reset on a
#                cold start / redeploy)
#    * durable : a Download / Restore pair in the sidebar lets you keep the
#                history JSON in OneDrive (or commit it to the repo) and reload
#                it, so nothing is ever truly lost.
# =========================================================================== #
_HISTORY_FILE = "compton_otb_history.json"


def _history_path():
    try:
        base = os.path.dirname(os.path.abspath(__file__))
    except NameError:
        base = os.getcwd()
    return os.path.join(base, _HISTORY_FILE)


def _agg_to_json(agg):
    """{(label, seg): {'rooms':x,'rev':y}} -> {'label|seg': [rooms, rev]}"""
    return {f"{k[0]}|{k[1]}": [round(float(v["rooms"]), 4), round(float(v["rev"]), 4)]
            for k, v in agg.items()}


def _json_to_agg(d):
    out = {}
    for k, v in (d or {}).items():
        if "|" not in k:
            continue
        label, seg = k.split("|", 1)
        try:
            out[(label, seg)] = {"rooms": float(v[0]), "rev": float(v[1])}
        except (TypeError, ValueError, IndexError):
            continue
    return out


def load_history():
    """Return the history dict {'snapshots': {'YYYY-MM-DD': {...}}}, preferring
    the in-session copy, then the on-disk mirror."""
    hist = st.session_state.get("_otb_history")
    if hist is None:
        try:
            with open(_history_path(), "r") as f:
                hist = json.load(f)
        except Exception:
            hist = {"snapshots": {}}
        st.session_state["_otb_history"] = hist
    hist.setdefault("snapshots", {})
    return hist


def save_history(hist):
    st.session_state["_otb_history"] = hist
    try:
        with open(_history_path(), "w") as f:
            json.dump(hist, f, indent=0)
    except Exception:
        pass  # read-only / ephemeral fs — session_state still holds it


def date_otb_from_extract(data):
    """Per-occupancy-date on-the-books from a Data Extract (the DD's source),
    for date-level pickup. Returns { 'YYYY-MM-DD': [total_rms, rev, trans, group] }."""
    if data is None or "Occupancy Date" not in getattr(data, "columns", []):
        return {}
    d = data.copy()
    d["Occupancy Date"] = pd.to_datetime(d["Occupancy Date"], errors="coerce")
    d = d.dropna(subset=["Occupancy Date"])

    def col(n):
        return (pd.to_numeric(d[n], errors="coerce").fillna(0)
                if n in d.columns else pd.Series(0.0, index=d.index))

    tot = col("Occupancy On Books This Year")
    rev = col("Booked Room Revenue This Year")
    tr = col("Rooms Sold - Transient This Year")
    gr = col("Rooms Sold - Group This Year")
    out = {}
    for dt, t, rv, x, g in zip(d["Occupancy Date"], tot, rev, tr, gr):
        out[dt.strftime("%Y-%m-%d")] = [float(t), float(rv), float(x), float(g)]
    return out


def record_snapshot(hist, report_date, capacity, year, source_name="",
                    monthly_agg=None, by_date=None):
    """Record (or MERGE into) the snapshot keyed by report_date. Each caller can
    contribute just the sections it has: `monthly_agg` (dashboard, by month/seg)
    and/or `by_date` (DD, by occupancy date). Existing sections are preserved
    when a caller doesn't supply them."""
    key = pd.Timestamp(report_date).strftime("%Y-%m-%d")
    snaps = hist.setdefault("snapshots", {})
    snap = snaps.get(key, {})
    snap["generated"] = pd.Timestamp.now().isoformat(timespec="seconds")
    snap["capacity"] = int(capacity)
    snap["year"] = int(year)
    if source_name:
        snap["source"] = source_name
    if monthly_agg is not None:
        snap["data"] = _agg_to_json(monthly_agg)
    if by_date is not None:
        snap["by_date"] = by_date
    snaps[key] = snap
    return hist


def snapshot_dates(hist, require=None):
    """Sorted list of snapshot dates (as date objects). `require` optionally
    filters to snapshots that carry a given section ('data' or 'by_date')."""
    out = []
    for k, snap in hist.get("snapshots", {}).items():
        if require and not snap.get(require):
            continue
        try:
            out.append(pd.Timestamp(k).normalize().date())
        except Exception:
            continue
    return sorted(out)


def get_snapshot_monthly(hist, date_key):
    snap = hist.get("snapshots", {}).get(pd.Timestamp(date_key).strftime("%Y-%m-%d"), {})
    return _json_to_agg(snap.get("data", {}))


def get_snapshot_by_date(hist, date_key):
    snap = hist.get("snapshots", {}).get(pd.Timestamp(date_key).strftime("%Y-%m-%d"), {})
    return snap.get("by_date", {}) or {}


def previous_snapshot(hist, report_date, require=None):
    """Most recent snapshot strictly BEFORE report_date. Returns (date, agg) or
    (None, None). `require` optionally restricts to snapshots with that section."""
    rd = pd.Timestamp(report_date).normalize().date()
    cands = [d for d in snapshot_dates(hist, require=require) if d < rd]
    if not cands:
        return None, None
    best = cands[-1]
    return best, get_snapshot_monthly(hist, best)


def nearest_snapshot(hist, target_date, before=None, require=None):
    """Snapshot date closest to target_date. If `before` is given, only snapshots
    strictly before it are considered. Returns a date or None."""
    tgt = pd.Timestamp(target_date).normalize().date()
    cands = snapshot_dates(hist, require=require)
    if before is not None:
        bd = pd.Timestamp(before).normalize().date()
        cands = [d for d in cands if d < bd]
    if not cands:
        return None
    return min(cands, key=lambda d: abs((pd.Timestamp(d) - pd.Timestamp(tgt)).days))


def _mseg_agg_from_upload(upload):
    """Parse an uploaded Market Segment .xlsx into the monthly/segment agg used
    by the dashboard. Returns (agg, years) or (None, None) on failure."""
    if upload is None:
        return None, None
    try:
        sheets = _read_all_sheets(upload)
        ms = [s for s in sheets if "market segment" in s.lower()]
        raw = sheets[ms[0]] if ms else sheets[list(sheets)[0]]
        return parse_market_segment(raw)
    except Exception as e:
        st.warning(f"Couldn't read **{getattr(upload, 'name', 'that file')}** ({e}).")
        return None, None


def render_dashboard_page(parts, as_of):
    """Streamlit page that reproduces the workbook 'Dashboard' tab."""
    st.title("The Compton — Owner Dashboard")
    st.caption("Total / Transient / Group on-the-books by month, versus last "
               "report, budget and the freeze forecast — the same view as the "
               "Dashboard tab of the Daily Report workbook.")

    cur_raw = parts.get("mseg_raw")
    if cur_raw is None:
        st.info("Upload your **Market Segment** export in the sidebar to build "
                "the dashboard (it supplies rooms and revenue by segment).")
        st.stop()

    cur_agg, years = parse_market_segment(cur_raw)
    if not cur_agg:
        st.error("Couldn't read rooms/revenue by segment from the Market Segment "
                 "file. Expected columns: 'Occupancy Date', 'Market Segment', "
                 "'Occupancy On Books This Year', 'Booked Room Revenue This Year'.")
        st.stop()

    # ---- capacity (physical rooms) ----
    cap_default = 142
    data = parts.get("data")
    if data is not None and "Physical Capacity This Year" in getattr(data, "columns", []):
        pc = pd.to_numeric(data["Physical Capacity This Year"], errors="coerce").dropna()
        if len(pc):
            cap_default = int(pc.max())

    # ---- pick the reporting year ----
    year = as_of.year if as_of.year in years else (min(years) if years else as_of.year)

    with st.sidebar:
        st.markdown("---")
        st.header("Dashboard settings")
        capacity = st.number_input("Physical room capacity", min_value=1,
                                   value=int(cap_default), step=1,
                                   help="Rooms available per night (drives Occ%).")
        if len(years) > 1:
            year = st.selectbox("Report year", sorted(years),
                                index=sorted(years).index(year))

    # ------------------------------------------------------------------ #
    #  Pickup source: automatic snapshot history (default) with a manual
    #  override + Download/Restore for durability on ephemeral hosts.
    # ------------------------------------------------------------------ #
    hist = load_history()
    report_date = pd.Timestamp(as_of).date()          # snapshot key = report as-of

    with st.sidebar:
        st.subheader("Pickup history")
        auto_save = st.checkbox(
            "Save this upload to history", value=True,
            help="Records tonight's on-the-books (by segment/month) under the "
                 "report as-of date so your NEXT upload can show pickup "
                 "automatically. Set the as-of date in Settings above.")
        snaps = sorted(hist.get("snapshots", {}).keys())
        if snaps:
            st.caption(f"{len(snaps)} snapshot(s) on file: "
                       f"{snaps[0]} → {snaps[-1]}")
        else:
            st.caption("No snapshots yet — this upload will seed the history.")

        with st.expander("Manage history (backup / restore)", expanded=False):
            st.caption("Streamlit Cloud storage resets on redeploy, so keep a "
                       "copy in OneDrive: download it here and re-upload to "
                       "restore. You can also commit it to your repo.")
            st.download_button(
                "⬇️ Download history (.json)",
                data=json.dumps(hist, indent=0).encode("utf-8"),
                file_name=_HISTORY_FILE, mime="application/json",
                use_container_width=True)
            restore = st.file_uploader("Restore history (.json)", type=["json"],
                                       accept_multiple_files=False, key="hist_restore")
            if restore is not None:
                try:
                    incoming = json.load(restore)
                    merged = hist.get("snapshots", {})
                    merged.update(incoming.get("snapshots", {}))
                    hist["snapshots"] = merged
                    save_history(hist)
                    st.success(f"Restored — {len(merged)} snapshot(s) now on file.")
                except Exception as e:
                    st.warning(f"Couldn't read that history file ({e}).")
            if st.button("🗑️ Clear all history", use_container_width=True):
                hist = {"snapshots": {}}
                save_history(hist)
                st.info("History cleared.")

    # ------------------------------------------------------------------ #
    #  Comparison files — SEPARATE upload areas so you can drop in whatever
    #  prior export you have on your laptop. Each falls back to saved history
    #  when left empty.
    #     * Last report  -> "Pickup Since Last Report"
    #     * 7 days ago   -> "7-Day Pickup" (Option A: full RN / ADR / Revenue)
    # ------------------------------------------------------------------ #
    with st.sidebar:
        st.subheader("Comparison files (pickup)")
        st.caption("Drop in a prior day's **Market Segment** export to drive "
                   "pickup. Leave a slot empty to fall back to saved history.")

        st.markdown("**① Last report**")
        lr_file = st.file_uploader("Last report — Market Segment (.xlsx)",
                                   type=["xlsx"], accept_multiple_files=False,
                                   key="lr_override")
        lr_date_manual = st.date_input("Last report date (optional)", value=None,
                                       key="lr_date",
                                       help="Only needed to label the header; "
                                            "otherwise inferred.")

        st.markdown("**② 7 days ago** (for the 7-day recap)")
        wk_file = st.file_uploader("7 days ago — Market Segment (.xlsx)",
                                   type=["xlsx"], accept_multiple_files=False,
                                   key="wk_override")
        wk_date_manual = st.date_input("7-days-ago date (optional)", value=None,
                                       key="wk_date",
                                       help="Defaults to 7 calendar days before "
                                            "the as-of date.")

        pickup_view = st.radio(
            "Pickup columns to show",
            ["Since last report", "7-day", "Both"], index=0, horizontal=True,
            help="Filter the pickup blocks on the dashboard. '7-day' options "
                 "require a 7-days-ago file (or a saved snapshot ~7 days back).")

    # ---- resolve the 'last report' (pickup since last report) ----
    last_agg, last_report_date, pickup_src = None, None, None
    if lr_file is not None:
        last_agg, _ = _mseg_agg_from_upload(lr_file)
        if last_agg:
            last_report_date = lr_date_manual or (pd.Timestamp(report_date) - pd.Timedelta(days=1)).date()
            pickup_src = "uploaded file"
    if last_agg is None:                       # fall back to saved history
        prev_date, prev_agg = previous_snapshot(hist, report_date)
        if prev_agg:
            last_agg, last_report_date, pickup_src = prev_agg, prev_date, "saved history"
    if last_agg is None:                       # nothing yet -> pickup 0
        last_agg = cur_agg
        last_report_date = report_date
        pickup_src = "none yet"

    # ---- resolve the '7 days ago' baseline (7-day pickup) ----
    week_agg, week_date, week_src = None, None, None
    if wk_file is not None:
        week_agg, _ = _mseg_agg_from_upload(wk_file)
        if week_agg:
            week_date = wk_date_manual or (pd.Timestamp(report_date) - pd.Timedelta(days=7)).date()
            week_src = "uploaded file"
    if week_agg is None:                       # fall back to a snapshot ~7 days back
        target = (pd.Timestamp(report_date) - pd.Timedelta(days=7)).date()
        wd = nearest_snapshot(hist, target, before=report_date, require="data")
        if wd is not None:
            week_agg = get_snapshot_monthly(hist, wd)
            week_date, week_src = wd, "saved history"

    # ---- Budget & Freeze: pull from the Data Extraction when it carries the
    #      Budget / Forecast columns, else fall back to the manual reference
    #      tables. Budget    = 'Budget …' columns.
    #                Freeze  = 'My Forecast …' (user) if present, else the
    #                          'Occupancy Forecast …' / 'Forecasted …' (system).
    bf_probe, _ = budget_freeze_from_extract(data, year)
    have_extract = bf_probe is not None

    budget_df, freeze_df = default_reference_frames()
    if not have_extract:
        with st.sidebar.expander("Budget & Freeze Forecast (manual)", expanded=False):
            st.caption("No Budget/Forecast columns were found in the Data "
                       "Extraction, so these fall back to editable reference "
                       "numbers. Re-export the Data Extraction with the Budget "
                       "and Forecast columns to pull them automatically.")
            st.markdown("**Budget**")
            budget_df = st.data_editor(budget_df, hide_index=True, key="dash_budget",
                                       use_container_width=True)
            st.markdown("**Freeze Forecast**")
            freeze_df = st.data_editor(freeze_df, hide_index=True, key="dash_freeze",
                                       use_container_width=True)

    tables = compute_dashboard(cur_agg, last_agg, capacity, year,
                               data_extract=data, budget_df=budget_df,
                               freeze_df=freeze_df, week_agg=week_agg)

    # ---- translate the pickup-view choice into render flags ----
    has_week = bool(tables.get("_has_week"))
    if pickup_view == "Since last report":
        show_lr, show_week = True, False
    elif pickup_view == "7-day":
        show_lr, show_week = False, True
    else:  # Both
        show_lr, show_week = True, True
    if (show_week and not show_lr) and not has_week:
        show_lr, show_week = True, False       # asked for 7-day but none available

    # ---- source banner so it's clear where Budget / Freeze came from ----
    if have_extract:
        fs = tables.get("_freeze_source", "System forecast")
        st.success(
            f"**Budget** pulled from the Data Extraction (Budget columns). "
            f"**Freeze Forecast** = {fs} "
            f"({'your My-Forecast values' if fs == 'Your forecast' else 'system forecast — no My-Forecast values found in the file'}). "
            "Note: the export only carries revenue at the total level, so within "
            "Budget/Freeze each segment shares that month's blended ADR while "
            "Transient + Group revenue still reconciles to the total.")
    else:
        st.warning("Budget & Freeze are using the manual reference tables "
                   "(sidebar). Re-export the Data Extraction with the Budget "
                   "and Forecast columns to pull them automatically.")

    # ---- pickup comparison banner(s) ----
    def _src_phrase(src):
        return {"uploaded file": "an uploaded file",
                "saved history": "saved history"}.get(src, src)

    days_gap = (pd.Timestamp(report_date) - pd.Timestamp(last_report_date)).days
    day_word = "day" if days_gap == 1 else "days"
    if pickup_src == "none yet":
        st.info("No prior on-the-books on file yet, so **pickup since last report "
                "shows 0**. Upload a **Last report** file in the sidebar, or rely "
                "on saved history next time.")
    elif show_lr:
        st.caption(f"📈 **Pickup since last report** compares "
                   f"**{pd.Timestamp(report_date):%b %d, %Y}** to "
                   f"**{pd.Timestamp(last_report_date):%b %d, %Y}** "
                   f"({days_gap} {day_word}, via {_src_phrase(pickup_src)}).")

    if show_week:
        if has_week:
            wgap = (pd.Timestamp(report_date) - pd.Timestamp(week_date)).days
            st.caption(f"🗓️ **7-day pickup** compares "
                       f"**{pd.Timestamp(report_date):%b %d, %Y}** to "
                       f"**{pd.Timestamp(week_date):%b %d, %Y}** "
                       f"({wgap} days, via {_src_phrase(week_src)}).")
        elif pickup_view != "Since last report":
            st.warning("No **7-days-ago** file or matching saved snapshot found, "
                       "so the 7-day pickup can't be shown. Upload a 7-days-ago "
                       "Market Segment export in the sidebar.")

    gap_bits = []
    if show_lr:
        gap_bits.append(f"vs last report {pd.Timestamp(last_report_date):%m/%d/%Y}")
    if show_week and has_week:
        gap_bits.append(f"vs 7-days-ago {pd.Timestamp(week_date):%m/%d/%Y}")
    header_line = (f"The Compton&nbsp;&nbsp;|&nbsp;&nbsp;{pd.Timestamp(report_date):%m/%d/%Y}"
                   + ("&nbsp;&nbsp;|&nbsp;&nbsp;" + " · ".join(gap_bits) if gap_bits else ""))

    st.markdown(render_dashboard_html(tables, header_line,
                                      show_lr=show_lr, show_week=show_week),
                unsafe_allow_html=True)

    st.download_button(
        "Download dashboard as Excel",
        data=dashboard_to_excel(tables, header_line.replace("&nbsp;", " "),
                                show_lr=show_lr, show_week=show_week),
        file_name=f"Compton_Dashboard_{pd.Timestamp(report_date):%Y%m%d}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # ---- record tonight's snapshot LAST (after pickup was computed against the
    #      prior one) so the next upload can show pickup automatically. We store
    #      BOTH the monthly agg (for this dashboard) and the date-level OTB (for
    #      the DD's 7-day pickup), merging into any snapshot the DD already made
    #      for the same as-of date. ----
    if auto_save:
        src_name = parts.get("mseg_file", "") if isinstance(parts, dict) else ""
        record_snapshot(hist, report_date, capacity, year, source_name=src_name,
                        monthly_agg=cur_agg,
                        by_date=date_otb_from_extract(data))
        save_history(hist)


st.title("The Compton — Daily Pick-Up Report Builder")
st.caption("Upload your four SynXis / rate-shop exports and get the full pick-up "
           "report instantly. No copy/paste required.")

with st.sidebar:
    page = st.radio("View", ["Pick-Up Report (DD)", "Dashboard"], index=0,
                    help="Switch between the daily pick-up report and the "
                         "owner dashboard (Total / Transient / Group by month).")
    st.markdown("---")
    st.header("1 - Upload your files")
    ups = st.file_uploader(
        "Drop in PCDC, Data Extract, Market Seg and the Rate-Shop file "
        "(all four at once is fine)",
        type=["xlsx"], accept_multiple_files=True,
    )
    pcdc7_file = st.file_uploader(
        "Optional: 7-day PCDC (for day-by-day 7-day pickup)",
        type=["xlsx"], accept_multiple_files=False, key="pcdc7_up",
        help="Run your PCDC change report over a 7-day window and drop it here. "
             "Its per-date room change becomes a '7-Day P/U' column set in the "
             "report below. Your normal daily/last-report PCDC is unaffected.")

if not ups:
    st.info("Upload your files in the sidebar to build the report.")
    st.stop()

parts = detect_and_load(ups)

# Parse the optional 7-day PCDC into its own slot (kept separate from the daily
# PCDC so the two never collide).
if pcdc7_file is not None:
    try:
        _p7 = _read_all_sheets(pcdc7_file)
        _cr = [s for s in _p7 if "changereport" in s.lower()]
        if _cr:
            parts["pcdc7"] = _parse_pcdc(_p7[_cr[0]])
        else:
            st.sidebar.warning("That 7-day PCDC didn't contain a ChangeReport "
                               "sheet — skipping the 7-day pickup column.")
    except Exception as e:
        st.sidebar.warning(f"Couldn't read the 7-day PCDC ({e}).")

# The date range actually covered by the PCDC (pace/pickup) report, if uploaded.
# Data Extract typically spans the WHOLE YEAR, but PCDC only covers its shorter
# pace window (e.g. ~60-90 days). We use this as the default view so the report
# isn't cluttered with the full year of Data-Extract-only dates.
_pcdc_min = _pcdc_max = None
if parts.get("pcdc") is not None and "Occupancy Date" in parts["pcdc"].columns:
    _pcdc_dates = pd.to_datetime(parts["pcdc"]["Occupancy Date"], errors="coerce").dropna()
    if len(_pcdc_dates):
        _pcdc_min, _pcdc_max = _pcdc_dates.min().date(), _pcdc_dates.max().date()

# 'Days Left' should always default to the REAL current date — not a date read
# out of the PCDC file — since the file could have been exported days ago and a
# stale reference date silently throws off every Days Left calculation.
#
# IMPORTANT Streamlit rule: once a widget with a given `key` has been created,
# you can NOT overwrite st.session_state[key] later in that same script run —
# doing so raises a StreamlitAPIException. The only safe ways to set a widget's
# value from code are (a) BEFORE the widget is instantiated in this run, or
# (b) inside a button's `on_click=` callback (callbacks run before the widget
# is re-created on the next rerun). Both are used below.
real_today = pd.Timestamp.today().date()

# (a) Seed/refresh the value BEFORE the widget exists this run:
#     - first time this session has ever seen this key, OR
#     - a new/different set of files was just uploaded (by name+size) —
#       this is what guarantees the date snaps back to today even if an
#       earlier session left a stale value sitting in session_state.
_file_sig = tuple(sorted((f.name, getattr(f, "size", None)) for f in ups))
if ("as_of_input" not in st.session_state
        or st.session_state.get("_uploaded_file_sig") != _file_sig):
    st.session_state["_uploaded_file_sig"] = _file_sig
    st.session_state["as_of_input"] = real_today

# (b) Callback for the manual reset button — safe because callbacks run
#     before the widget is redrawn, never in the same pass as the widget.
def _reset_as_of_to_today():
    st.session_state["as_of_input"] = pd.Timestamp.today().date()

with st.sidebar:
    st.header("2 - Settings")
    st.caption(f"🕒 Server's current date right now: **{real_today:%b %d, %Y}**")
    as_of = st.date_input(
        "Report as-of date (drives 'Days Left')",
        key="as_of_input",
        help="Defaults to today's real date every time you upload files. Change "
             "it here only if you want to look at pace as of a different date.",
    )
    st.button("🔄 Reset as-of date to today", use_container_width=True,
              on_click=_reset_as_of_to_today)
    if parts.get("pcdc_asof"):
        st.caption(f"ℹ️ For reference, this PCDC report was generated on "
                   f"**{pd.Timestamp(parts['pcdc_asof']):%b %d, %Y}**.")
    if page != "Dashboard":
        comp_method = st.radio("Comp Set Avg method", ["Average", "Median"], horizontal=True)
        show_change = st.radio(
            "Show competitor rate change vs.",
            ["None", "Yesterday", "3 days ago", "7 days ago"],
            help="Adds a 'Change' column beside each competitor, pulled from the "
                 "matching 'vs.' tab in the rate-shop file.",
        )
        show_own_change = st.checkbox(
            "Show our BAR change (7 days)", value=True,
            help="Adds a 'BAR Chg 7d' column showing how our own rate moved over the "
                 "last 7 days (from the rate-shop file).",
        )
    
        st.markdown("---")
        st.header("3 - Signal sensitivity")
        st.caption("The Compton prices as the market leader (#1 in the comp set). "
                   "The ⚡ Action column is purely about rate-leadership position.")
        with st.expander("🔴 Rate leadership (#1 vs #2)", expanded=False):
            st.caption("🔴 Not #1 = a competitor is at or above our BAR.  "
                       "🟠 Too close to #2 = we're #1, but by less than the cushion below.")
            lead_buffer = st.slider("Minimum $ cushion we want over the #2 (top competitor)",
                                    5, 150, 20, 5,
                                    help="If we're #1 but ahead by less than this, flag "
                                         "'Too close to #2'. If a competitor is at/above our "
                                         "BAR, that's always flagged as 'Not #1' regardless "
                                         "of this slider.")
        with st.expander("🔥 Spike alert", expanded=False):
            st.caption("This self-calibrates to this property's own pickup pattern instead of "
                       "a fixed number — it flags days with unusually high pickup relative to "
                       "everything else in the current report.")
            pickup_pctile = st.slider("Flag pickup at or above this percentile of all pickup days",
                                      50, 99, 75, 1,
                                      help="Lower = flags MORE days (e.g. 75 flags the busiest "
                                           "quarter of pickup days). Higher = flags fewer, more "
                                           "extreme outliers.")
            pickup_min = st.slider("...but never flag below this many rooms (floor)",
                                   1, 30, 3, 1,
                                   help="Keeps quiet reports from flagging tiny 1-room pickups "
                                        "just because they're 'relatively' high.")
    
        thr = {"lead_buffer": lead_buffer,
               "pickup_pctile": pickup_pctile, "pickup_min": pickup_min}
    
        st.markdown("---")
        st.caption("Budget & forecast are pulled automatically from the Data Extract "
                   "(Budget / Forecast columns) into the KPI Summary tab.")
        st.caption("Tip: leave the file names as SynXis exports them — the app "
                   "auto-detects each report by its sheets.")

# --------------------------------------------------------------------------- #
#  Page router: render the Dashboard and stop before the pick-up report runs.
# --------------------------------------------------------------------------- #
if page == "Dashboard":
    render_dashboard_page(parts, as_of)
    st.stop()

c1, c2, c3, c4 = st.columns(4)
c1.metric("PCDC",         "OK" if parts["pcdc"] is not None else "-")
c2.metric("Data Extract", "OK" if parts["data"] is not None else "-")
c3.metric("Market Seg",   "OK" if parts["mseg"] is not None else "-")
c4.metric("Rate Shop",    "OK" if parts["shop"] is not None else "-")

st.caption(f"📅 'Days Left' counts from **{pd.Timestamp(as_of):%b %d, %Y}** "
           f"(sidebar 'Report as-of date'). Defaults to today; change it in the "
           f"sidebar if you want a different reference date.")

report = build_report(parts, as_of, comp_method, show_change, thr, show_own_change)

# =============================================================== #
#  Monthly recap (owner view) — current / next / following month
# =============================================================== #
st.subheader("Monthly Recap — On the Books vs Forecast vs Budget vs Last Year")
_raw = parts.get("data")
if _raw is not None and "Occupancy Date" in getattr(_raw, "columns", []):
    labels, metrics, fc_kind = compute_monthly_recap(_raw, as_of)

    _pct_metrics = {"Occupancy %"}
    _money_metrics = {"ADR", "RevPAR", "Room Revenue"}

    def _fmt(metric, v):
        if not isinstance(v, (int, float)) or pd.isna(v):
            return "–"
        if metric in _pct_metrics:
            return f"{v*100:,.1f}%"
        if metric in _money_metrics:
            return f"${v:,.0f}"
        return f"{v:,.0f}"

    def _fmt_var(metric, v):
        if not isinstance(v, (int, float)) or pd.isna(v):
            return "–"
        if metric in _pct_metrics:
            return f"{v*100:+,.1f} pts"
        if metric in _money_metrics:
            # sign must come BEFORE the $ (e.g. "+$64", not "$+64")
            sign = "+" if v >= 0 else "-"
            return f"{sign}${abs(v):,.0f}"
        return f"{v:+,.0f}"

    # one table per month, metrics as rows, scenarios + variances as columns
    cols_order = ["On the Books", "Forecast", "Budget", "Last Year",
                  "Fcst vs Budget", "OTB vs Budget"]
    for mi, label in enumerate(labels):
        rows = {}
        for metric, per_month in metrics.items():
            m = per_month[mi]
            fvb = (m["fcst"] - m["budget"]) if pd.notna(m["fcst"]) and pd.notna(m["budget"]) else np.nan
            ovb = (m["otb"] - m["budget"]) if pd.notna(m["otb"]) and pd.notna(m["budget"]) else np.nan
            rows[metric] = {
                "On the Books": _fmt(metric, m["otb"]),
                "Forecast": _fmt(metric, m["fcst"]),
                "Budget": _fmt(metric, m["budget"]),
                "Last Year": _fmt(metric, m["stly"]),
                "Fcst vs Budget": _fmt_var(metric, fvb),
                "OTB vs Budget": _fmt_var(metric, ovb),
                "_fvb": fvb, "_ovb": ovb,
            }
        tdf = pd.DataFrame(rows).T[cols_order]

        def _color_var(col):
            key = "_fvb" if col.name == "Fcst vs Budget" else "_ovb"
            styles = []
            for metric in col.index:
                raw = rows[metric][key]
                if isinstance(raw, (int, float)) and pd.notna(raw):
                    styles.append("color:#1a7f37;font-weight:600" if raw >= 0 else "color:#cf222e;font-weight:600")
                else:
                    styles.append("")
            return styles

        st.markdown(f"**{label}**")
        sty = tdf.style.apply(_color_var, subset=["Fcst vs Budget"]) \
                       .apply(_color_var, subset=["OTB vs Budget"])
        st.dataframe(sty, use_container_width=True)

    st.caption(f"Forecast source: {fc_kind}. RevPAR & Occupancy use physical capacity "
               "(available room-nights). Budget, forecast, and last-year all pulled from the "
               "Data Extract. The Excel export's KPI Summary tab adds the Group vs Transient split.")
else:
    st.info("Upload the Data Extract with Budget/Forecast columns to see the monthly recap.")

# =============================================================== #
#  Filters
# =============================================================== #
st.subheader("Filters")
min_d, max_d = report["Date"].min(), report["Date"].max()

_window_options = ["Custom", "Next 7 days", "Next 14 days", "Next 30 days",
                  "Next 60 days", "Next 90 days", "All future dates", "All dates (incl. past)"]
_default_window_idx = 7
if _pcdc_min is not None:
    _window_options = ["PCDC report range"] + _window_options
    _default_window_idx = 0     # default to exactly the dates covered by PCDC

fa, fb, fc = st.columns([1.2, 1, 1])
window = fa.selectbox("Booking window", _window_options, index=_default_window_idx,
                      help="'PCDC report range' (default, when a PCDC file is uploaded) "
                           "shows exactly the dates covered by your PCDC pace report — "
                           f"{_pcdc_min:%b %d} to {_pcdc_max:%b %d, %Y}"
                           if _pcdc_min is not None else
                           "'All future dates' or any 'Next X days' option will HIDE "
                           "dates before your as-of date (negative Days Left) — use "
                           "the look-back option below to also see recent past dates.")
dow_pick = fb.multiselect("Day of week", ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"],
                          default=["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"])
action_pick = fc.multiselect("Action", ["🔴 Not #1", "🟠 Too close to #2", "🟠 At ceiling", "✅ Leading"],
                             default=["🔴 Not #1", "🟠 Too close to #2", "🟠 At ceiling", "✅ Leading"])

lookback = st.number_input(
    "Also include the last N days before the as-of date (negative Days Left)",
    min_value=0, max_value=90, value=0, step=1,
    help="0 = forward-looking only (default). Set this above 0 to also pull in recent "
         "past dates alongside 'Next X days' / 'All future dates', e.g. to review last "
         "week's pickup. Ignored when 'Custom' or 'All dates (incl. past)' is selected.")

fd, fe, ff, fg = st.columns([1, 1, 1, 1])
events_only = fd.checkbox("Events only", value=False)
occ_band = fe.select_slider("Occ Forecast % band",
                            options=[0, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100],
                            value=(0, 100))
var_thresh = ff.number_input("Min |variance vs STLY| (rooms)", min_value=0, value=0, step=1)
pickup_only = fg.checkbox("🔥 Spike days only", value=False)

# Pickup focus — filter rows to just the dates that moved, since last report or
# over the 7-day window (the latter needs the optional 7-day PCDC).
_has_wk_col = "7-Day P/U | Total Hotel" in report.columns
_focus_opts = ["All dates", "Picked up since last report"]
if _has_wk_col:
    _focus_opts.append("7-day pickup only")
pickup_focus = st.selectbox(
    "Pickup focus", _focus_opts, index=0,
    help="'Picked up since last report' keeps dates whose Total-Hotel rooms "
         "changed since your last report. '7-day pickup only' (needs a 7-day "
         "PCDC) keeps dates that moved over the last 7 days — handy for the "
         "7-day recap.")

# custom date range (used when window == Custom)
start, end = min_d, max_d          # defaults so the filename always has a range
if window == "Custom":
    d1, d2 = st.columns(2)
    default_start = max(min_d, as_of)
    start = d1.date_input("From", value=default_start, min_value=min_d, max_value=max_d)
    end   = d2.date_input("To",   value=max_d,         min_value=min_d, max_value=max_d)

# ---- apply filters ---- #
v = report.copy()
win_map = {"Next 7 days": 7, "Next 14 days": 14, "Next 30 days": 30,
           "Next 60 days": 60, "Next 90 days": 90}
floor_days = -int(lookback)          # e.g. lookback=7 -> also show Days Left down to -7
if window == "PCDC report range" and _pcdc_min is not None:
    v = v[(v["Date"] >= _pcdc_min) & (v["Date"] <= _pcdc_max)]
elif window in win_map:
    v = v[(v["Days Left"] >= floor_days) & (v["Days Left"] <= win_map[window])]
elif window == "All future dates":
    v = v[v["Days Left"] >= floor_days]
elif window == "All dates (incl. past)":
    pass  # no Days Left filter at all — show everything in the uploaded range
else:  # Custom
    v = v[(v["Date"] >= start) & (v["Date"] <= end)]

dow_full = {"Mon": "Monday", "Tue": "Tuesday", "Wed": "Wednesday", "Thu": "Thursday",
            "Fri": "Friday", "Sat": "Saturday", "Sun": "Sunday"}
keep_dow = [dow_full[d] for d in dow_pick]
v = v[v["DOW"].isin(keep_dow)]

if events_only:
    v = v[v["Events"].astype(str).str.strip() != ""]

if "⚡ Action" in v:
    a = v["⚡ Action"].fillna("")
    sel = pd.Series(False, index=v.index)
    if "🔴 Not #1" in action_pick:          sel |= a.str.contains("Not #1")
    if "🟠 Too close to #2" in action_pick: sel |= a.str.contains("Too close")
    if "🟠 At ceiling" in action_pick:      sel |= a.str.contains("At ceiling")
    if "✅ Leading" in action_pick:         sel |= a.str.contains("Leading")
    sel |= (a.str.strip() == "")      # keep past/blank rows
    v = v[sel]

ofp_col = pd.to_numeric(v.get("Occ Forecast % | Total Hotel"), errors="coerce") * 100
lo, hi = occ_band
v = v[(ofp_col.isna()) | ((ofp_col >= lo) & (ofp_col <= hi))]

if var_thresh > 0 and "Rooms OTB STLY | Total Hotel | Change" in v:
    var_col = pd.to_numeric(v["Rooms OTB STLY | Total Hotel | Change"], errors="coerce").abs()
    v = v[var_col >= var_thresh]

if pickup_only and "🔥 Spike" in v:
    v = v[v["🔥 Spike"].astype(str).str.strip() != ""]

if pickup_focus == "Picked up since last report" and "Rooms Sold | Total Hotel | Change" in v:
    _pc = pd.to_numeric(v["Rooms Sold | Total Hotel | Change"], errors="coerce").fillna(0)
    v = v[_pc != 0]
elif pickup_focus == "7-day pickup only" and "7-Day P/U | Total Hotel" in v:
    _pc = pd.to_numeric(v["7-Day P/U | Total Hotel"], errors="coerce").fillna(0)
    v = v[_pc != 0]

view = v.reset_index(drop=True)

st.subheader(f"Pick-Up Report — {len(view)} dates")

pct_cols   = [c for c in view.columns if c.startswith("Occ Forecast %")]
shop_all   = [c for c in view.columns if c.startswith("Competitor Shops")]
shop_chg   = [c for c in shop_all if c.endswith("Change")]
shop_cur   = [c for c in shop_all if c not in shop_chg]
money_cols = (["BAR", "Comp Set Avg", "Last Room Value", "Estimated ADR"]
              + [c for c in view.columns if c.startswith("Booked ADR")])
rooms_chg  = [c for c in view.columns
              if (c.startswith("Rooms Sold") or c.startswith("Rooms OTB STLY"))
              and c.endswith("Change")]
wk_pu_cols = [c for c in view.columns if c.startswith("7-Day P/U")]
var_cols   = ([c for c in view.columns if "Variance" in c] + shop_chg + rooms_chg
              + wk_pu_cols
              + [c for c in ("Lead Gap vs Comp", "BAR Chg 7d") if c in view.columns])
otb_sold   = "Rooms Left to Sell"      # heat-map column

def _money(v):
    if isinstance(v, (int, float)) and not pd.isna(v):
        return f"${v:,.0f}"
    return "" if v is None or (isinstance(v, float) and pd.isna(v)) else str(v)

def _signed(v):
    if isinstance(v, (int, float)) and not pd.isna(v) and v != 0:
        return f"{v:+,.0f}"
    return ""

fmt = {}
for c in pct_cols:
    fmt[c] = "{:.1%}"
for c in money_cols:
    fmt[c] = "${:,.0f}"
for c in shop_cur:
    fmt[c] = _money
for c in shop_chg:
    fmt[c] = _signed
for c in ("Lead Gap vs Comp", "BAR Chg 7d"):
    if c in view.columns:
        fmt[c] = _signed
for c in wk_pu_cols:
    fmt[c] = _signed
for c in view.columns:
    if c in ("Days Left", "OOO", "Ovrbk") or c.startswith(("Rooms Sold", "Rooms OTB STLY",
             "Remaining Demand", "Occ Forecast |")) and c not in pct_cols:
        fmt.setdefault(c, "{:,.0f}")

styler = (view.style
          .format(fmt, na_rep="")
          .map(lambda v: "color:#1a7f37;font-weight:600" if isinstance(v, (int, float)) and v > 0
               else ("color:#cf222e;font-weight:600" if isinstance(v, (int, float)) and v < 0 else ""),
               subset=var_cols))


# Heat map on the OTB rooms sold column — pure-Python red->yellow->green
# gradient (no matplotlib needed, so it works on Streamlit Cloud out of the box).
def _heat_bg(series):
    # For "Rooms Left to Sell": GREEN = lots of rooms left (wide open),
    # RED = few rooms left (near sell-out).  Low value -> red, high value -> green.
    vals = pd.to_numeric(series, errors="coerce")
    lo, hi = vals.min(), vals.max()
    span = (hi - lo) or 1
    out = []
    for v in vals:
        if pd.isna(v):
            out.append("")
            continue
        t = (v - lo) / span                    # low rooms left -> red, high -> green
        if t < 0.5:
            f = t / 0.5
            red, grn, blu = 248, int(105 + f * (235 - 105)), int(107 + f * (132 - 107))
        else:
            f = (t - 0.5) / 0.5
            red, grn, blu = int(255 - f * (255 - 99)), int(235 - f * (235 - 190)), int(132 - f * (132 - 123))
        out.append(f"background-color: #{red:02X}{grn:02X}{blu:02X}")
    return out

# Heat map is applied directly inside the pick-up report on the OTB column
if otb_sold in view.columns and view[otb_sold].notna().any():
    styler = styler.apply(_heat_bg, subset=[otb_sold])

# Colour the ⚡ Action column so recommendations pop
def _action_bg(series):
    out = []
    for v in series.astype(str):
        if "Not #1" in v:
            out.append("background-color:#F8D7DA;color:#842029;font-weight:600")
        elif "Too close" in v:
            out.append("background-color:#FFF3CD;color:#664D03;font-weight:600")
        elif "At ceiling" in v:
            out.append("background-color:#FFE5B4;color:#7A4A00;font-weight:600")
        elif "Leading" in v:
            out.append("background-color:#D1E7DD;color:#0F5132;font-weight:600")
        else:
            out.append("")
    return out

if "⚡ Action" in view.columns:
    styler = styler.apply(_action_bg, subset=["⚡ Action"])

# Colour the 🔥 Spike column so big pickup days pop
def _pickup_bg(series):
    return ["background-color:#FDE7CE;color:#B45309;font-weight:600" if "Spike" in str(v)
            else "" for v in series]

if "🔥 Spike" in view.columns:
    styler = styler.apply(_pickup_bg, subset=["🔥 Spike"])

st.dataframe(styler, use_container_width=True, height=560)

# filename reflects the actual filtered date range (falls back to the full range)
if len(view) > 0:
    _fn_start, _fn_end = view["Date"].min(), view["Date"].max()
else:
    _fn_start, _fn_end = start, end

st.download_button(
    "Download report as Excel",
    data=to_excel(view, {"data_raw": parts.get("data"), "as_of": as_of}),
    file_name=f"Compton_PickUp_{_fn_start:%Y%m%d}_{_fn_end:%Y%m%d}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

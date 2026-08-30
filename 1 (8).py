"""
The Compton — Daily Pick-Up Report Builder
Run it with:   streamlit run app.py
"""

import io
import numpy as np
import pandas as pd
import streamlit as st

st.set_page_config(page_title="Compton Daily Pick-Up Report", layout="wide")


def _num(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return np.nan


def _read_all_sheets(uploaded_file):
    xls = pd.ExcelFile(uploaded_file, engine="openpyxl")
    return {s: xls.parse(s, header=None) for s in xls.sheet_names}


def detect_and_load(uploaded_files):
    found = {"pcdc": None, "data": None, "mseg": None, "shop": None,
             "shop_chg_1": None, "shop_chg_3": None, "shop_chg_7": None,
             "pcdc_asof": None}

    for f in uploaded_files:
        try:
            sheets = _read_all_sheets(f)
        except Exception as e:
            st.warning(f"Could not open **{f.name}** ({e}).")
            continue
        names = [s.lower() for s in sheets.keys()]

        if any("rate" in n for n in names) and any("overview" in n for n in names):
            raw = sheets[[s for s in sheets if s.lower() == "rates"][0]]
            found["shop"] = _parse_rateshop(raw)
            for tab, key in [("vs. yesterday", "shop_chg_1"),
                             ("vs. 3 days ago", "shop_chg_3"),
                             ("vs. 7 days ago", "shop_chg_7")]:
                match = [s for s in sheets if s.lower() == tab]
                if match:
                    found[key] = _parse_shop_change(sheets[match[0]])

        elif any("changereport" in n for n in names):
            raw = sheets[[s for s in sheets if "changereport" in s.lower()][0]]
            found["pcdc"] = _parse_pcdc(raw)
            crit = [s for s in sheets if s.lower() == "report criteria"]
            if crit:
                found["pcdc_asof"] = _parse_pcdc_asof(sheets[crit[0]])

        elif any("market segment" in n for n in names):
            raw = sheets[[s for s in sheets if s.lower() == "market segment"][0]]
            found["mseg"] = _parse_mseg(raw)

        elif any(n == "property" for n in names):
            raw = sheets[[s for s in sheets if s.lower() == "property"][0]]
            found["data"] = _parse_data(raw)

        else:
            st.warning(f"Didn't recognise **{f.name}** — skipping. "
                       f"(sheets: {', '.join(sheets.keys())})")
    return found


def _parse_data(raw):
    df = raw.copy()
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)
    df["Occupancy Date"] = pd.to_datetime(df["Occupancy Date"], errors="coerce")
    df = df.dropna(subset=["Occupancy Date"])
    return df


def _parse_pcdc(raw):
    df = raw.iloc[3:].copy()
    df[0] = pd.to_datetime(df[0], errors="coerce")
    df = df.dropna(subset=[0]).reset_index(drop=True)
    out = pd.DataFrame({"Occupancy Date": df[0]})
    out["pc_event"]      = df[2]
    out["trans_cur"]     = pd.to_numeric(df[3], errors="coerce")
    out["trans_chg"]     = pd.to_numeric(df[4], errors="coerce")
    out["grp_cur"]       = pd.to_numeric(df[5], errors="coerce")
    out["grp_chg"]       = pd.to_numeric(df[6], errors="coerce")
    out["grp_blocked"]   = pd.to_numeric(df[7], errors="coerce")
    out["grp_avail"]     = pd.to_numeric(df[8], errors="coerce")
    out["grp_pickup"]    = pd.to_numeric(df[9], errors="coerce")
    out["fcst_trans"]    = pd.to_numeric(df[10], errors="coerce")
    out["fcst_grp"]      = pd.to_numeric(df[12], errors="coerce")
    out["rev_trans"]     = pd.to_numeric(df[14], errors="coerce")
    out["rev_grp"]       = pd.to_numeric(df[16], errors="coerce")
    out["adr_trans"]     = pd.to_numeric(df[22], errors="coerce")
    out["adr_grp"]       = pd.to_numeric(df[24], errors="coerce")
    return out


def _parse_pcdc_asof(raw):
    """
    Read the PCDC 'Report Criteria' sheet and return the reference date for
    'Days Left'.  Prefers 'Analysis Start Date'; falls back to 'Generated On'
    (the export/created date).  Returns a datetime.date or None.
    """
    try:
        # find the header row that contains the labels, then the value row below it
        for i in range(len(raw) - 1):
            rowvals = [str(x).strip().lower() for x in raw.iloc[i].tolist()]
            if "analysis start date" in rowvals or "generated on" in rowvals:
                header = raw.iloc[i].tolist()
                values = raw.iloc[i + 1].tolist()
                lut = {str(h).strip().lower(): v for h, v in zip(header, values)}
                for key in ("analysis start date", "activity start date", "generated on"):
                    if key in lut and lut[key] not in (None, ""):
                        d = pd.to_datetime(lut[key], errors="coerce")
                        if pd.notna(d):
                            return d.date()
    except Exception:
        pass
    return None


def _parse_mseg(raw):
    df = raw.copy()
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)
    df["Occupancy Date"] = pd.to_datetime(df["Occupancy Date"], errors="coerce")
    df = df.dropna(subset=["Occupancy Date"])
    df["Occupancy On Books This Year"] = pd.to_numeric(
        df["Occupancy On Books This Year"], errors="coerce").fillna(0)
    df["is_trans"] = df["Market Segment"].astype(str).str.startswith("Transient")
    agg = df.groupby("Occupancy Date").apply(
        lambda g: pd.Series({
            "ms_trans": g.loc[g["is_trans"], "Occupancy On Books This Year"].sum(),
            "ms_group": g.loc[~g["is_trans"], "Occupancy On Books This Year"].sum(),
        })
    ).reset_index()
    return agg


def _clean_shop(x):
    if x is None:
        return None
    v = _num(x)
    if not np.isnan(v):
        return int(v) if float(v).is_integer() else v
    s = str(x).strip()
    return s if s else None


def _parse_rateshop(raw):
    df = raw.iloc[5:].copy()
    df[2] = pd.to_datetime(df[2], errors="coerce")
    df = df.dropna(subset=[2]).reset_index(drop=True)
    out = pd.DataFrame({"Occupancy Date": df[2]})
    out["own_bar"] = df[4].map(_num)
    for key, col in [("c_21c", 5), ("c_motto", 6), ("c_ac", 7), ("c_dt", 8)]:
        out[key] = df[col].map(_clean_shop)
        out[key + "_n"] = df[col].map(_num)
    return out


def _parse_shop_change(raw):
    df = raw.iloc[5:].copy()
    df[2] = pd.to_datetime(df[2], errors="coerce")
    df = df.dropna(subset=[2]).reset_index(drop=True)
    out = pd.DataFrame({"Occupancy Date": df[2]})
    for key, chg_col in [("c_21c", 8), ("c_motto", 10), ("c_ac", 12), ("c_dt", 14)]:
        out[key + "_chg"] = df[chg_col].map(_num) if chg_col in df.columns else np.nan
    return out


def build_report(parts, as_of, comp_method="Average", show_change="None", thr=None):
    data, pcdc, mseg, shop = parts["data"], parts["pcdc"], parts["mseg"], parts["shop"]
    if data is None:
        st.error("The **Data Extract** file is required (couldn't find a 'Property' sheet).")
        st.stop()

    df = data.copy()
    for extra in (pcdc, mseg, shop):
        if extra is not None:
            df = df.merge(extra, on="Occupancy Date", how="left")

    chg_key = {"Yesterday": "shop_chg_1", "3 days ago": "shop_chg_3",
               "7 days ago": "shop_chg_7"}.get(show_change)
    chg_df = parts.get(chg_key) if chg_key else None
    if chg_df is not None:
        df = df.merge(chg_df, on="Occupancy Date", how="left")

    def dnum(col):
        return pd.to_numeric(df.get(col), errors="coerce")

    cap        = dnum("Physical Capacity This Year")
    occ_ty     = dnum("Occupancy On Books This Year")
    occ_stly   = dnum("Occupancy On Books STLY")
    tr_ty      = dnum("Rooms Sold - Transient This Year")
    tr_stly    = dnum("Rooms Sold - Transient STLY")
    gr_ty      = dnum("Rooms Sold - Group This Year")
    gr_stly    = dnum("Rooms Sold - Group STLY")
    dem_tot    = dnum("User Total Demand - Total This Year")
    dem_grp    = dnum("User Constrained Total Demand - Group This Year")
    dem_tr     = dnum("User Unconstrained Total Demand - Transient This Year")

    r = pd.DataFrame()
    r["DOW"]  = df["Day of Week"]
    r["Date"] = df["Occupancy Date"].dt.date
    r["Days Left"] = (df["Occupancy Date"] - pd.Timestamp(as_of)).dt.days

    ev = df.get("Special Event This Year")
    ev = ev.fillna("") if ev is not None else ""
    if "pc_event" in df:
        ev = ev.where(ev.astype(str).str.strip() != "", df["pc_event"].fillna(""))
    r["Events"] = pd.Series(ev, index=df.index).replace("nan", "")

    r["Rooms Left to Sell"] = dnum("Remaining Capacity This Year")
    r["BAR"]                = dnum("BAR")

    if "c_21c_n" in df:
        comps_n = df[["c_21c_n", "c_motto_n", "c_ac_n", "c_dt_n"]]
        r["Comp Set Avg"] = (comps_n.mean(axis=1) if comp_method == "Average"
                             else comps_n.median(axis=1)).round(0)
    else:
        r["Comp Set Avg"] = np.nan

    r["Last Room Value"] = dnum("Last Room Value This Year")

    add_chg = chg_df is not None
    for name, key in [("21c Museum Hotel Bentonville - MGallery", "c_21c"),
                      ("Motto By Hilton Bentonville Downtown",    "c_motto"),
                      ("AC Hotel by Marriott Bentonville",        "c_ac"),
                      ("DoubleTree Suites by Hilton Bentonville", "c_dt")]:
        base = "Competitor Shops | " + name
        if add_chg:
            r[base + " | Current"] = df.get(key)
            r[base + " | Change"]  = df.get(key + "_chg")
        else:
            r[base] = df.get(key)

    r["OOO"]   = dnum("Rooms N/A - Out of Order This Year")
    r["Ovrbk"] = dnum("Overbooking This Year")

    tc, tch = df.get("trans_cur"), df.get("trans_chg")
    gc, gch = df.get("grp_cur"),   df.get("grp_chg")
    r["Rooms Sold | Total Hotel | Current"]   = (tc.fillna(0) + gc.fillna(0)) if tc is not None else occ_ty
    r["Rooms Sold | Total Hotel | Change"]    = (tch.fillna(0) + gch.fillna(0)) if tch is not None else np.nan
    r["Rooms Sold | Total Transient | Current"] = tc
    r["Rooms Sold | Total Transient | Change"]  = tch
    r["Rooms Sold | Total Group | Current"]     = gc
    r["Rooms Sold | Total Group | Change"]      = gch
    r["Rooms Sold | Total Group | Blocked"]     = df.get("grp_blocked")
    r["Rooms Sold | Total Group | P/U"]         = df.get("grp_pickup")
    r["Rooms Sold | Total Group | Remaining"]   = df.get("grp_avail")

    r["Rooms OTB STLY | Total Hotel | Current"]     = occ_stly
    r["Rooms OTB STLY | Total Hotel | Change"]      = occ_ty - occ_stly
    r["Rooms OTB STLY | Total Transient | Current"] = tr_stly
    r["Rooms OTB STLY | Total Transient | Change"]  = tr_ty - tr_stly
    r["Rooms OTB STLY | Total Group | Current"]     = gr_stly
    r["Rooms OTB STLY | Total Group | Change"]      = gr_ty - gr_stly

    r["Remaining Demand | Total Hotel"]     = (dem_tot - occ_ty).clip(lower=0)
    r["Remaining Demand | Total Transient"] = (dem_tr - tr_ty).clip(lower=0)
    r["Remaining Demand | Total Group"]     = (dem_grp - gr_ty).clip(lower=0)

    ft, fg = df.get("fcst_trans"), df.get("fcst_grp")
    fcst_tot = (ft.fillna(0) + fg.fillna(0)) if ft is not None else dnum("Forecasted Room Revenue This Year")*np.nan
    r["Occ Forecast | Total Hotel"]     = fcst_tot.round(1) if ft is not None else np.nan
    r["Occ Forecast | Total Transient"] = ft.round(1) if ft is not None else np.nan
    r["Occ Forecast | Total Group"]     = fg.round(1) if fg is not None else np.nan

    r["Occ Forecast % | Total Hotel"]     = (fcst_tot / cap) if ft is not None else np.nan
    r["Occ Forecast % | Total Transient"] = (ft / cap) if ft is not None else np.nan
    r["Occ Forecast % | Total Group"]     = (fg / cap) if fg is not None else np.nan

    rev_t, rev_g = df.get("rev_trans"), df.get("rev_grp")
    if tc is not None:
        total_rooms = (tc.fillna(0) + gc.fillna(0)).replace(0, np.nan)
        r["Booked ADR | Total Hotel"] = ((rev_t.fillna(0) + rev_g.fillna(0)) / total_rooms).round(2)
    else:
        r["Booked ADR | Total Hotel"] = dnum("ADR On Books This Year")
    r["Booked ADR | Total Transient"] = df.get("adr_trans")
    r["Booked ADR | Total Group"]     = df.get("adr_grp")

    r["Estimated ADR"] = dnum("ADR Forecast This Year")

    # ================================================================= #
    #  PRICING DECISION SIGNALS  (VP-of-Revenue toolkit)
    # ================================================================= #
    bar   = r["BAR"]
    comp  = r["Comp Set Avg"]
    rl    = r["Rooms Left to Sell"]
    rd    = r["Remaining Demand | Total Hotel"]
    ofp   = r["Occ Forecast % | Total Hotel"]
    pace  = r["Rooms OTB STLY | Total Hotel | Change"]      # TY − STLY (+ = ahead)
    dleft = r["Days Left"]
    floor   = dnum("Floor")
    ceiling = dnum("Ceiling")

    # --- Rate Rank badge:  where our BAR sits in the comp set (1 = highest) --- #
    rank_cols = [c for c in ("c_21c_n", "c_motto_n", "c_ac_n", "c_dt_n") if c in df]
    if rank_cols:
        price_df = pd.DataFrame({"own": pd.to_numeric(r["BAR"], errors="coerce").values})
        for cc in rank_cols:
            price_df[cc] = pd.to_numeric(df[cc], errors="coerce").values

        def _rank_row(row):
            own = row["own"]
            allv = [v for v in row.tolist() if pd.notna(v)]
            if pd.isna(own) or not allv:
                return ""
            higher = sum(1 for v in allv if v > own)
            return f"{higher + 1} of {len(allv)}"

        r["Rate Rank"] = [_rank_row(price_df.iloc[i]) for i in range(len(price_df))]
    else:
        r["Rate Rank"] = ""

    # --- Signal thresholds (tunable from the sidebar) --- #
    thr = thr or {}
    comp_occ   = thr.get("comp_occ", 0.90)     # occupancy % that means "compression"
    comp_rooms = thr.get("comp_rooms", 0.10)   # rooms-left as % of capacity for compression
    comp_dem   = thr.get("comp_dem", 5)        # min remaining demand for the rooms-left path
    soft_window = thr.get("soft_window", 21)   # booking window (days) for a "soft" date
    soft_occ    = thr.get("soft_occ", 0.60)    # occupancy % below which a near date is "soft"

    # --- ⚡ Action:  the one-glance recommendation --- #
    cap_safe = cap.replace(0, np.nan)
    # effective occupancy %: use the forecast, fall back to on-the-books ÷ capacity
    ofp_eff = ofp.copy()
    ofp_eff = ofp_eff.where(ofp_eff.notna(), occ_ty / cap_safe)
    compression = (ofp_eff >= comp_occ) | ((rl <= (comp_rooms * cap_safe)) & (rd >= comp_dem))
    opportunity = (bar < comp) & (pace > 0)
    soft        = (dleft >= 0) & (dleft <= soft_window) & (pace < 0) & (ofp_eff < soft_occ)
    at_ceiling  = bar >= ceiling
    at_floor    = bar <= floor

    def _label(i):
        if pd.isna(dleft.iloc[i]) or dleft.iloc[i] < 0:
            return ""                                    # past date
        if bool(compression.iloc[i]):
            return "🔴 Raise · compression" if not bool(at_ceiling.iloc[i]) else "🟠 At ceiling"
        if bool(opportunity.iloc[i]):
            return "🔴 Raise · below comp"
        if bool(soft.iloc[i]):
            return "🟢 Stimulate · soft" if not bool(at_floor.iloc[i]) else "🟠 At floor"
        return "⚪ Hold"

    r["⚡ Action"] = [_label(i) for i in range(len(r))]

    # ---- Enforce the exact original report column order ---- #
    lead = ["⚡ Action", "DOW", "Date", "Days Left", "Events", "Rooms Left to Sell",
            "BAR", "Comp Set Avg", "Rate Rank", "Last Room Value"]
    comp = [c for c in r.columns if c.startswith("Competitor Shops")]
    mid  = ["OOO", "Ovrbk"]
    groups = ["Rooms Sold", "Rooms OTB STLY", "Remaining Demand",
              "Occ Forecast |", "Occ Forecast %", "Booked ADR"]
    grouped = []
    for g in groups:
        grouped += [c for c in r.columns if c.startswith(g)]
    tail = ["Estimated ADR"]
    order = lead + comp + mid + grouped + tail
    order = [c for c in order if c in r.columns]          # keep only existing
    order += [c for c in r.columns if c not in order]     # safety: append any strays
    r = r[order]

    return r


def _levels(col):
    parts = [p.strip() for p in str(col).split("|")]
    while len(parts) < 3:
        parts.append("")
    return tuple(parts[:3])


def to_excel(r):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    lv = [_levels(c) for c in r.columns]
    n = len(lv)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pick-Up Report"

    NAVY, TEAL, ORANGE, PURPLE, GREEN = "16365C", "215967", "E26B0A", "60497A", "76933C"
    white = Font(color="FFFFFF", bold=True, size=9)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

    def PF(hexcol):
        return PatternFill("solid", fgColor=hexcol)

    def seg_color(txt):
        if "Total Hotel" in txt:
            return ORANGE
        if "Total Transient" in txt:
            return PURPLE
        if "Total Group" in txt:
            return GREEN
        return None

    def band_fill(l1, l2):
        return PF(NAVY) if l2 == "" else PF(TEAL)

    def lvl2_fill(l1, l2):
        sc = seg_color(l2)
        if sc:
            return PF(sc)
        if l1 == "Competitor Shops":
            return PF(NAVY)
        return PF(TEAL)

    def lvl3_fill(l1, l2):
        sc = seg_color(l2)
        return PF(sc) if sc else PF(NAVY)

    j = 0
    while j < n:
        l1 = lv[j][0]
        k = j
        while k + 1 < n and lv[k + 1][0] == l1 and lv[j][1]:
            k += 1
        fill = band_fill(l1, lv[j][1])
        if lv[j][1] == "" and l1 in ("Estimated ADR",):
            fill = PF(ORANGE)
        cell = ws.cell(row=1, column=j + 1, value=l1)
        cell.font = white
        cell.alignment = center
        cell.fill = fill
        if lv[j][1] == "":
            ws.merge_cells(start_row=1, start_column=j + 1, end_row=3, end_column=j + 1)
        elif k > j:
            ws.merge_cells(start_row=1, start_column=j + 1, end_row=1, end_column=k + 1)
        j = k + 1

    j = 0
    while j < n:
        l1, l2, l3 = lv[j]
        if l2:
            k = j
            while k + 1 < n and lv[k + 1][0] == l1 and lv[k + 1][1] == l2:
                k += 1
            cell = ws.cell(row=2, column=j + 1, value=l2)
            cell.font = white
            cell.alignment = center
            cell.fill = lvl2_fill(l1, l2)
            if l3 == "":
                ws.merge_cells(start_row=2, start_column=j + 1, end_row=3, end_column=j + 1)
            elif k > j:
                ws.merge_cells(start_row=2, start_column=j + 1, end_row=2, end_column=k + 1)
            j = k + 1
        else:
            j += 1

    for j, (l1, l2, l3) in enumerate(lv, start=1):
        if l3:
            cell = ws.cell(row=3, column=j, value=l3)
            cell.font = white
            cell.alignment = center
            cell.fill = lvl3_fill(l1, l2)

    _hotelnames = ("Bentonville", "MGallery")
    DATA0 = 4
    pct = {i for i, c0 in enumerate(r.columns) if c0.startswith("Occ Forecast %")}
    chg = {i for i, c0 in enumerate(r.columns)
           if any(h in c0 for h in _hotelnames) and c0.endswith("Change")}
    money = {i for i, c0 in enumerate(r.columns)
             if (c0 in ("BAR", "Comp Set Avg", "Last Room Value", "Estimated ADR")
                 or "ADR" in c0 or any(h in c0 for h in _hotelnames))
             and i not in chg}
    otb_sold_col = None
    for i, c0 in enumerate(r.columns):
        if c0 == "Rooms Left to Sell":
            otb_sold_col = i

    def _action_fill(txt):
        t = str(txt)
        if "Raise" in t:     return PatternFill("solid", fgColor="F8D7DA")
        if "Stimulate" in t: return PatternFill("solid", fgColor="D1E7DD")
        if "At " in t:       return PatternFill("solid", fgColor="FFF3CD")
        if "Hold" in t:      return PatternFill("solid", fgColor="E9ECEF")
        return None

    for ri, (_, row) in enumerate(r.iterrows(), start=DATA0):
        for ci2, col in enumerate(r.columns, start=1):
            v = row[col]
            cell = ws.cell(row=ri, column=ci2, value=(None if pd.isna(v) else v))
            cell.border = thin
            cell.font = Font(size=9)
            # center everything; left-align the long-text "Events" column
            halign = "left" if col == "Events" else "center"
            cell.alignment = Alignment(horizontal=halign, vertical="center")
            if col == "⚡ Action":
                af = _action_fill(v)
                if af is not None:
                    cell.fill = af
            idx = ci2 - 1
            if idx in pct:
                cell.number_format = "0.0%"
            elif idx in chg:
                cell.number_format = "+#,##0;-#,##0;\"\""
            elif idx in money:
                cell.number_format = "$#,##0"
            elif isinstance(v, float) and float(v).is_integer():
                cell.number_format = "#,##0"

    if otb_sold_col is not None and len(r) > 0:
        L = get_column_letter(otb_sold_col + 1)
        rng = f"{L}{DATA0}:{L}{DATA0 + len(r) - 1}"
        # Rooms Left to Sell: RED = few left (near sell-out), GREEN = lots left
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(start_type="min", start_color="F8696B",         # red (few left)
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",  # yellow
                           end_type="max", end_color="63BE7B"))            # green (lots left)

    ws.freeze_panes = "C4"
    for j in range(1, n + 1):
        ws.column_dimensions[get_column_letter(j)].width = 12
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["D"].width = 22

    # ================================================================= #
    #  KPI Summary tab (self-contained snapshot for sharing)
    # ================================================================= #
    _build_kpi_sheet(wb, r, NAVY, TEAL, ORANGE, PURPLE, GREEN)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_kpi_sheet(wb, r, NAVY, TEAL, ORANGE, PURPLE, GREEN):
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    def col(name):
        return pd.to_numeric(r[name], errors="coerce") if name in r.columns else pd.Series(dtype=float)

    # future-only slice
    dleft = pd.to_numeric(r.get("Days Left"), errors="coerce")
    fut = r[dleft >= 0] if dleft is not None else r

    def fcol(name):
        return pd.to_numeric(fut[name], errors="coerce") if name in fut.columns else pd.Series(dtype=float)

    otb_ty   = fcol("Rooms Sold | Total Hotel | Current").sum()
    otb_stly = fcol("Rooms OTB STLY | Total Hotel | Current").sum()
    pace_abs = otb_ty - otb_stly
    pace_pct = (pace_abs / otb_stly) if otb_stly else np.nan

    fdl = pd.to_numeric(fut.get("Days Left"), errors="coerce")
    n7  = fut[fdl <= 7]
    n30 = fut[fdl <= 30]
    occ7  = pd.to_numeric(n7.get("Occ Forecast % | Total Hotel"), errors="coerce").mean()
    occ30 = pd.to_numeric(n30.get("Occ Forecast % | Total Hotel"), errors="coerce").mean()

    act = fut.get("⚡ Action", pd.Series(dtype=str)).fillna("")
    n_raise     = int(act.str.contains("Raise").sum())
    n_stimulate = int(act.str.contains("Stimulate").sum())
    n_limit     = int(act.str.contains("At ").sum())
    n_hold      = int(act.str.contains("Hold").sum())

    rate_gap = (fcol("BAR") - fcol("Comp Set Avg")).mean()

    # transient / group pace
    tr_var = fcol("Rooms OTB STLY | Total Transient | Change").sum()
    gr_var = fcol("Rooms OTB STLY | Total Group | Change").sum()

    ks = wb.create_sheet("KPI Summary", 0)   # put it first
    white = Font(color="FFFFFF", bold=True, size=12)
    big   = Font(bold=True, size=20)
    lbl   = Font(size=10, color="555555")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left", vertical="center")

    ks.merge_cells("A1:F1")
    t = ks["A1"]; t.value = "The Compton — Pick-Up Snapshot"
    t.font = white; t.alignment = center
    t.fill = PatternFill("solid", fgColor=NAVY)
    ks.row_dimensions[1].height = 30

    ks.merge_cells("A2:F2")
    s = ks["A2"]
    dmin = r["Date"].min() if "Date" in r.columns else ""
    dmax = r["Date"].max() if "Date" in r.columns else ""
    s.value = f"Report window: {dmin} to {dmax}   •   {len(fut)} future dates"
    s.font = lbl; s.alignment = center

    # KPI cards: (label, value, hint_color)
    def pct(x):  return f"{x*100:,.0f}%" if pd.notna(x) else "–"
    def signed(x): return f"{x:+,.0f}" if pd.notna(x) else "–"
    cards = [
        ("Rooms OTB (future)", f"{otb_ty:,.0f}", TEAL),
        ("Pace vs STLY", f"{signed(pace_abs)}  ({pct(pace_pct)})",
         GREEN if (pd.notna(pace_abs) and pace_abs >= 0) else "C00000"),
        ("Rate vs Comp Set", f"${rate_gap:,.0f}" if pd.notna(rate_gap) else "–", NAVY),
        ("Next 7 Occ %", pct(occ7), ORANGE),
        ("Next 30 Occ %", pct(occ30), ORANGE),
        ("Transient pace / Group pace", f"{signed(tr_var)} / {signed(gr_var)}", PURPLE),
    ]
    row0 = 4
    for i, (label, value, color) in enumerate(cards):
        rr = row0 + (i // 3) * 3
        cc = 1 + (i % 3) * 2                       # columns A/C/E (each card spans 2)
        c1 = ks.cell(row=rr, column=cc, value=label)
        c1.font = Font(bold=True, size=10, color="FFFFFF")
        c1.alignment = center
        c1.fill = PatternFill("solid", fgColor=color)
        ks.merge_cells(start_row=rr, start_column=cc, end_row=rr, end_column=cc + 1)
        c2 = ks.cell(row=rr + 1, column=cc, value=value)
        c2.font = big; c2.alignment = center
        ks.merge_cells(start_row=rr + 1, start_column=cc, end_row=rr + 1, end_column=cc + 1)
        ks.row_dimensions[rr + 1].height = 34

    # Action breakdown table
    ar = row0 + 7
    ks.cell(row=ar, column=1, value="Action breakdown (future dates)").font = Font(bold=True, size=12)
    hdr = ["Action", "Count"]
    for j, h in enumerate(hdr, start=1):
        c = ks.cell(row=ar + 1, column=j, value=h)
        c.font = white; c.alignment = center
        c.fill = PatternFill("solid", fgColor=NAVY)
    breakdown = [("🔴 Raise", n_raise, "F8D7DA"),
                 ("🟢 Stimulate", n_stimulate, "D1E7DD"),
                 ("🟠 At limit", n_limit, "FFF3CD"),
                 ("⚪ Hold", n_hold, "E9ECEF")]
    for k, (name, cnt, fill) in enumerate(breakdown, start=1):
        c1 = ks.cell(row=ar + 1 + k, column=1, value=name)
        c1.alignment = left
        c1.fill = PatternFill("solid", fgColor=fill)
        c2 = ks.cell(row=ar + 1 + k, column=2, value=cnt)
        c2.alignment = center
        c2.fill = PatternFill("solid", fgColor=fill)

    # Top dates needing attention (Raise or Stimulate), soonest first
    tr_ = ar
    flagged = fut[act.str.contains("Raise|Stimulate", regex=True)].copy()
    if "Days Left" in flagged.columns:
        flagged = flagged.sort_values("Days Left")
    tcol = 4
    ks.cell(row=tr_, column=tcol, value="Dates needing action (soonest first)").font = Font(bold=True, size=12)
    heads = ["Date", "DOW", "Action", "Rooms Left", "BAR", "Comp"]
    for j, h in enumerate(heads):
        c = ks.cell(row=tr_ + 1, column=tcol + j, value=h)
        c.font = white; c.alignment = center
        c.fill = PatternFill("solid", fgColor=NAVY)
    show = flagged.head(12)
    for k in range(len(show)):
        rowd = show.iloc[k]
        vals = [str(rowd.get("Date", "")), rowd.get("DOW", ""), rowd.get("⚡ Action", ""),
                rowd.get("Rooms Left to Sell", ""), rowd.get("BAR", ""), rowd.get("Comp Set Avg", "")]
        for j, v in enumerate(vals):
            c = ks.cell(row=tr_ + 2 + k, column=tcol + j,
                        value=(None if (isinstance(v, float) and pd.isna(v)) else v))
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(size=9)

    # widths
    for cc in "ABCDEF":
        ks.column_dimensions[cc].width = 16
    ks.column_dimensions["D"].width = 13
    ks.sheet_view.showGridLines = False
    return ks


st.title("The Compton — Daily Pick-Up Report Builder")
st.caption("Upload your four SynXis / rate-shop exports and get the full pick-up "
           "report instantly. No copy/paste required.")

with st.sidebar:
    st.header("1 - Upload your files")
    ups = st.file_uploader(
        "Drop in PCDC, Data Extract, Market Seg and the Rate-Shop file "
        "(all four at once is fine)",
        type=["xlsx"], accept_multiple_files=True,
    )

if not ups:
    st.info("Upload your files in the sidebar to build the report.")
    st.stop()

parts = detect_and_load(ups)

# 'Days Left' reference date, pulled from the PCDC report (Analysis Start Date /
# Generated On).  Falls back to today's date if the PCDC criteria isn't found.
pcdc_asof = parts.get("pcdc_asof") or pd.Timestamp.today().date()

with st.sidebar:
    st.header("2 - Settings")
    as_of = st.date_input(
        "Report as-of date (drives 'Days Left')",
        value=pcdc_asof,
        help="Auto-filled from the PCDC report's Analysis Start Date. "
             "Override here if you want a different reference date.",
    )
    comp_method = st.radio("Comp Set Avg method", ["Average", "Median"], horizontal=True)
    show_change = st.radio(
        "Show competitor rate change vs.",
        ["None", "Yesterday", "3 days ago", "7 days ago"],
        help="Adds a 'Change' column beside each competitor, pulled from the "
             "matching 'vs.' tab in the rate-shop file.",
    )

    st.markdown("---")
    st.header("3 - Signal sensitivity")
    st.caption("Tune when each ⚡ Action fires to match Bentonville's patterns.")
    with st.expander("🔴 Compression (Raise)", expanded=False):
        comp_occ = st.slider("Occupancy % that means compression", 70, 100, 90, 1,
                             help="At or above this forecast occupancy, flag Raise.") / 100
        comp_rooms = st.slider("…or rooms left ≤ this % of capacity", 2, 25, 10, 1,
                               help="Few rooms left also triggers compression.") / 100
        comp_dem = st.slider("…with at least this much remaining demand (rooms)", 0, 30, 5, 1)
    with st.expander("🟢 Soft date (Stimulate)", expanded=False):
        soft_window = st.slider("Only flag dates within this many days out", 3, 60, 21, 1)
        soft_occ = st.slider("…and forecast occupancy below this %", 30, 90, 60, 5,
                             help="Near-term + behind pace + under this occupancy = Stimulate.") / 100

    thr = {"comp_occ": comp_occ, "comp_rooms": comp_rooms, "comp_dem": comp_dem,
           "soft_window": soft_window, "soft_occ": soft_occ}

    st.markdown("---")
    st.caption("Tip: leave the file names as SynXis exports them — the app "
               "auto-detects each report by its sheets.")

c1, c2, c3, c4 = st.columns(4)
c1.metric("PCDC",         "OK" if parts["pcdc"] is not None else "-")
c2.metric("Data Extract", "OK" if parts["data"] is not None else "-")
c3.metric("Market Seg",   "OK" if parts["mseg"] is not None else "-")
c4.metric("Rate Shop",    "OK" if parts["shop"] is not None else "-")

if parts.get("pcdc_asof"):
    st.caption(f"📅 'Days Left' counts from the PCDC report date: "
               f"**{pd.Timestamp(pcdc_asof):%b %d, %Y}** (Analysis Start Date). "
               f"You can change this in the sidebar.")

report = build_report(parts, as_of, comp_method, show_change, thr)

# =============================================================== #
#  KPI strip — portfolio health at a glance
# =============================================================== #
_fut = report[report["Days Left"] >= 0].copy()
def _s(col):
    return pd.to_numeric(_fut.get(col), errors="coerce") if col in _fut else pd.Series(dtype=float)

otb_ty   = _s("Rooms Sold | Total Hotel | Current").sum()
otb_stly = _s("Rooms OTB STLY | Total Hotel | Current").sum()
pace_abs = otb_ty - otb_stly
pace_pct = (pace_abs / otb_stly * 100) if otb_stly else np.nan

n7  = _fut[_fut["Days Left"] <= 7]
n30 = _fut[_fut["Days Left"] <= 30]
occ7  = pd.to_numeric(n7.get("Occ Forecast % | Total Hotel"), errors="coerce").mean()
occ30 = pd.to_numeric(n30.get("Occ Forecast % | Total Hotel"), errors="coerce").mean()

act = _fut.get("⚡ Action", pd.Series(dtype=str)).fillna("")
n_raise     = int(act.str.contains("Raise").sum())
n_stimulate = int(act.str.contains("Stimulate").sum())

bar_ty  = pd.to_numeric(_fut.get("BAR"), errors="coerce")
comp_ty = pd.to_numeric(_fut.get("Comp Set Avg"), errors="coerce")
rate_gap = (bar_ty - comp_ty).mean()

st.subheader("Portfolio snapshot")
k = st.columns(6)
k[0].metric("OTB vs STLY", f"{otb_ty:,.0f}",
            f"{pace_abs:+,.0f} ({pace_pct:+.0f}%)" if pd.notna(pace_pct) else None)
k[1].metric("Next 7 Occ %",  f"{occ7*100:,.0f}%" if pd.notna(occ7) else "–")
k[2].metric("Next 30 Occ %", f"{occ30*100:,.0f}%" if pd.notna(occ30) else "–")
k[3].metric("🔴 Raise dates", f"{n_raise}")
k[4].metric("🟢 Stimulate dates", f"{n_stimulate}")
k[5].metric("Rate vs comp", f"{rate_gap:+,.0f}" if pd.notna(rate_gap) else "–")

# =============================================================== #
#  Filters
# =============================================================== #
st.subheader("Filters")
min_d, max_d = report["Date"].min(), report["Date"].max()

fa, fb, fc = st.columns([1.2, 1, 1])
window = fa.selectbox("Booking window",
                      ["Custom", "Next 7 days", "Next 14 days", "Next 30 days",
                       "Next 60 days", "Next 90 days", "All future dates"],
                      index=6)
dow_pick = fb.multiselect("Day of week", ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"],
                          default=["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"])
action_pick = fc.multiselect("Action", ["🔴 Raise", "🟢 Stimulate", "🟠 At limit", "⚪ Hold"],
                             default=["🔴 Raise", "🟢 Stimulate", "🟠 At limit", "⚪ Hold"])

fd, fe, ff = st.columns([1, 1, 1])
events_only = fd.checkbox("Events only", value=False)
occ_band = fe.select_slider("Occ Forecast % band",
                            options=[0, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100],
                            value=(0, 100))
var_thresh = ff.number_input("Min |variance vs STLY| (rooms)", min_value=0, value=0, step=1)

# custom date range (used when window == Custom)
if window == "Custom":
    d1, d2 = st.columns(2)
    default_start = max(min_d, as_of)
    start = d1.date_input("From", value=default_start, min_value=min_d, max_value=max_d)
    end   = d2.date_input("To",   value=max_d,         min_value=min_d, max_value=max_d)

# ---- apply filters ---- #
v = report.copy()
win_map = {"Next 7 days": 7, "Next 14 days": 14, "Next 30 days": 30,
           "Next 60 days": 60, "Next 90 days": 90}
if window in win_map:
    v = v[(v["Days Left"] >= 0) & (v["Days Left"] <= win_map[window])]
elif window == "All future dates":
    v = v[v["Days Left"] >= 0]
else:  # Custom
    v = v[(v["Date"] >= start) & (v["Date"] <= end)]

dow_full = {"Mon": "Monday", "Tue": "Tuesday", "Wed": "Wednesday", "Thu": "Thursday",
            "Fri": "Friday", "Sat": "Saturday", "Sun": "Sunday"}
keep_dow = [dow_full[d] for d in dow_pick]
v = v[v["DOW"].isin(keep_dow)]

if events_only:
    v = v[v["Events"].astype(str).str.strip() != ""]

if "⚡ Action" in v:
    a = v["⚡ Action"].fillna("")
    sel = pd.Series(False, index=v.index)
    if "🔴 Raise" in action_pick:     sel |= a.str.contains("Raise")
    if "🟢 Stimulate" in action_pick: sel |= a.str.contains("Stimulate")
    if "🟠 At limit" in action_pick:  sel |= a.str.contains("At ")
    if "⚪ Hold" in action_pick:      sel |= a.str.contains("Hold")
    sel |= (a.str.strip() == "")      # keep past/blank rows
    v = v[sel]

ofp_col = pd.to_numeric(v.get("Occ Forecast % | Total Hotel"), errors="coerce") * 100
lo, hi = occ_band
v = v[(ofp_col.isna()) | ((ofp_col >= lo) & (ofp_col <= hi))]

if var_thresh > 0 and "Rooms OTB STLY | Total Hotel | Change" in v:
    var_col = pd.to_numeric(v["Rooms OTB STLY | Total Hotel | Change"], errors="coerce").abs()
    v = v[var_col >= var_thresh]

view = v.reset_index(drop=True)

st.subheader(f"Pick-Up Report — {len(view)} dates")

pct_cols   = [c for c in view.columns if c.startswith("Occ Forecast %")]
shop_all   = [c for c in view.columns if c.startswith("Competitor Shops")]
shop_chg   = [c for c in shop_all if c.endswith("Change")]
shop_cur   = [c for c in shop_all if c not in shop_chg]
money_cols = (["BAR", "Comp Set Avg", "Last Room Value", "Estimated ADR"]
              + [c for c in view.columns if c.startswith("Booked ADR")])
rooms_chg  = [c for c in view.columns
              if (c.startswith("Rooms Sold") or c.startswith("Rooms OTB STLY"))
              and c.endswith("Change")]
var_cols   = [c for c in view.columns if "Variance" in c] + shop_chg + rooms_chg
otb_sold   = "Rooms Left to Sell"      # heat-map column

def _money(v):
    if isinstance(v, (int, float)) and not pd.isna(v):
        return f"${v:,.0f}"
    return "" if v is None or (isinstance(v, float) and pd.isna(v)) else str(v)

def _signed(v):
    if isinstance(v, (int, float)) and not pd.isna(v) and v != 0:
        return f"{v:+,.0f}"
    return ""

fmt = {}
for c in pct_cols:
    fmt[c] = "{:.1%}"
for c in money_cols:
    fmt[c] = "${:,.0f}"
for c in shop_cur:
    fmt[c] = _money
for c in shop_chg:
    fmt[c] = _signed
for c in view.columns:
    if c in ("Days Left", "OOO", "Ovrbk") or c.startswith(("Rooms Sold", "Rooms OTB STLY",
             "Remaining Demand", "Occ Forecast |")) and c not in pct_cols:
        fmt.setdefault(c, "{:,.0f}")

styler = (view.style
          .format(fmt, na_rep="")
          .map(lambda v: "color:#1a7f37;font-weight:600" if isinstance(v, (int, float)) and v > 0
               else ("color:#cf222e;font-weight:600" if isinstance(v, (int, float)) and v < 0 else ""),
               subset=var_cols))


# Heat map on the OTB rooms sold column — pure-Python red->yellow->green
# gradient (no matplotlib needed, so it works on Streamlit Cloud out of the box).
def _heat_bg(series):
    # For "Rooms Left to Sell": GREEN = lots of rooms left (wide open),
    # RED = few rooms left (near sell-out).  Low value -> red, high value -> green.
    vals = pd.to_numeric(series, errors="coerce")
    lo, hi = vals.min(), vals.max()
    span = (hi - lo) or 1
    out = []
    for v in vals:
        if pd.isna(v):
            out.append("")
            continue
        t = (v - lo) / span                    # low rooms left -> red, high -> green
        if t < 0.5:
            f = t / 0.5
            red, grn, blu = 248, int(105 + f * (235 - 105)), int(107 + f * (132 - 107))
        else:
            f = (t - 0.5) / 0.5
            red, grn, blu = int(255 - f * (255 - 99)), int(235 - f * (235 - 190)), int(132 - f * (132 - 123))
        out.append(f"background-color: #{red:02X}{grn:02X}{blu:02X}")
    return out

# Heat map is applied directly inside the pick-up report on the OTB column
if otb_sold in view.columns and view[otb_sold].notna().any():
    styler = styler.apply(_heat_bg, subset=[otb_sold])

# Colour the ⚡ Action column so recommendations pop
def _action_bg(series):
    out = []
    for v in series.astype(str):
        if "Raise" in v:
            out.append("background-color:#F8D7DA;color:#842029;font-weight:600")
        elif "Stimulate" in v:
            out.append("background-color:#D1E7DD;color:#0F5132;font-weight:600")
        elif "At " in v:
            out.append("background-color:#FFF3CD;color:#664D03;font-weight:600")
        elif "Hold" in v:
            out.append("background-color:#E9ECEF;color:#495057")
        else:
            out.append("")
    return out

if "⚡ Action" in view.columns:
    styler = styler.apply(_action_bg, subset=["⚡ Action"])

st.dataframe(styler, use_container_width=True, height=560)

st.download_button(
    "Download report as Excel",
    data=to_excel(view),
    file_name=f"Compton_PickUp_{start:%Y%m%d}_{end:%Y%m%d}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

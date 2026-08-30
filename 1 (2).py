"""
The Compton — Daily Pick-Up Report Builder
Run it with:   streamlit run app.py
"""

import io
import numpy as np
import pandas as pd
import streamlit as st

st.set_page_config(page_title="Compton Daily Pick-Up Report", layout="wide")


def _num(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return np.nan


def _read_all_sheets(uploaded_file):
    xls = pd.ExcelFile(uploaded_file, engine="openpyxl")
    return {s: xls.parse(s, header=None) for s in xls.sheet_names}


def detect_and_load(uploaded_files):
    found = {"pcdc": None, "data": None, "mseg": None, "shop": None,
             "shop_chg_1": None, "shop_chg_3": None, "shop_chg_7": None}

    for f in uploaded_files:
        try:
            sheets = _read_all_sheets(f)
        except Exception as e:
            st.warning(f"Could not open **{f.name}** ({e}).")
            continue
        names = [s.lower() for s in sheets.keys()]

        if any("rate" in n for n in names) and any("overview" in n for n in names):
            raw = sheets[[s for s in sheets if s.lower() == "rates"][0]]
            found["shop"] = _parse_rateshop(raw)
            for tab, key in [("vs. yesterday", "shop_chg_1"),
                             ("vs. 3 days ago", "shop_chg_3"),
                             ("vs. 7 days ago", "shop_chg_7")]:
                match = [s for s in sheets if s.lower() == tab]
                if match:
                    found[key] = _parse_shop_change(sheets[match[0]])

        elif any("changereport" in n for n in names):
            raw = sheets[[s for s in sheets if "changereport" in s.lower()][0]]
            found["pcdc"] = _parse_pcdc(raw)

        elif any("market segment" in n for n in names):
            raw = sheets[[s for s in sheets if s.lower() == "market segment"][0]]
            found["mseg"] = _parse_mseg(raw)

        elif any(n == "property" for n in names):
            raw = sheets[[s for s in sheets if s.lower() == "property"][0]]
            found["data"] = _parse_data(raw)

        else:
            st.warning(f"Didn't recognise **{f.name}** — skipping. "
                       f"(sheets: {', '.join(sheets.keys())})")
    return found


def _parse_data(raw):
    df = raw.copy()
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)
    df["Occupancy Date"] = pd.to_datetime(df["Occupancy Date"], errors="coerce")
    df = df.dropna(subset=["Occupancy Date"])
    return df


def _parse_pcdc(raw):
    df = raw.iloc[3:].copy()
    df[0] = pd.to_datetime(df[0], errors="coerce")
    df = df.dropna(subset=[0]).reset_index(drop=True)
    out = pd.DataFrame({"Occupancy Date": df[0]})
    out["pc_event"]      = df[2]
    out["trans_cur"]     = pd.to_numeric(df[3], errors="coerce")
    out["trans_chg"]     = pd.to_numeric(df[4], errors="coerce")
    out["grp_cur"]       = pd.to_numeric(df[5], errors="coerce")
    out["grp_chg"]       = pd.to_numeric(df[6], errors="coerce")
    out["grp_blocked"]   = pd.to_numeric(df[7], errors="coerce")
    out["grp_avail"]     = pd.to_numeric(df[8], errors="coerce")
    out["grp_pickup"]    = pd.to_numeric(df[9], errors="coerce")
    out["fcst_trans"]    = pd.to_numeric(df[10], errors="coerce")
    out["fcst_grp"]      = pd.to_numeric(df[12], errors="coerce")
    out["rev_trans"]     = pd.to_numeric(df[14], errors="coerce")
    out["rev_grp"]       = pd.to_numeric(df[16], errors="coerce")
    out["adr_trans"]     = pd.to_numeric(df[22], errors="coerce")
    out["adr_grp"]       = pd.to_numeric(df[24], errors="coerce")
    return out


def _parse_mseg(raw):
    df = raw.copy()
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)
    df["Occupancy Date"] = pd.to_datetime(df["Occupancy Date"], errors="coerce")
    df = df.dropna(subset=["Occupancy Date"])
    df["Occupancy On Books This Year"] = pd.to_numeric(
        df["Occupancy On Books This Year"], errors="coerce").fillna(0)
    df["is_trans"] = df["Market Segment"].astype(str).str.startswith("Transient")
    agg = df.groupby("Occupancy Date").apply(
        lambda g: pd.Series({
            "ms_trans": g.loc[g["is_trans"], "Occupancy On Books This Year"].sum(),
            "ms_group": g.loc[~g["is_trans"], "Occupancy On Books This Year"].sum(),
        })
    ).reset_index()
    return agg


def _clean_shop(x):
    if x is None:
        return None
    v = _num(x)
    if not np.isnan(v):
        return int(v) if float(v).is_integer() else v
    s = str(x).strip()
    return s if s else None


def _parse_rateshop(raw):
    df = raw.iloc[5:].copy()
    df[2] = pd.to_datetime(df[2], errors="coerce")
    df = df.dropna(subset=[2]).reset_index(drop=True)
    out = pd.DataFrame({"Occupancy Date": df[2]})
    out["own_bar"] = df[4].map(_num)
    for key, col in [("c_21c", 5), ("c_motto", 6), ("c_ac", 7), ("c_dt", 8)]:
        out[key] = df[col].map(_clean_shop)
        out[key + "_n"] = df[col].map(_num)
    return out


def _parse_shop_change(raw):
    df = raw.iloc[5:].copy()
    df[2] = pd.to_datetime(df[2], errors="coerce")
    df = df.dropna(subset=[2]).reset_index(drop=True)
    out = pd.DataFrame({"Occupancy Date": df[2]})
    for key, chg_col in [("c_21c", 8), ("c_motto", 10), ("c_ac", 12), ("c_dt", 14)]:
        out[key + "_chg"] = df[chg_col].map(_num) if chg_col in df.columns else np.nan
    return out


def build_report(parts, as_of, comp_method="Average", show_change="None"):
    data, pcdc, mseg, shop = parts["data"], parts["pcdc"], parts["mseg"], parts["shop"]
    if data is None:
        st.error("The **Data Extract** file is required (couldn't find a 'Property' sheet).")
        st.stop()

    df = data.copy()
    for extra in (pcdc, mseg, shop):
        if extra is not None:
            df = df.merge(extra, on="Occupancy Date", how="left")

    chg_key = {"Yesterday": "shop_chg_1", "3 days ago": "shop_chg_3",
               "7 days ago": "shop_chg_7"}.get(show_change)
    chg_df = parts.get(chg_key) if chg_key else None
    if chg_df is not None:
        df = df.merge(chg_df, on="Occupancy Date", how="left")

    def dnum(col):
        return pd.to_numeric(df.get(col), errors="coerce")

    cap        = dnum("Physical Capacity This Year")
    occ_ty     = dnum("Occupancy On Books This Year")
    occ_stly   = dnum("Occupancy On Books STLY")
    tr_ty      = dnum("Rooms Sold - Transient This Year")
    tr_stly    = dnum("Rooms Sold - Transient STLY")
    gr_ty      = dnum("Rooms Sold - Group This Year")
    gr_stly    = dnum("Rooms Sold - Group STLY")
    dem_tot    = dnum("User Total Demand - Total This Year")
    dem_grp    = dnum("User Constrained Total Demand - Group This Year")
    dem_tr     = dnum("User Unconstrained Total Demand - Transient This Year")

    r = pd.DataFrame()
    r["DOW"]  = df["Day of Week"]
    r["Date"] = df["Occupancy Date"].dt.date
    r["Days Left"] = (df["Occupancy Date"] - pd.Timestamp(as_of)).dt.days

    ev = df.get("Special Event This Year")
    ev = ev.fillna("") if ev is not None else ""
    if "pc_event" in df:
        ev = ev.where(ev.astype(str).str.strip() != "", df["pc_event"].fillna(""))
    r["Events"] = pd.Series(ev, index=df.index).replace("nan", "")

    r["Rooms Left to Sell"] = dnum("Remaining Capacity This Year")
    r["BAR"]                = dnum("BAR")

    if "c_21c_n" in df:
        comps_n = df[["c_21c_n", "c_motto_n", "c_ac_n", "c_dt_n"]]
        r["Comp Set Avg"] = (comps_n.mean(axis=1) if comp_method == "Average"
                             else comps_n.median(axis=1)).round(0)
    else:
        r["Comp Set Avg"] = np.nan

    r["Last Room Value"] = dnum("Last Room Value This Year")

    add_chg = chg_df is not None
    for name, key in [("21c Museum Hotel Bentonville - MGallery", "c_21c"),
                      ("Motto By Hilton Bentonville Downtown",    "c_motto"),
                      ("AC Hotel by Marriott Bentonville",        "c_ac"),
                      ("DoubleTree Suites by Hilton Bentonville", "c_dt")]:
        base = "Competitor Shops | " + name
        if add_chg:
            r[base + " | Current"] = df.get(key)
            r[base + " | Change"]  = df.get(key + "_chg")
        else:
            r[base] = df.get(key)

    r["OOO"]   = dnum("Rooms N/A - Out of Order This Year")
    r["Ovrbk"] = dnum("Overbooking This Year")

    tc, tch = df.get("trans_cur"), df.get("trans_chg")
    gc, gch = df.get("grp_cur"),   df.get("grp_chg")
    r["Rooms Sold | Total Hotel | Current"]   = (tc.fillna(0) + gc.fillna(0)) if tc is not None else occ_ty
    r["Rooms Sold | Total Hotel | Change"]    = (tch.fillna(0) + gch.fillna(0)) if tch is not None else np.nan
    r["Rooms Sold | Total Transient | Current"] = tc
    r["Rooms Sold | Total Transient | Change"]  = tch
    r["Rooms Sold | Total Group | Current"]     = gc
    r["Rooms Sold | Total Group | Change"]      = gch
    r["Rooms Sold | Total Group | Blocked"]     = df.get("grp_blocked")
    r["Rooms Sold | Total Group | P/U"]         = df.get("grp_pickup")
    r["Rooms Sold | Total Group | Remaining"]   = df.get("grp_avail")

    r["Rooms OTB STLY | Total Hotel | Current"]     = occ_stly
    r["Rooms OTB STLY | Total Hotel | Change"]      = occ_ty - occ_stly
    r["Rooms OTB STLY | Total Transient | Current"] = tr_stly
    r["Rooms OTB STLY | Total Transient | Change"]  = tr_ty - tr_stly
    r["Rooms OTB STLY | Total Group | Current"]     = gr_stly
    r["Rooms OTB STLY | Total Group | Change"]      = gr_ty - gr_stly

    r["Remaining Demand | Total Hotel"]     = (dem_tot - occ_ty).clip(lower=0)
    r["Remaining Demand | Total Transient"] = (dem_tr - tr_ty).clip(lower=0)
    r["Remaining Demand | Total Group"]     = (dem_grp - gr_ty).clip(lower=0)

    ft, fg = df.get("fcst_trans"), df.get("fcst_grp")
    fcst_tot = (ft.fillna(0) + fg.fillna(0)) if ft is not None else dnum("Forecasted Room Revenue This Year")*np.nan
    r["Occ Forecast | Total Hotel"]     = fcst_tot.round(1) if ft is not None else np.nan
    r["Occ Forecast | Total Transient"] = ft.round(1) if ft is not None else np.nan
    r["Occ Forecast | Total Group"]     = fg.round(1) if fg is not None else np.nan

    r["Occ Forecast % | Total Hotel"]     = (fcst_tot / cap) if ft is not None else np.nan
    r["Occ Forecast % | Total Transient"] = (ft / cap) if ft is not None else np.nan
    r["Occ Forecast % | Total Group"]     = (fg / cap) if fg is not None else np.nan

    rev_t, rev_g = df.get("rev_trans"), df.get("rev_grp")
    if tc is not None:
        total_rooms = (tc.fillna(0) + gc.fillna(0)).replace(0, np.nan)
        r["Booked ADR | Total Hotel"] = ((rev_t.fillna(0) + rev_g.fillna(0)) / total_rooms).round(2)
    else:
        r["Booked ADR | Total Hotel"] = dnum("ADR On Books This Year")
    r["Booked ADR | Total Transient"] = df.get("adr_trans")
    r["Booked ADR | Total Group"]     = df.get("adr_grp")

    r["Estimated ADR"] = dnum("ADR Forecast This Year")

    return r


def _levels(col):
    parts = [p.strip() for p in str(col).split("|")]
    while len(parts) < 3:
        parts.append("")
    return tuple(parts[:3])


def to_excel(r):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    lv = [_levels(c) for c in r.columns]
    n = len(lv)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pick-Up Report"

    NAVY, TEAL, ORANGE, PURPLE, GREEN = "16365C", "215967", "E26B0A", "60497A", "76933C"
    white = Font(color="FFFFFF", bold=True, size=9)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

    def PF(hexcol):
        return PatternFill("solid", fgColor=hexcol)

    def seg_color(txt):
        if "Total Hotel" in txt:
            return ORANGE
        if "Total Transient" in txt:
            return PURPLE
        if "Total Group" in txt:
            return GREEN
        return None

    def band_fill(l1, l2):
        return PF(NAVY) if l2 == "" else PF(TEAL)

    def lvl2_fill(l1, l2):
        sc = seg_color(l2)
        if sc:
            return PF(sc)
        if l1 == "Competitor Shops":
            return PF(NAVY)
        return PF(TEAL)

    def lvl3_fill(l1, l2):
        sc = seg_color(l2)
        return PF(sc) if sc else PF(NAVY)

    j = 0
    while j < n:
        l1 = lv[j][0]
        k = j
        while k + 1 < n and lv[k + 1][0] == l1 and lv[j][1]:
            k += 1
        fill = band_fill(l1, lv[j][1])
        if lv[j][1] == "" and l1 in ("Estimated ADR",):
            fill = PF(ORANGE)
        cell = ws.cell(row=1, column=j + 1, value=l1)
        cell.font = white
        cell.alignment = center
        cell.fill = fill
        if lv[j][1] == "":
            ws.merge_cells(start_row=1, start_column=j + 1, end_row=3, end_column=j + 1)
        elif k > j:
            ws.merge_cells(start_row=1, start_column=j + 1, end_row=1, end_column=k + 1)
        j = k + 1

    j = 0
    while j < n:
        l1, l2, l3 = lv[j]
        if l2:
            k = j
            while k + 1 < n and lv[k + 1][0] == l1 and lv[k + 1][1] == l2:
                k += 1
            cell = ws.cell(row=2, column=j + 1, value=l2)
            cell.font = white
            cell.alignment = center
            cell.fill = lvl2_fill(l1, l2)
            if l3 == "":
                ws.merge_cells(start_row=2, start_column=j + 1, end_row=3, end_column=j + 1)
            elif k > j:
                ws.merge_cells(start_row=2, start_column=j + 1, end_row=2, end_column=k + 1)
            j = k + 1
        else:
            j += 1

    for j, (l1, l2, l3) in enumerate(lv, start=1):
        if l3:
            cell = ws.cell(row=3, column=j, value=l3)
            cell.font = white
            cell.alignment = center
            cell.fill = lvl3_fill(l1, l2)

    _hotelnames = ("Bentonville", "MGallery")
    DATA0 = 4
    pct = {i for i, c0 in enumerate(r.columns) if c0.startswith("Occ Forecast %")}
    chg = {i for i, c0 in enumerate(r.columns)
           if any(h in c0 for h in _hotelnames) and c0.endswith("Change")}
    money = {i for i, c0 in enumerate(r.columns)
             if (c0 in ("BAR", "Comp Set Avg", "Last Room Value", "Estimated ADR")
                 or "ADR" in c0 or any(h in c0 for h in _hotelnames))
             and i not in chg}
    otb_sold_col = None
    for i, c0 in enumerate(r.columns):
        if c0 == "Rooms Sold | Total Hotel | Current":
            otb_sold_col = i

    for ri, (_, row) in enumerate(r.iterrows(), start=DATA0):
        for ci2, col in enumerate(r.columns, start=1):
            v = row[col]
            cell = ws.cell(row=ri, column=ci2, value=(None if pd.isna(v) else v))
            cell.border = thin
            cell.font = Font(size=9)
            idx = ci2 - 1
            if idx in pct:
                cell.number_format = "0.0%"
            elif idx in chg:
                cell.number_format = "+#,##0;-#,##0;\"\""
            elif idx in money:
                cell.number_format = "$#,##0"
            elif isinstance(v, float) and float(v).is_integer():
                cell.number_format = "#,##0"

    if otb_sold_col is not None and len(r) > 0:
        L = get_column_letter(otb_sold_col + 1)
        rng = f"{L}{DATA0}:{L}{DATA0 + len(r) - 1}"
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(start_type="min", start_color="F8696B",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="max", end_color="63BE7B"))

    ws.freeze_panes = "C4"
    for j in range(1, n + 1):
        ws.column_dimensions[get_column_letter(j)].width = 12
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["D"].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


st.title("The Compton — Daily Pick-Up Report Builder")
st.caption("Upload your four SynXis / rate-shop exports and get the full pick-up "
           "report instantly. No copy/paste required.")

with st.sidebar:
    st.header("1 - Upload your files")
    ups = st.file_uploader(
        "Drop in PCDC, Data Extract, Market Seg and the Rate-Shop file "
        "(all four at once is fine)",
        type=["xlsx"], accept_multiple_files=True,
    )
    st.header("2 - Settings")
    as_of = st.date_input("Report as-of date (drives 'Days Left')",
                          value=pd.Timestamp.today().date())
    comp_method = st.radio("Comp Set Avg method", ["Average", "Median"], horizontal=True)
    show_change = st.radio(
        "Show competitor rate change vs.",
        ["None", "Yesterday", "3 days ago", "7 days ago"],
        help="Adds a 'Change' column beside each competitor, pulled from the "
             "matching 'vs.' tab in the rate-shop file.",
    )
    st.markdown("---")
    st.caption("Tip: leave the file names as SynXis exports them — the app "
               "auto-detects each report by its sheets.")

if not ups:
    st.info("Upload your files in the sidebar to build the report.")
    st.stop()

parts = detect_and_load(ups)

c1, c2, c3, c4 = st.columns(4)
c1.metric("PCDC",         "OK" if parts["pcdc"] is not None else "-")
c2.metric("Data Extract", "OK" if parts["data"] is not None else "-")
c3.metric("Market Seg",   "OK" if parts["mseg"] is not None else "-")
c4.metric("Rate Shop",    "OK" if parts["shop"] is not None else "-")

report = build_report(parts, as_of, comp_method, show_change)

st.subheader("Filter dates")
min_d, max_d = report["Date"].min(), report["Date"].max()
default_start = max(min_d, as_of)
d1, d2 = st.columns(2)
start = d1.date_input("From", value=default_start, min_value=min_d, max_value=max_d)
end   = d2.date_input("To",   value=max_d,         min_value=min_d, max_value=max_d)
mask = (report["Date"] >= start) & (report["Date"] <= end)
view = report.loc[mask].reset_index(drop=True)

st.subheader(f"Pick-Up Report — {start:%b %d, %Y} to {end:%b %d, %Y}  ({len(view)} days)")

pct_cols   = [c for c in view.columns if c.startswith("Occ Forecast %")]
shop_all   = [c for c in view.columns if c.startswith("Competitor Shops")]
shop_chg   = [c for c in shop_all if c.endswith("Change")]
shop_cur   = [c for c in shop_all if c not in shop_chg]
money_cols = (["BAR", "Comp Set Avg", "Last Room Value", "Estimated ADR"]
              + [c for c in view.columns if c.startswith("Booked ADR")])
rooms_chg  = [c for c in view.columns
              if (c.startswith("Rooms Sold") or c.startswith("Rooms OTB STLY"))
              and c.endswith("Change")]
var_cols   = [c for c in view.columns if "Variance" in c] + shop_chg + rooms_chg
otb_sold   = "Rooms Sold | Total Hotel | Current"

def _money(v):
    if isinstance(v, (int, float)) and not pd.isna(v):
        return f"${v:,.0f}"
    return "" if v is None or (isinstance(v, float) and pd.isna(v)) else str(v)

def _signed(v):
    if isinstance(v, (int, float)) and not pd.isna(v) and v != 0:
        return f"{v:+,.0f}"
    return ""

fmt = {}
for c in pct_cols:
    fmt[c] = "{:.1%}"
for c in money_cols:
    fmt[c] = "${:,.0f}"
for c in shop_cur:
    fmt[c] = _money
for c in shop_chg:
    fmt[c] = _signed
for c in view.columns:
    if c in ("Days Left", "OOO", "Ovrbk") or c.startswith(("Rooms Sold", "Rooms OTB STLY",
             "Remaining Demand", "Occ Forecast |")) and c not in pct_cols:
        fmt.setdefault(c, "{:,.0f}")

styler = (view.style
          .format(fmt, na_rep="")
          .map(lambda v: "color:#1a7f37;font-weight:600" if isinstance(v, (int, float)) and v > 0
               else ("color:#cf222e;font-weight:600" if isinstance(v, (int, float)) and v < 0 else ""),
               subset=var_cols))


# Heat map on the OTB rooms sold column — pure-Python red->yellow->green
# gradient (no matplotlib needed, so it works on Streamlit Cloud out of the box).
def _heat_bg(series):
    vals = pd.to_numeric(series, errors="coerce")
    lo, hi = vals.min(), vals.max()
    span = (hi - lo) or 1
    out = []
    for v in vals:
        if pd.isna(v):
            out.append("")
            continue
        t = (v - lo) / span
        if t < 0.5:
            f = t / 0.5
            red, grn, blu = 248, int(105 + f * (235 - 105)), int(107 + f * (132 - 107))
        else:
            f = (t - 0.5) / 0.5
            red, grn, blu = int(255 - f * (255 - 99)), int(235 - f * (235 - 190)), int(132 - f * (132 - 123))
        out.append(f"background-color: #{red:02X}{grn:02X}{blu:02X}")
    return out

if otb_sold in view.columns and view[otb_sold].notna().any():
    styler = styler.apply(_heat_bg, subset=[otb_sold])

st.dataframe(styler, use_container_width=True, height=560)

if otb_sold in view.columns and view[otb_sold].notna().any():
    st.subheader("OTB Rooms Sold — heat calendar")
    import altair as alt
    hc = view[["Date", "DOW", otb_sold]].copy()
    hc = hc.rename(columns={otb_sold: "Rooms Sold"})
    hc["Date"] = pd.to_datetime(hc["Date"])
    hc["Week of"] = (hc["Date"] - pd.to_timedelta(hc["Date"].dt.weekday, unit="D"))
    hc["Week label"] = hc["Week of"].dt.strftime("%b %d")
    dow_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    chart = (
        alt.Chart(hc)
        .mark_rect(stroke="white", strokeWidth=1)
        .encode(
            x=alt.X("DOW:N", sort=dow_order, title=None,
                    axis=alt.Axis(labelAngle=0, orient="top")),
            y=alt.Y("Week label:N", sort=list(hc.sort_values("Week of")["Week label"].unique()),
                    title="Week of"),
            color=alt.Color("Rooms Sold:Q", scale=alt.Scale(scheme="redyellowgreen"),
                            legend=alt.Legend(title="Rooms Sold")),
            tooltip=[alt.Tooltip("Date:T", title="Date"),
                     alt.Tooltip("DOW:N", title="Day"),
                     alt.Tooltip("Rooms Sold:Q", title="Rooms Sold")],
        )
        .properties(height=max(120, 26 * hc["Week label"].nunique()))
    )
    text = (
        alt.Chart(hc)
        .mark_text(baseline="middle", fontSize=10)
        .encode(
            x=alt.X("DOW:N", sort=dow_order),
            y=alt.Y("Week label:N",
                    sort=list(hc.sort_values("Week of")["Week label"].unique())),
            text=alt.Text("Rooms Sold:Q", format=".0f"),
            color=alt.condition("datum['Rooms Sold'] > 100",
                                alt.value("white"), alt.value("black")),
        )
    )
    st.altair_chart(chart + text, use_container_width=True)

st.download_button(
    "Download report as Excel",
    data=to_excel(view),
    file_name=f"Compton_PickUp_{start:%Y%m%d}_{end:%Y%m%d}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

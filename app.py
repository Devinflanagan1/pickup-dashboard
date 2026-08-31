"""
The Compton — Daily Pick-Up Report Builder
Run it with:   streamlit run app.py

Two views (switch with the "View" radio at the top of the sidebar):
  1. Pick-Up Report (DD) — the original daily pick-up report (unchanged).
  2. Dashboard          — reproduces the workbook 'Dashboard' tab: Total /
                          Transient / Group on-the-books by month vs Last
                          Report, Budget and Freeze Forecast.
"""

import io
import calendar
import numpy as np
import pandas as pd
import streamlit as st

st.set_page_config(page_title="Compton Daily Pick-Up Report", layout="wide")


def _num(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return np.nan


def _read_all_sheets(uploaded_file):
    xls = pd.ExcelFile(uploaded_file, engine="openpyxl")
    return {s: xls.parse(s, header=None) for s in xls.sheet_names}


def detect_and_load(uploaded_files):
    found = {"pcdc": None, "data": None, "mseg": None, "shop": None,
             "shop_chg_1": None, "shop_chg_3": None, "shop_chg_7": None,
             "pcdc_asof": None, "mseg_raw": None, "mseg_file": None}

    for f in uploaded_files:
        try:
            sheets = _read_all_sheets(f)
        except Exception as e:
            st.warning(f"Could not open **{f.name}** ({e}).")
            continue
        names = [s.lower() for s in sheets.keys()]

        if any("rate" in n for n in names) and any("overview" in n for n in names):
            raw = sheets[[s for s in sheets if s.lower() == "rates"][0]]
            found["shop"] = _parse_rateshop(raw)
            for tab, key in [("vs. yesterday", "shop_chg_1"),
                             ("vs. 3 days ago", "shop_chg_3"),
                             ("vs. 7 days ago", "shop_chg_7")]:
                match = [s for s in sheets if s.lower() == tab]
                if match:
                    found[key] = _parse_shop_change(sheets[match[0]])

        elif any("changereport" in n for n in names):
            raw = sheets[[s for s in sheets if "changereport" in s.lower()][0]]
            found["pcdc"] = _parse_pcdc(raw)
            crit = [s for s in sheets if s.lower() == "report criteria"]
            if crit:
                found["pcdc_asof"] = _parse_pcdc_asof(sheets[crit[0]])

        elif any("market segment" in n for n in names):
            raw = sheets[[s for s in sheets if s.lower() == "market segment"][0]]
            found["mseg"] = _parse_mseg(raw)
            found["mseg_raw"] = raw          # keep raw sheet for the Dashboard page
            found["mseg_file"] = getattr(f, "name", "")  # remember source filename

        elif any(n == "property" for n in names):
            raw = sheets[[s for s in sheets if s.lower() == "property"][0]]
            found["data"] = _parse_data(raw)

        else:
            st.warning(f"Didn't recognise **{f.name}** — skipping. "
                       f"(sheets: {', '.join(sheets.keys())})")
    return found


def _parse_data(raw):
    df = raw.copy()
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)
    df["Occupancy Date"] = pd.to_datetime(df["Occupancy Date"], errors="coerce")
    df = df.dropna(subset=["Occupancy Date"])
    return df


def _parse_pcdc(raw):
    df = raw.iloc[3:].copy()
    df[0] = pd.to_datetime(df[0], errors="coerce")
    df = df.dropna(subset=[0]).reset_index(drop=True)
    out = pd.DataFrame({"Occupancy Date": df[0]})
    out["pc_event"]      = df[2]
    out["trans_cur"]     = pd.to_numeric(df[3], errors="coerce")
    out["trans_chg"]     = pd.to_numeric(df[4], errors="coerce")
    out["grp_cur"]       = pd.to_numeric(df[5], errors="coerce")
    out["grp_chg"]       = pd.to_numeric(df[6], errors="coerce")
    out["grp_blocked"]   = pd.to_numeric(df[7], errors="coerce")
    out["grp_avail"]     = pd.to_numeric(df[8], errors="coerce")
    out["grp_pickup"]    = pd.to_numeric(df[9], errors="coerce")
    out["fcst_trans"]    = pd.to_numeric(df[10], errors="coerce")
    out["fcst_grp"]      = pd.to_numeric(df[12], errors="coerce")
    out["rev_trans"]     = pd.to_numeric(df[14], errors="coerce")
    out["rev_grp"]       = pd.to_numeric(df[16], errors="coerce")
    out["adr_trans"]     = pd.to_numeric(df[22], errors="coerce")
    out["adr_grp"]       = pd.to_numeric(df[24], errors="coerce")
    return out


def _parse_pcdc_asof(raw):
    """
    Read the PCDC 'Report Criteria' sheet and return the reference date for
    'Days Left' (i.e. "today" for pace purposes).

    'Generated On' is the actual timestamp the report was run/exported, so it's
    the most reliable proxy for "today". 'Analysis Start Date' and 'Activity
    Start Date' are report-WINDOW parameters the revenue manager can set to any
    date (e.g. to look at pace from a few days back) — they are NOT guaranteed
    to equal today, so they're only used as a fallback if 'Generated On' is
    missing. Returns a datetime.date or None.
    """
    try:
        # find the header row that contains the labels, then the value row below it
        for i in range(len(raw) - 1):
            rowvals = [str(x).strip().lower() for x in raw.iloc[i].tolist()]
            if "analysis start date" in rowvals or "generated on" in rowvals:
                header = raw.iloc[i].tolist()
                values = raw.iloc[i + 1].tolist()
                lut = {str(h).strip().lower(): v for h, v in zip(header, values)}
                for key in ("generated on", "activity start date", "analysis start date"):
                    if key in lut and lut[key] not in (None, ""):
                        d = pd.to_datetime(lut[key], errors="coerce")
                        if pd.notna(d):
                            return d.date()
    except Exception:
        pass
    return None


def _parse_mseg(raw):
    df = raw.copy()
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)
    df["Occupancy Date"] = pd.to_datetime(df["Occupancy Date"], errors="coerce")
    df = df.dropna(subset=["Occupancy Date"])
    df["Occupancy On Books This Year"] = pd.to_numeric(
        df["Occupancy On Books This Year"], errors="coerce").fillna(0)
    df["is_trans"] = df["Market Segment"].astype(str).str.startswith("Transient")
    agg = df.groupby("Occupancy Date").apply(
        lambda g: pd.Series({
            "ms_trans": g.loc[g["is_trans"], "Occupancy On Books This Year"].sum(),
            "ms_group": g.loc[~g["is_trans"], "Occupancy On Books This Year"].sum(),
        })
    ).reset_index()
    return agg


def _clean_shop(x):
    if x is None:
        return None
    v = _num(x)
    if not np.isnan(v):
        return int(v) if float(v).is_integer() else v
    s = str(x).strip()
    return s if s else None


def _parse_rateshop(raw):
    df = raw.iloc[5:].copy()
    df[2] = pd.to_datetime(df[2], errors="coerce")
    df = df.dropna(subset=[2]).reset_index(drop=True)
    out = pd.DataFrame({"Occupancy Date": df[2]})
    out["own_bar"] = df[4].map(_num)
    for key, col in [("c_21c", 5), ("c_motto", 6), ("c_ac", 7), ("c_dt", 8)]:
        out[key] = df[col].map(_clean_shop)
        out[key + "_n"] = df[col].map(_num)
    return out


def _parse_shop_change(raw):
    df = raw.iloc[5:].copy()
    df[2] = pd.to_datetime(df[2], errors="coerce")
    df = df.dropna(subset=[2]).reset_index(drop=True)
    out = pd.DataFrame({"Occupancy Date": df[2]})
    # our own BAR change sits in column 6 (right of our rate in column 5)
    out["own_chg"] = df[6].map(_num) if 6 in df.columns else np.nan
    for key, chg_col in [("c_21c", 8), ("c_motto", 10), ("c_ac", 12), ("c_dt", 14)]:
        out[key + "_chg"] = df[chg_col].map(_num) if chg_col in df.columns else np.nan
    return out




# Physical length-of-stay restrictions we can't export from IDeaS, so they're
# entered by hand and carried forward via the prior day's Compton export.
RESTRICTION_TYPES = ["MinLOS 2", "MinLOS 3", "MinLOS 4", "CTA", "Closed"]


def _strip_restr_annotation(s):
    """Recover the base restriction from an exported/annotated cell, e.g.
    'MinLOS 2 (Previous MinLOS 3)' -> 'MinLOS 2'; 'Open (Previous MinLOS 3)' -> ''."""
    s = str(s or "").strip()
    if not s or s.lower() == "nan":
        return ""
    base = s.split(" (Previous")[0].strip()
    return "" if base.lower() in ("", "open") else base


def annotate_restriction(cur, prev, has_prev):
    """Format the Restrictions cell, flagging changes vs the last report:
        changed:  'MinLOS 2 (Previous MinLOS 3)'
        lifted:   'Open (Previous MinLOS 3)'
        new:      'MinLOS 3'      (no annotation — nothing there before)
        same:     'MinLOS 3'      (no annotation)
    `has_prev` is False when there's no comparison file, so we just show `cur`."""
    cur = str(cur or "").strip()
    prev = str(prev or "").strip()
    if not has_prev:
        return cur
    if cur and prev and cur != prev:
        return f"{cur} (Previous {prev})"
    if not cur and prev:
        return f"Open (Previous {prev})"
    return cur


def build_report(parts, as_of, comp_method="Average", show_change="None", thr=None,
                 show_own_change=True, restrictions=None, prev_restrictions=None,
                 restr_annotate=False):
    data, pcdc, mseg, shop = parts["data"], parts["pcdc"], parts["mseg"], parts["shop"]
    if data is None:
        st.error("The **Data Extract** file is required (couldn't find a 'Property' sheet).")
        st.stop()

    restrictions = restrictions or {}
    prev_restrictions = prev_restrictions or {}

    # Optional date-level comparison baselines (from a prior day's uploaded file,
    # either a Compton export's Pick-Up Report tab or a Market Segment export):
    #   pu_dates -> "since last report" pickup; wk_dates -> "7-day" pickup.
    # Each is { 'YYYY-MM-DD': [total, rev, trans, group] }.
    pu_dates = (parts.get("cmp_lr") or {}).get("by_date") or {}
    wk_dates = (parts.get("cmp_wk") or {}).get("by_date") or {}

    df = data.copy()
    for extra in (pcdc, mseg, shop):
        if extra is not None:
            df = df.merge(extra, on="Occupancy Date", how="left")

    chg_key = {"Yesterday": "shop_chg_1", "3 days ago": "shop_chg_3",
               "7 days ago": "shop_chg_7"}.get(show_change)
    chg_df = parts.get(chg_key) if chg_key else None
    if chg_df is not None:
        # avoid duplicate own_chg if we also pull the 7-day own change below
        df = df.merge(chg_df.drop(columns=[c for c in ["own_chg"] if c in chg_df], errors="ignore"),
                      on="Occupancy Date", how="left")

    # Always pull OUR OWN 7-day BAR change (independent of the competitor toggle)
    own7 = parts.get("shop_chg_7")
    if own7 is not None and "own_chg" in own7:
        df = df.merge(own7[["Occupancy Date", "own_chg"]].rename(columns={"own_chg": "bar_chg_7d"}),
                      on="Occupancy Date", how="left")

    # Optional 7-day PCDC (a change report run over a 7-day window): its per-date
    # transient/group room change IS the day-by-day 7-day pickup. Merge it in with
    # distinct names so it doesn't collide with the daily PCDC change columns.
    pcdc7 = parts.get("pcdc7")
    has_wk_pu = pcdc7 is not None and "Occupancy Date" in getattr(pcdc7, "columns", [])
    if has_wk_pu:
        w = pcdc7[["Occupancy Date", "trans_chg", "grp_chg"]].rename(
            columns={"trans_chg": "w_trans_chg", "grp_chg": "w_grp_chg"})
        df = df.merge(w, on="Occupancy Date", how="left")

    def dnum(col):
        return pd.to_numeric(df.get(col), errors="coerce")

    cap        = dnum("Physical Capacity This Year")
    occ_ty     = dnum("Occupancy On Books This Year")
    occ_stly   = dnum("Occupancy On Books STLY")
    tr_ty      = dnum("Rooms Sold - Transient This Year")
    tr_stly    = dnum("Rooms Sold - Transient STLY")
    gr_ty      = dnum("Rooms Sold - Group This Year")
    gr_stly    = dnum("Rooms Sold - Group STLY")
    dem_tot    = dnum("User Total Demand - Total This Year")
    dem_grp    = dnum("User Constrained Total Demand - Group This Year")
    dem_tr     = dnum("User Unconstrained Total Demand - Transient This Year")

    r = pd.DataFrame()
    r["DOW"]  = df["Day of Week"]
    r["Date"] = df["Occupancy Date"].dt.date
    r["Days Left"] = (df["Occupancy Date"] - pd.Timestamp(as_of)).dt.days

    ev = df.get("Special Event This Year")
    ev = ev.fillna("") if ev is not None else ""
    if "pc_event" in df:
        ev = ev.where(ev.astype(str).str.strip() != "", df["pc_event"].fillna(""))
    r["Events"] = pd.Series(ev, index=df.index).replace("nan", "")

    r["Rooms Left to Sell"] = dnum("Remaining Capacity This Year")
    r["BAR"]                = dnum("BAR")
    if show_own_change:
        r["BAR Chg 7d"]     = dnum("bar_chg_7d")     # our own rate move over last 7 days

    if "c_21c_n" in df:
        comps_n = df[["c_21c_n", "c_motto_n", "c_ac_n", "c_dt_n"]]
        r["Comp Set Avg"] = (comps_n.mean(axis=1) if comp_method == "Average"
                             else comps_n.median(axis=1)).round(0)
    else:
        r["Comp Set Avg"] = np.nan

    r["Last Room Value"] = dnum("Last Room Value This Year")

    # Manual physical restrictions (entered in the app), placed next to Last Room
    # Value, with a "(Previous …)" note when they changed since the last report.
    _rds = df["Occupancy Date"].dt.strftime("%Y-%m-%d")
    r["Restrictions"] = [
        annotate_restriction(restrictions.get(ds, ""), prev_restrictions.get(ds, ""),
                             restr_annotate)
        for ds in _rds
    ]

    add_chg = chg_df is not None
    for name, key in [("21c Museum Hotel Bentonville - MGallery", "c_21c"),
                      ("Motto By Hilton Bentonville Downtown",    "c_motto"),
                      ("AC Hotel by Marriott Bentonville",        "c_ac"),
                      ("DoubleTree Suites by Hilton Bentonville", "c_dt")]:
        base = "Competitor Shops | " + name
        if add_chg:
            r[base + " | Current"] = df.get(key)
            r[base + " | Change"]  = df.get(key + "_chg")
        else:
            r[base] = df.get(key)

    r["OOO"]   = dnum("Rooms N/A - Out of Order This Year")
    r["Ovrbk"] = dnum("Overbooking This Year")

    # per-date baseline series aligned to df rows, from a prior day's file
    _dstr = df["Occupancy Date"].dt.strftime("%Y-%m-%d")

    def _base(baseline, idx):
        if not baseline:
            return pd.Series(np.nan, index=df.index)
        return _dstr.map(lambda ds: (baseline.get(ds) or [np.nan, np.nan, np.nan, np.nan])[idx])

    tc, tch = df.get("trans_cur"), df.get("trans_chg")
    gc, gch = df.get("grp_cur"),   df.get("grp_chg")

    # CURRENT on-the-books: prefer the PCDC's per-segment current; otherwise fall
    # back to the Data Extract's own transient/group/total OTB.
    cur_tot   = (tc.fillna(0) + gc.fillna(0)) if tc is not None else occ_ty
    cur_trans = tc if tc is not None else tr_ty
    cur_group = gc if gc is not None else gr_ty

    # CHANGE (pickup since last report): prefer the PCDC's own change; otherwise
    # compute it from an uploaded prior-day baseline (current − last report).
    if tch is not None:
        chg_tot, chg_trans, chg_group = (tch.fillna(0) + gch.fillna(0)), tch, gch
    elif pu_dates:
        chg_tot   = cur_tot   - _base(pu_dates, 0)
        chg_trans = cur_trans - _base(pu_dates, 2)
        chg_group = cur_group - _base(pu_dates, 3)
    else:
        chg_tot = chg_trans = chg_group = np.nan

    r["Rooms Sold | Total Hotel | Current"]     = cur_tot
    r["Rooms Sold | Total Hotel | Change"]      = chg_tot
    r["Rooms Sold | Total Transient | Current"] = cur_trans
    r["Rooms Sold | Total Transient | Change"]  = chg_trans
    r["Rooms Sold | Total Group | Current"]     = cur_group
    r["Rooms Sold | Total Group | Change"]      = chg_group
    r["Rooms Sold | Total Group | Blocked"]     = df.get("grp_blocked")
    r["Rooms Sold | Total Group | P/U"]         = df.get("grp_pickup")
    r["Rooms Sold | Total Group | Remaining"]   = df.get("grp_avail")

    # 7-day pickup per occupancy date. Prefer a dedicated 7-day PCDC; otherwise
    # compute it from an uploaded '7 days ago' baseline (current − 7-days-ago).
    if has_wk_pu:
        wtc = pd.to_numeric(df.get("w_trans_chg"), errors="coerce")
        wgc = pd.to_numeric(df.get("w_grp_chg"), errors="coerce")
        r["7-Day P/U | Total Hotel"]     = wtc.fillna(0) + wgc.fillna(0)
        r["7-Day P/U | Total Transient"] = wtc
        r["7-Day P/U | Total Group"]     = wgc
    elif wk_dates:
        r["7-Day P/U | Total Hotel"]     = cur_tot   - _base(wk_dates, 0)
        r["7-Day P/U | Total Transient"] = cur_trans - _base(wk_dates, 2)
        r["7-Day P/U | Total Group"]     = cur_group - _base(wk_dates, 3)

    r["Rooms OTB STLY | Total Hotel | Current"]     = occ_stly
    r["Rooms OTB STLY | Total Hotel | Change"]      = occ_ty - occ_stly
    r["Rooms OTB STLY | Total Transient | Current"] = tr_stly
    r["Rooms OTB STLY | Total Transient | Change"]  = tr_ty - tr_stly
    r["Rooms OTB STLY | Total Group | Current"]     = gr_stly
    r["Rooms OTB STLY | Total Group | Change"]      = gr_ty - gr_stly

    r["Remaining Demand | Total Hotel"]     = (dem_tot - occ_ty).clip(lower=0)
    r["Remaining Demand | Total Transient"] = (dem_tr - tr_ty).clip(lower=0)
    r["Remaining Demand | Total Group"]     = (dem_grp - gr_ty).clip(lower=0)

    ft, fg = df.get("fcst_trans"), df.get("fcst_grp")
    fcst_tot = (ft.fillna(0) + fg.fillna(0)) if ft is not None else dnum("Forecasted Room Revenue This Year")*np.nan
    r["Occ Forecast | Total Hotel"]     = fcst_tot.round(1) if ft is not None else np.nan
    r["Occ Forecast | Total Transient"] = ft.round(1) if ft is not None else np.nan
    r["Occ Forecast | Total Group"]     = fg.round(1) if fg is not None else np.nan

    r["Occ Forecast % | Total Hotel"]     = (fcst_tot / cap) if ft is not None else np.nan
    r["Occ Forecast % | Total Transient"] = (ft / cap) if ft is not None else np.nan
    r["Occ Forecast % | Total Group"]     = (fg / cap) if fg is not None else np.nan

    rev_t, rev_g = df.get("rev_trans"), df.get("rev_grp")
    if tc is not None:
        total_rooms = (tc.fillna(0) + gc.fillna(0)).replace(0, np.nan)
        r["Booked ADR | Total Hotel"] = ((rev_t.fillna(0) + rev_g.fillna(0)) / total_rooms).round(2)
    else:
        r["Booked ADR | Total Hotel"] = dnum("ADR On Books This Year")
    r["Booked ADR | Total Transient"] = df.get("adr_trans")
    r["Booked ADR | Total Group"]     = df.get("adr_grp")

    r["Estimated ADR"] = dnum("ADR Forecast This Year")

    # ================================================================= #
    #  PRICING DECISION SIGNALS  (VP-of-Revenue toolkit)
    # ================================================================= #
    bar   = r["BAR"]
    comp  = r["Comp Set Avg"]
    rl    = r["Rooms Left to Sell"]
    rd    = r["Remaining Demand | Total Hotel"]
    ofp   = r["Occ Forecast % | Total Hotel"]
    pace  = r["Rooms OTB STLY | Total Hotel | Change"]      # TY − STLY (+ = ahead)
    dleft = r["Days Left"]
    floor   = dnum("Floor")
    ceiling = dnum("Ceiling")

    # --- Rate Rank badge:  where our BAR sits in the comp set (1 = highest) --- #
    rank_cols = [c for c in ("c_21c_n", "c_motto_n", "c_ac_n", "c_dt_n") if c in df]
    if rank_cols:
        price_df = pd.DataFrame({"own": pd.to_numeric(r["BAR"], errors="coerce").values})
        for cc in rank_cols:
            price_df[cc] = pd.to_numeric(df[cc], errors="coerce").values

        def _rank_row(row):
            own = row["own"]
            allv = [v for v in row.tolist() if pd.notna(v)]
            if pd.isna(own) or not allv:
                return ""
            higher = sum(1 for v in allv if v > own)
            return f"{higher + 1} of {len(allv)}"

        r["Rate Rank"] = [_rank_row(price_df.iloc[i]) for i in range(len(price_df))]
    else:
        r["Rate Rank"] = ""

    # --- Signal thresholds (tunable from the sidebar) --- #
    thr = thr or {}
    lead_buffer   = thr.get("lead_buffer", 20)    # min $ cushion over the #2 to be "safely" #1
    pickup_pctile = thr.get("pickup_pctile", 75)  # percentile of positive pickups = "big"
    pickup_min    = thr.get("pickup_min", 3)      # floor so quiet reports don't flag tiny pickup

    # Strategy: The Compton prices as the market LEADER. The Action column is
    # purely about rate-leadership position — two flags only:
    #   🔴 Not #1        — a competitor is at or above our BAR
    #   🟠 Too close #2  — we're #1, but our cushion over the top competitor
    #                      is thinner than we want (within "lead_buffer" $)

    # top competitor rate per date (max of the 4 shops) — effectively "the #2"
    if rank_cols:
        comp_matrix = pd.DataFrame({cc: pd.to_numeric(df[cc], errors="coerce").values
                                    for cc in rank_cols})
        max_comp = comp_matrix.max(axis=1)
        max_comp.index = r.index
    else:
        max_comp = pd.Series(np.nan, index=r.index)

    # gap to the #2 (top competitor): + = we're above them, 0/− = not #1
    lead_gap = (bar - max_comp)
    r["Lead Gap vs Comp"] = lead_gap.round(0)

    not_leader = pd.notna(max_comp) & (lead_gap <= 0)             # someone matches/beats us
    too_close  = pd.notna(max_comp) & (lead_gap > 0) & (lead_gap < lead_buffer)
    at_ceiling = bar >= ceiling

    def _label(i):
        if pd.isna(dleft.iloc[i]) or dleft.iloc[i] < 0:
            return ""                                     # past date
        if pd.isna(max_comp.iloc[i]):
            return ""                                     # no comp-set data for this date
        if bool(not_leader.iloc[i]):
            return "🔴 Not #1" if not bool(at_ceiling.iloc[i]) else "🟠 Not #1 · at ceiling"
        if bool(too_close.iloc[i]):
            return "🟠 Too close to #2" if not bool(at_ceiling.iloc[i]) else "🟠 At ceiling"
        return "✅ Leading"

    r["⚡ Action"] = [_label(i) for i in range(len(r))]

    # --- 🔥 Spike: flags an unusually large day's pickup ------------- #
    # "Big" is relative, not a fixed number, so this self-calibrates off this
    # property's own pickup pattern: it flags a date when its pickup sits at or
    # above the Nth percentile of all POSITIVE daily pickups in this report
    # (default 90th = top 10%), with a small floor so a quiet report full of
    # 1-2 room days doesn't get flagged just because it's "relatively" high.
    pickup = r["Rooms Sold | Total Hotel | Change"]
    pos_pickup = pickup[pickup > 0]
    if len(pos_pickup) >= 5:
        pctile_val = pos_pickup.quantile(pickup_pctile / 100)
        pickup_threshold = max(pctile_val, pickup_min)
    else:
        pickup_threshold = pickup_min          # not enough data to calibrate — use the floor

    is_big_pickup = (pickup >= pickup_threshold) & (dleft >= 0)

    def _pickup_label(i):
        if bool(is_big_pickup.iloc[i]):
            v = pickup.iloc[i]
            return f"🔥 Spike (+{v:,.0f})"
        return ""

    r["🔥 Spike"] = [_pickup_label(i) for i in range(len(r))]

    # ---- Enforce the exact original report column order ---- #
    lead = ["⚡ Action", "🔥 Spike", "DOW", "Date", "Days Left", "Events",
            "Rooms Left to Sell", "BAR", "BAR Chg 7d", "Comp Set Avg", "Rate Rank",
            "Lead Gap vs Comp", "Last Room Value", "Restrictions"]
    comp = [c for c in r.columns if c.startswith("Competitor Shops")]
    mid  = ["OOO", "Ovrbk"]
    groups = ["Rooms Sold", "7-Day P/U", "Rooms OTB STLY", "Remaining Demand",
              "Occ Forecast |", "Occ Forecast %", "Booked ADR"]
    grouped = []
    for g in groups:
        grouped += [c for c in r.columns if c.startswith(g)]
    tail = ["Estimated ADR"]
    order = lead + comp + mid + grouped + tail
    order = [c for c in order if c in r.columns]          # keep only existing
    order += [c for c in r.columns if c not in order]     # safety: append any strays
    r = r[order]

    return r


def _levels(col):
    parts = [p.strip() for p in str(col).split("|")]
    while len(parts) < 3:
        parts.append("")
    return tuple(parts[:3])


def to_excel(r, wb=None):
    """Write the Pick-Up Report into a workbook.

    When `wb` is None this creates a fresh workbook and returns a BytesIO buffer
    (the standalone DD download). When an existing `wb` is passed, the sheet is
    ADDED to it and the workbook object is returned (used by the combined export).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    lv = [_levels(c) for c in r.columns]
    n = len(lv)

    standalone = wb is None
    if standalone:
        wb = Workbook()
        ws = wb.active
        ws.title = "Pick-Up Report"
    else:
        ws = wb.create_sheet("Pick-Up Report")

    NAVY, TEAL, ORANGE, PURPLE, GREEN = "16365C", "215967", "E26B0A", "60497A", "76933C"
    white = Font(color="FFFFFF", bold=True, size=9)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

    # Light segment tints for the "Current" data cells (match the header colors)
    TINT_HOTEL, TINT_TRANS, TINT_GROUP = "FCE4D6", "E4DFEC", "EBF1DE"
    WEEKEND_FILL = "EEF1F5"        # subtle grey-blue band for weekend rows
    WEEK_LINE = Side(style="medium", color="9AA7B4")   # week separator line

    def seg_tint(colname):
        """Light background tint for a segment 'Current' column."""
        c = str(colname)
        if not c.endswith("Current"):
            return None
        if "Total Hotel" in c:
            return TINT_HOTEL
        if "Total Transient" in c:
            return TINT_TRANS
        if "Total Group" in c:
            return TINT_GROUP
        return None

    def PF(hexcol):
        return PatternFill("solid", fgColor=hexcol)

    def seg_color(txt):
        if "Total Hotel" in txt:
            return ORANGE
        if "Total Transient" in txt:
            return PURPLE
        if "Total Group" in txt:
            return GREEN
        return None

    def band_fill(l1, l2):
        return PF(NAVY) if l2 == "" else PF(TEAL)

    def lvl2_fill(l1, l2):
        sc = seg_color(l2)
        if sc:
            return PF(sc)
        if l1 == "Competitor Shops":
            return PF(NAVY)
        return PF(TEAL)

    def lvl3_fill(l1, l2):
        sc = seg_color(l2)
        return PF(sc) if sc else PF(NAVY)

    j = 0
    while j < n:
        l1 = lv[j][0]
        k = j
        while k + 1 < n and lv[k + 1][0] == l1 and lv[j][1]:
            k += 1
        fill = band_fill(l1, lv[j][1])
        if lv[j][1] == "" and l1 in ("Estimated ADR",):
            fill = PF(ORANGE)
        cell = ws.cell(row=1, column=j + 1, value=l1)
        cell.font = white
        cell.alignment = center
        cell.fill = fill
        if lv[j][1] == "":
            ws.merge_cells(start_row=1, start_column=j + 1, end_row=3, end_column=j + 1)
        elif k > j:
            ws.merge_cells(start_row=1, start_column=j + 1, end_row=1, end_column=k + 1)
        j = k + 1

    j = 0
    while j < n:
        l1, l2, l3 = lv[j]
        if l2:
            k = j
            while k + 1 < n and lv[k + 1][0] == l1 and lv[k + 1][1] == l2:
                k += 1
            cell = ws.cell(row=2, column=j + 1, value=l2)
            cell.font = white
            cell.alignment = center
            cell.fill = lvl2_fill(l1, l2)
            if l3 == "":
                ws.merge_cells(start_row=2, start_column=j + 1, end_row=3, end_column=j + 1)
            elif k > j:
                ws.merge_cells(start_row=2, start_column=j + 1, end_row=2, end_column=k + 1)
            j = k + 1
        else:
            j += 1

    for j, (l1, l2, l3) in enumerate(lv, start=1):
        if l3:
            cell = ws.cell(row=3, column=j, value=l3)
            cell.font = white
            cell.alignment = center
            cell.fill = lvl3_fill(l1, l2)

    _hotelnames = ("Bentonville", "MGallery")
    DATA0 = 4
    pct = {i for i, c0 in enumerate(r.columns) if c0.startswith("Occ Forecast %")}
    signed_cols = ("BAR Chg 7d", "Lead Gap vs Comp")
    chg = {i for i, c0 in enumerate(r.columns)
           if (any(h in c0 for h in _hotelnames) and c0.endswith("Change"))
           or c0 in signed_cols or c0.startswith("7-Day P/U")}
    money = {i for i, c0 in enumerate(r.columns)
             if (c0 in ("BAR", "Comp Set Avg", "Last Room Value", "Estimated ADR")
                 or "ADR" in c0 or any(h in c0 for h in _hotelnames))
             and i not in chg}
    otb_sold_col = None
    for i, c0 in enumerate(r.columns):
        if c0 == "Rooms Left to Sell":
            otb_sold_col = i

    def _action_fill(txt):
        t = str(txt)
        if "Not #1" in t:          return PatternFill("solid", fgColor="F8D7DA")
        if "Too close" in t:       return PatternFill("solid", fgColor="FFF3CD")
        if "At ceiling" in t:      return PatternFill("solid", fgColor="FFE5B4")
        if "Leading" in t:         return PatternFill("solid", fgColor="D1E7DD")
        return None

    def _pickup_fill(txt):
        return PatternFill("solid", fgColor="FDE7CE") if "Spike" in str(txt) else None

    # Per-row day-of-week (for weekend banding + week-separator lines)
    dow_series = r["DOW"].astype(str).tolist() if "DOW" in r.columns else [""] * len(r)
    _weekend = {"Friday", "Saturday", "Sunday"}

    for ri, (_, row) in enumerate(r.iterrows(), start=DATA0):
        dow = dow_series[ri - DATA0]
        is_weekend = dow in _weekend
        is_monday = (dow == "Monday")     # start of a new week -> draw a top line
        for ci2, col in enumerate(r.columns, start=1):
            v = row[col]
            cell = ws.cell(row=ri, column=ci2, value=(None if pd.isna(v) else v))
            # base border; add a medium TOP line on Mondays to bracket each week
            if is_monday:
                cell.border = Border(left=Side(style="thin", color="D9D9D9"),
                                     right=Side(style="thin", color="D9D9D9"),
                                     bottom=Side(style="thin", color="D9D9D9"),
                                     top=WEEK_LINE)
            else:
                cell.border = thin
            cell.font = Font(size=9)
            # center everything; left-align the long-text "Events"/"Restrictions"
            halign = "left" if col in ("Events", "Restrictions") else "center"
            cell.alignment = Alignment(horizontal=halign, vertical="center")

            # ---- fill precedence: special > restriction > segment tint > weekend ----
            tint = seg_tint(col)
            if col == "⚡ Action":
                af = _action_fill(v)
                if af is not None:
                    cell.fill = af
            elif col == "🔥 Spike":
                pf = _pickup_fill(v)
                if pf is not None:
                    cell.fill = pf
                    cell.font = Font(size=9, bold=True, color="B45309")
            elif col == "Restrictions" and str(v).strip():
                # amber highlight; deeper amber + note when it changed vs last report
                changed = "(Previous" in str(v)
                cell.fill = PF("FFE08A" if changed else "FFF3CD")
                cell.font = Font(size=9, bold=True, color="7A4A00")
            elif tint is not None:
                cell.fill = PF(tint)               # segment-coloured "Current" cell
            elif is_weekend and col not in ("Rooms Left to Sell", "Restrictions"):
                # weekend band on structural/other cells (skip the heat-map column)
                cell.fill = PF(WEEKEND_FILL)

            idx = ci2 - 1
            if idx in pct:
                cell.number_format = "0.0%"
            elif idx in chg:
                cell.number_format = "+#,##0;-#,##0;\"\""
            elif idx in money:
                cell.number_format = "$#,##0"
            elif isinstance(v, float) and float(v).is_integer():
                cell.number_format = "#,##0"

    if otb_sold_col is not None and len(r) > 0:
        L = get_column_letter(otb_sold_col + 1)
        rng = f"{L}{DATA0}:{L}{DATA0 + len(r) - 1}"
        # Rooms Left to Sell: RED = few left (near sell-out), GREEN = lots left
        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(start_type="min", start_color="F8696B",         # red (few left)
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",  # yellow
                           end_type="max", end_color="63BE7B"))            # green (lots left)

    ws.freeze_panes = "C4"
    # Tight, content-aware widths so day-by-day pickup fits without scrolling.
    def _col_width(name):
        c = str(name)
        if c == "Events":
            return 16
        if c == "Restrictions":
            return 18
        if c == "⚡ Action":
            return 13
        if c == "🔥 Spike":
            return 11
        if c == "DOW":
            return 5
        if c == "Date":
            return 9
        if c == "Days Left":
            return 6
        if c.startswith("Competitor Shops"):
            return 8
        if "ADR" in c or "Revenue" in c or c in ("BAR", "Comp Set Avg",
                                                 "Last Room Value", "Estimated ADR"):
            return 8
        return 7                     # default: tight numeric columns
    for j, name in enumerate(r.columns, start=1):
        ws.column_dimensions[get_column_letter(j)].width = _col_width(name)

    # Excel AutoFilter so the team can filter/sort right in Excel.
    # Anchored on the bottom header row (row 3) through the last data row.
    last_row = DATA0 + len(r) - 1
    ws.auto_filter.ref = f"A3:{get_column_letter(n)}{last_row}"

    if not standalone:
        return wb

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf




# =========================================================================== #
#  DASHBOARD PAGE  — reproduces the workbook 'Dashboard' tab
# =========================================================================== #
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# --------------------------------------------------------------------------- #
#  Static reference tables (pulled straight from the workbook Dashboard tab).
#  Tuple = (Budget RN, Budget ADR, Freeze RN, Freeze ADR).
# --------------------------------------------------------------------------- #
_REF_TRANSIENT = {
    "Jan": (900, 329.85, 699, 304.42),
    "Feb": (1510, 343.97, 1041, 330.88),
    "Mar": (2006, 369.38, 1670, 306.26),
    "Apr": (1982, 376.04, 2007, 340.52),
    "May": (2254, 385.27, 1279, 415.08),
    "Jun": (2441, 382.19, 1532, 429.48),
    "Jul": (2182, 376.15, 1257, 369.70),
    "Aug": (1989, 361.19, 1579, 381.60),
    "Sep": (2190, 390.36, 1849, 477.54),
    "Oct": (2233, 407.01, 2233, 407.01),
    "Nov": (2003, 352.54, 2003, 352.54),
    "Dec": (1990, 344.98, 1990, 344.98),
}
_REF_GROUP = {
    "Jan": (212, 206.81, 232, 208.85),
    "Feb": (275, 249.48, 275, 249.48),
    "Mar": (600, 278.00, 580, 316.77),
    "Apr": (500, 279.00, 764, 269.00),
    "May": (650, 289.00, 650, 289.00),
    "Jun": (700, 297.57, 694, 272.25),
    "Jul": (500, 224.00, 250, 294.54),
    "Aug": (525, 239.00, 333, 381.70),
    "Sep": (650, 275.00, 950, 300.42),
    "Oct": (600, 294.45, 600, 294.45),
    "Nov": (420, 229.00, 420, 229.00),
    "Dec": (350, 243.00, 350, 243.00),
}


def default_reference_frames():
    """Return (budget_df, freeze_df) editable tables indexed like the sidebar
    data-editors expect: columns [Month, Trans RN, Trans ADR, Group RN, Group ADR]."""
    brows, frows = [], []
    for m in MONTHS:
        tb_rn, tb_adr, tf_rn, tf_adr = _REF_TRANSIENT[m]
        gb_rn, gb_adr, gf_rn, gf_adr = _REF_GROUP[m]
        brows.append({"Month": m, "Trans RN": tb_rn, "Trans ADR": tb_adr,
                      "Group RN": gb_rn, "Group ADR": gb_adr})
        frows.append({"Month": m, "Trans RN": tf_rn, "Trans ADR": tf_adr,
                      "Group RN": gf_rn, "Group ADR": gf_adr})
    return pd.DataFrame(brows), pd.DataFrame(frows)


def classify_segment(market_segment):
    """Map a detailed SynXis market-segment name to Transient / Group / Comp / Other,
    exactly matching how the workbook's SEGMENT helper column rolls things up."""
    m = str(market_segment).strip()
    if m.startswith("Transient") or m == "Contract":
        return "Transient"
    if m.startswith("Group"):
        return "Group"
    if m.startswith("Comp"):
        return "Comp"
    return "Other"


def parse_market_segment(raw):
    """
    Aggregate a Market Segment export (raw sheet read with header=None) into a
    dict keyed by (month_label, segment) -> {'rooms': x, 'rev': y}, where
    month_label is like 'Jan-2026' and segment is 'Transient' or 'Group'.
    Returns (aggregation_dict, set_of_years).
    """
    df = raw.copy()
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)

    def find(*names):
        for n in names:
            for c in df.columns:
                if str(c).strip().lower() == n.lower():
                    return c
        return None

    c_date = find("Occupancy Date")
    c_seg = find("Market Segment")
    c_rms = find("Occupancy On Books This Year", "Occupancy on Books This Year")
    c_rev = find("Booked Room Revenue This Year")
    if c_date is None or c_seg is None or c_rms is None:
        return {}, set()

    dates = pd.to_datetime(df[c_date], errors="coerce")
    rooms = pd.to_numeric(df[c_rms], errors="coerce").fillna(0)
    rev = (pd.to_numeric(df[c_rev], errors="coerce").fillna(0)
           if c_rev is not None else pd.Series(0.0, index=df.index))
    seg = df[c_seg].map(classify_segment)

    agg = {}
    years = set()
    for d, s, rm, rv in zip(dates, seg, rooms, rev):
        if pd.isna(d) or s not in ("Transient", "Group"):
            continue
        label = f"{calendar.month_abbr[d.month]}-{d.year}"
        years.add(d.year)
        key = (label, s)
        cur = agg.setdefault(key, {"rooms": 0.0, "rev": 0.0})
        cur["rooms"] += float(rm)
        cur["rev"] += float(rv)
    return agg, years


def _safe_div(a, b):
    return (a / b) if b else 0.0


def _month_days(month_abbr, year):
    m = MONTHS.index(month_abbr) + 1
    return calendar.monthrange(year, m)[1]


def budget_freeze_from_extract(data_extract, year):
    """
    Derive per-month, per-segment Budget and Freeze Forecast figures straight
    from a Data Extraction export (the 'Property' sheet, already parsed so its
    first row is the header).

      * Budget    -> 'Budget Occupancy - Group/Transient This Year' (rooms) and
                     'Budget Room Revenue This Year' (total revenue).
      * Freeze    -> the USER forecast when it is populated, otherwise the SYSTEM
                     forecast:
                        user   = 'My Forecast Occupancy - …' + 'My Forecast Revenue This Year'
                        system = 'Occupancy Forecast - …'   + 'Forecasted Room Revenue This Year'

    The export only carries revenue at the TOTAL level (no segment revenue), and
    Budget/Forecast occupancy Total == Group + Transient exactly, so segment
    revenue is split using the blended monthly ADR (total revenue / total rooms).
    Each segment therefore shares that scenario's blended ADR while Group +
    Transient revenue still reconciles to the total.

    Returns (bf, freeze_source) where:
        bf[(month_abbr, seg)] = (bud_rn, bud_adr, frz_rn, frz_adr)   seg in {Transient, Group}
        freeze_source         = 'Your forecast' or 'System forecast'
    Returns (None, None) if the required columns are not present.
    """
    if data_extract is None:
        return None, None
    d = data_extract.copy()
    if "Occupancy Date" not in d.columns:
        return None, None
    d["Occupancy Date"] = pd.to_datetime(d["Occupancy Date"], errors="coerce")
    d = d.dropna(subset=["Occupancy Date"])
    d = d[d["Occupancy Date"].dt.year == year]
    if d.empty:
        return None, None

    def col(name):
        return (pd.to_numeric(d[name], errors="coerce").fillna(0)
                if name in d.columns else None)

    bud_rev = col("Budget Room Revenue This Year")
    bud_g = col("Budget Occupancy - Group This Year")
    bud_t = col("Budget Occupancy - Transient This Year")
    if bud_rev is None or bud_g is None or bud_t is None:
        return None, None

    # Freeze source: user forecast if it carries any value, else system forecast.
    my_rev = col("My Forecast Revenue This Year")
    my_g = col("My Forecast Occupancy - Group This Year")
    my_t = col("My Forecast Occupancy - Transient This Year")
    use_user = (my_rev is not None and float(my_rev.abs().sum()) > 0
                and my_g is not None and my_t is not None
                and float((my_g.abs() + my_t.abs()).sum()) > 0)
    if use_user:
        frz_rev, frz_g, frz_t = my_rev, my_g, my_t
        freeze_source = "Your forecast"
    else:
        frz_rev = col("Forecasted Room Revenue This Year")
        frz_g = col("Occupancy Forecast - Group This Year")
        frz_t = col("Occupancy Forecast - Transient This Year")
        freeze_source = "System forecast"
        if frz_rev is None or frz_g is None or frz_t is None:
            frz_rev = bud_rev * 0.0
            frz_g = bud_g * 0.0
            frz_t = bud_t * 0.0

    work = pd.DataFrame({
        "_mo": d["Occupancy Date"].dt.month.values,
        "bud_rev": bud_rev.values, "bud_g": bud_g.values, "bud_t": bud_t.values,
        "frz_rev": frz_rev.values, "frz_g": frz_g.values, "frz_t": frz_t.values,
    })
    grp = work.groupby("_mo").sum()

    bf = {}
    for mo, row in grp.iterrows():
        mabbr = MONTHS[int(mo) - 1]
        bud_rn_tot = row["bud_g"] + row["bud_t"]
        bud_adr = _safe_div(row["bud_rev"], bud_rn_tot)   # blended budget ADR
        frz_rn_tot = row["frz_g"] + row["frz_t"]
        frz_adr = _safe_div(row["frz_rev"], frz_rn_tot)   # blended freeze ADR
        bf[(mabbr, "Transient")] = (float(row["bud_t"]), bud_adr,
                                    float(row["frz_t"]), frz_adr)
        bf[(mabbr, "Group")] = (float(row["bud_g"]), bud_adr,
                                float(row["frz_g"]), frz_adr)
    return bf, freeze_source


def compute_dashboard(cur_agg, last_agg, capacity, year,
                      data_extract=None, budget_df=None, freeze_df=None,
                      week_agg=None):
    """
    Build the three dashboard tables. Returns a dict:
        {'Total': rows, 'Transient': rows, 'Group': rows, '_freeze_source': str}
    where rows is a list of 13 dicts (12 months + Total), each holding every
    metric cell needed to render the table.

    Budget and Freeze Forecast are pulled from `data_extract` (the Data
    Extraction export) when it carries the Budget/Forecast columns; otherwise
    they fall back to the static `budget_df` / `freeze_df` reference tables.

    When `week_agg` (the monthly/segment aggregation from ~7 days ago) is given,
    each row also carries a 7-days-ago baseline (wk_rn/wk_adr/wk_rev) and the
    7-day pickup (wpu_rn/wpu_adr/wpu_rev = current minus 7-days-ago).
    """
    has_week = bool(week_agg)
    bf, freeze_source = budget_freeze_from_extract(data_extract, year)
    use_extract = bf is not None

    if budget_df is None or freeze_df is None:
        budget_df, freeze_df = default_reference_frames()
    bud = {r["Month"]: r for _, r in budget_df.iterrows()}
    frz = {r["Month"]: r for _, r in freeze_df.iterrows()}

    def ref(month, seg, table, kind):
        if use_extract:
            b_rn, b_adr, f_rn, f_adr = bf.get((month, seg), (0.0, 0.0, 0.0, 0.0))
            return (b_rn, b_adr) if table == "budget" else (f_rn, f_adr)
        row = (bud if table == "budget" else frz)[month]
        if seg == "Transient":
            rn = row["Trans RN"]; adr = row["Trans ADR"]
        else:
            rn = row["Group RN"]; adr = row["Group ADR"]
        return (float(rn or 0), float(adr or 0))

    def seg_month(agg, month, seg):
        key = (f"{month}-{year}", seg)
        d = agg.get(key, {"rooms": 0.0, "rev": 0.0})
        return float(d["rooms"]), float(d["rev"])

    def week_month(m, seg):
        """7-days-ago rooms & rev for a month/segment (Total = Trans + Group)."""
        if not has_week:
            return 0.0, 0.0
        if seg in ("Transient", "Group"):
            return seg_month(week_agg, m, seg)
        t = seg_month(week_agg, m, "Transient"); g = seg_month(week_agg, m, "Group")
        return t[0] + g[0], t[1] + g[1]

    def build_rows(seg):
        rows = []
        tot = {k: 0.0 for k in ("otb_rn", "otb_rev", "lr_rn", "lr_rev",
                                "bud_rn", "bud_rev", "frz_rn", "frz_rev",
                                "wk_rn", "wk_rev")}
        for m in MONTHS:
            if seg in ("Transient", "Group"):
                otb_rn, otb_rev = seg_month(cur_agg, m, seg)
                lr_rn, lr_rev = seg_month(last_agg, m, seg)
                b_rn, b_adr = ref(m, seg, "budget", None)
                f_rn, f_adr = ref(m, seg, "freeze", None)
            else:  # Total = Transient + Group
                t = seg_month(cur_agg, m, "Transient"); g = seg_month(cur_agg, m, "Group")
                otb_rn, otb_rev = t[0] + g[0], t[1] + g[1]
                tl = seg_month(last_agg, m, "Transient"); gl = seg_month(last_agg, m, "Group")
                lr_rn, lr_rev = tl[0] + gl[0], tl[1] + gl[1]
                tb = ref(m, "Transient", "budget", None); gb = ref(m, "Group", "budget", None)
                b_rn = tb[0] + gb[0]
                b_rev_tmp = tb[0] * tb[1] + gb[0] * gb[1]
                b_adr = _safe_div(b_rev_tmp, b_rn)
                tf = ref(m, "Transient", "freeze", None); gf = ref(m, "Group", "freeze", None)
                f_rn = tf[0] + gf[0]
                f_rev_tmp = tf[0] * tf[1] + gf[0] * gf[1]
                f_adr = _safe_div(f_rev_tmp, f_rn)

            bud_rev = b_rn * b_adr
            frz_rev = f_rn * f_adr
            otb_adr = _safe_div(otb_rev, otb_rn)
            lr_adr = _safe_div(lr_rev, lr_rn)
            days = _month_days(m, year)
            occ = _safe_div(otb_rn, capacity * days)

            row = {
                "month": f"{m}-{year}",
                "occ": occ,
                "otb_rn": otb_rn, "otb_adr": otb_adr, "otb_rev": otb_rev,
                "lr_rn": lr_rn, "lr_adr": lr_adr, "lr_rev": lr_rev,
                "pu_rn": otb_rn - lr_rn, "pu_adr": otb_adr - lr_adr, "pu_rev": otb_rev - lr_rev,
                "bud_rn": b_rn, "bud_adr": b_adr, "bud_rev": bud_rev,
                "bv_rn": otb_rn - b_rn, "bv_adr": otb_adr - b_adr, "bv_rev": otb_rev - bud_rev,
                "frz_rn": f_rn, "frz_adr": f_adr, "frz_rev": frz_rev,
                "fv_rn": otb_rn - f_rn, "fv_adr": otb_adr - f_adr, "fv_rev": otb_rev - frz_rev,
            }
            if has_week:
                wk_rn, wk_rev = week_month(m, seg)
                wk_adr = _safe_div(wk_rev, wk_rn)
                row.update({
                    "wk_rn": wk_rn, "wk_adr": wk_adr, "wk_rev": wk_rev,
                    "wpu_rn": otb_rn - wk_rn, "wpu_adr": otb_adr - wk_adr,
                    "wpu_rev": otb_rev - wk_rev,
                })
                tot["wk_rn"] += wk_rn; tot["wk_rev"] += wk_rev
            rows.append(row)
            tot["otb_rn"] += otb_rn; tot["otb_rev"] += otb_rev
            tot["lr_rn"] += lr_rn; tot["lr_rev"] += lr_rev
            tot["bud_rn"] += b_rn; tot["bud_rev"] += bud_rev
            tot["frz_rn"] += f_rn; tot["frz_rev"] += frz_rev

        # ---- Total row (bottom) ----
        year_days = 366 if calendar.isleap(year) else 365
        otb_adr = _safe_div(tot["otb_rev"], tot["otb_rn"])
        lr_adr = _safe_div(tot["lr_rev"], tot["lr_rn"])
        bud_adr = _safe_div(tot["bud_rev"], tot["bud_rn"])
        frz_adr = _safe_div(tot["frz_rev"], tot["frz_rn"])
        trow = {
            "month": "Total",
            "occ": _safe_div(tot["otb_rn"], capacity * year_days),
            "otb_rn": tot["otb_rn"], "otb_adr": otb_adr, "otb_rev": tot["otb_rev"],
            "lr_rn": tot["lr_rn"], "lr_adr": lr_adr, "lr_rev": tot["lr_rev"],
            "pu_rn": tot["otb_rn"] - tot["lr_rn"], "pu_adr": otb_adr - lr_adr,
            "pu_rev": tot["otb_rev"] - tot["lr_rev"],
            "bud_rn": tot["bud_rn"], "bud_adr": bud_adr, "bud_rev": tot["bud_rev"],
            "bv_rn": tot["otb_rn"] - tot["bud_rn"], "bv_adr": otb_adr - bud_adr,
            "bv_rev": tot["otb_rev"] - tot["bud_rev"],
            "frz_rn": tot["frz_rn"], "frz_adr": frz_adr, "frz_rev": tot["frz_rev"],
            "fv_rn": tot["otb_rn"] - tot["frz_rn"], "fv_adr": otb_adr - frz_adr,
            "fv_rev": tot["otb_rev"] - tot["frz_rev"],
        }
        if has_week:
            wk_adr_t = _safe_div(tot["wk_rev"], tot["wk_rn"])
            trow.update({
                "wk_rn": tot["wk_rn"], "wk_adr": wk_adr_t, "wk_rev": tot["wk_rev"],
                "wpu_rn": tot["otb_rn"] - tot["wk_rn"], "wpu_adr": otb_adr - wk_adr_t,
                "wpu_rev": tot["otb_rev"] - tot["wk_rev"],
            })
        rows.append(trow)
        return rows

    return {"Total": build_rows("Total"),
            "Transient": build_rows("Transient"),
            "Group": build_rows("Group"),
            "_has_week": has_week,
            "_freeze_source": (freeze_source if use_extract else "Static table"),
            "_budget_source": ("Data Extraction" if use_extract else "Static table")}


# --------------------------------------------------------------------------- #
#  Formatting helpers
# --------------------------------------------------------------------------- #
def _f_pct(v):
    return f"{v*100:,.1f}%"

def _f_int(v):
    return f"{v:,.0f}"

def _f_money0(v):
    return f"${v:,.0f}"

def _f_money2(v):
    return f"${v:,.2f}"

def _f_delta_int(v):
    v = round(v)
    return f"({abs(v):,.0f})" if v < 0 else f"{v:,.0f}"

def _f_delta_money0(v):
    return f"(${abs(v):,.0f})" if v < 0 else f"${v:,.0f}"

def _f_delta_money2(v):
    return f"(${abs(v):,.2f})" if v < 0 else f"${v:,.2f}"


# Excel number formats reused by the layout builder
_NF_INT = "#,##0"
_NF_M2 = '"$"#,##0.00'
_NF_M0 = '"$"#,##0'
_NF_DINT = "#,##0_);[Red](#,##0)"
_NF_D2 = '"$"#,##0.00_);[Red]("$"#,##0.00)'
_NF_D0 = '"$"#,##0_);[Red]("$"#,##0)'


def _dashboard_layout(title, freeze_label="Freeze Forecast",
                      otbvar_label="OTB Var to Freeze Fcst",
                      show_lr=True, show_week=False):
    """Single source of truth for the dashboard's columns. Returns an ordered
    list of column dicts: {key, fmt, delta, group, sub, numfmt, budget}. The
    'Pickup Since Last Report' block is optional (show_lr) and a '7-Day Pickup'
    block (with its own 7-days-ago baseline) is optional (show_week), so the
    same renderer drives every pickup view."""
    cols = []

    def add(key, fmt, delta, group, sub, nf, budget=False):
        cols.append({"key": key, "fmt": fmt, "delta": delta, "group": group,
                     "sub": sub, "numfmt": nf, "budget": budget})

    add("occ", _f_pct, False, title, "Occ% OTB", "0.0%")
    add("otb_rn", _f_int, False, title, "RN OTB", _NF_INT)
    add("otb_adr", _f_money2, False, title, "ADR OTB", _NF_M2)
    add("otb_rev", _f_money0, False, title, "Revenue OTB", _NF_M0)
    if show_lr:
        add("lr_rn", _f_int, False, "Last Report OTB", "RN", _NF_INT)
        add("lr_adr", _f_money2, False, "Last Report OTB", "ADR", _NF_M2)
        add("lr_rev", _f_money0, False, "Last Report OTB", "Revenue", _NF_M0)
        add("pu_rn", _f_delta_int, True, "Pickup Since Last Report", "RN", _NF_DINT)
        add("pu_adr", _f_delta_money2, True, "Pickup Since Last Report", "ADR Chg", _NF_D2)
        add("pu_rev", _f_delta_money0, True, "Pickup Since Last Report", "Revenue", _NF_D0)
    if show_week:
        add("wk_rn", _f_int, False, "7 Days Ago OTB", "RN", _NF_INT)
        add("wk_adr", _f_money2, False, "7 Days Ago OTB", "ADR", _NF_M2)
        add("wk_rev", _f_money0, False, "7 Days Ago OTB", "Revenue", _NF_M0)
        add("wpu_rn", _f_delta_int, True, "7-Day Pickup", "RN", _NF_DINT)
        add("wpu_adr", _f_delta_money2, True, "7-Day Pickup", "ADR Chg", _NF_D2)
        add("wpu_rev", _f_delta_money0, True, "7-Day Pickup", "Revenue", _NF_D0)
    add("bud_rn", _f_int, False, "Budget", "RN", _NF_INT, True)
    add("bud_adr", _f_money2, False, "Budget", "ADR", _NF_M2, True)
    add("bud_rev", _f_money0, False, "Budget", "Rev", _NF_M0, True)
    add("bv_rn", _f_delta_int, True, "Budget Variance", "RN", _NF_DINT)
    add("bv_adr", _f_delta_money2, True, "Budget Variance", "ADR", _NF_D2)
    add("bv_rev", _f_delta_money0, True, "Budget Variance", "Rev", _NF_D0)
    add("frz_rn", _f_int, False, freeze_label, "RN", _NF_INT)
    add("frz_adr", _f_money2, False, freeze_label, "ADR", _NF_M2)
    add("frz_rev", _f_money0, False, freeze_label, "Rev", _NF_M0)
    add("fv_rn", _f_delta_int, True, otbvar_label, "RN", _NF_DINT)
    add("fv_adr", _f_delta_money2, True, otbvar_label, "ADR", _NF_D2)
    add("fv_rev", _f_delta_money0, True, otbvar_label, "Rev", _NF_D0)
    return cols


def _groups_from_cols(cols):
    """Collapse a column list into (group_name, span) banner segments."""
    groups = []
    for c in cols:
        if groups and groups[-1][0] == c["group"]:
            groups[-1] = (c["group"], groups[-1][1] + 1)
        else:
            groups.append((c["group"], 1))
    return groups


def dashboard_css(colors=None):
    c = {
        "band":   "#F4CCCC",   # section + group header band (salmon)
        "sub":    "#FBE4E4",   # sub-header row
        "budget": "#EDE7D3",   # budget/freeze group shading (tan)
        "total":  "#F2F2F2",   # total row
        "neg":    "#C00000",   # negative text
        "grid":   "#D9D9D9",
        "text":   "#1a1a1a",
    }
    if colors:
        c.update(colors)
    return f"""
<style>
.cmp-dash {{ font-family: Segoe UI, Arial, sans-serif; margin-bottom: 26px; }}
.cmp-dash table {{ border-collapse: collapse; width: 100%; font-size: 12px; }}
.cmp-dash th, .cmp-dash td {{ border: 1px solid {c['grid']}; padding: 3px 6px;
    text-align: right; white-space: nowrap; color: {c['text']}; }}
.cmp-dash .title {{ background: {c['band']}; font-weight: 700; text-align: center;
    font-size: 13px; padding: 5px; }}
.cmp-dash .band {{ background: {c['band']}; font-weight: 700; text-align: center; }}
.cmp-dash .sub  {{ background: {c['sub']};  font-weight: 700; text-align: center; }}
.cmp-dash td.month, .cmp-dash th.month {{ text-align: left; font-weight: 600; }}
.cmp-dash tr.total td {{ background: {c['total']}; font-weight: 700;
    border-top: 2px solid #999; }}
.cmp-dash td.budget {{ background: {c['budget']}; }}
.cmp-dash tr.total td.budget {{ background: #E4DEC8; }}
.cmp-dash .neg {{ color: {c['neg']}; }}
.cmp-dash .hdrline {{ font-size: 13px; font-weight: 700; margin: 4px 0 8px 0; }}
</style>
"""


def _freeze_labels(tables):
    """Map the freeze-forecast source to the two group headers that reference it,
    so the sheet says 'My Forecast' vs 'System Forecast' instead of a generic
    'Freeze Forecast'."""
    fs = (tables or {}).get("_freeze_source", "")
    if fs == "Your forecast":
        return "My Forecast", "OTB Var to My Fcst"
    if fs == "System forecast":
        return "System Forecast", "OTB Var to System Fcst"
    return "Freeze Forecast", "OTB Var to Freeze Fcst"


def _table_html(title, rows, cols):
    groups = _groups_from_cols(cols)

    h = ['<div class="cmp-dash"><table>']
    # banner row
    h.append('<tr><th class="title" rowspan="2">Month</th>')
    for name, span in groups:
        cls = "title" if name == title else "band"
        h.append(f'<th class="{cls}" colspan="{span}">{name}</th>')
    h.append('</tr>')
    # sub-header row
    h.append('<tr>')
    for c in cols:
        cls = "sub budget" if c["budget"] else "sub"
        h.append(f'<th class="{cls}">{c["sub"]}</th>')
    h.append('</tr>')
    # data rows
    for r in rows:
        is_total = r["month"] == "Total"
        h.append('<tr class="total">' if is_total else '<tr>')
        h.append(f'<td class="month">{r["month"]}</td>')
        for c in cols:
            v = r.get(c["key"])
            txt = c["fmt"](v) if isinstance(v, (int, float)) else ""
            classes = []
            if c["budget"]:
                classes.append("budget")
            if c["delta"] and isinstance(v, (int, float)) and v < 0:
                classes.append("neg")
            cls = f' class="{" ".join(classes)}"' if classes else ""
            h.append(f"<td{cls}>{txt}</td>")
        h.append('</tr>')
    h.append('</table></div>')
    return "".join(h)


def render_dashboard_html(tables, header_line=None, colors=None,
                          show_lr=True, show_week=None):
    parts = [dashboard_css(colors)]
    if header_line:
        parts.append(f'<div class="cmp-dash"><div class="hdrline">{header_line}</div></div>')
    freeze_label, otbvar_label = _freeze_labels(tables)
    has_week = bool(tables.get("_has_week"))
    show_week = has_week if show_week is None else (show_week and has_week)
    if not show_lr and not show_week:
        show_lr = True                     # never render zero pickup blocks
    for seg, title in [("Total", "Total OTB"),
                       ("Transient", "Transient OTB"),
                       ("Group", "Group OTB")]:
        cols = _dashboard_layout(title, freeze_label, otbvar_label,
                                 show_lr=show_lr, show_week=show_week)
        parts.append(_table_html(title, tables[seg], cols))
    return "\n".join(parts)


# --------------------------------------------------------------------------- #
#  Excel export (mirrors the on-screen layout)
# --------------------------------------------------------------------------- #
def dashboard_to_excel(tables, header_line=None, show_lr=True, show_week=None,
                       wb=None, sheet_title="Dashboard"):
    """Write the owner Dashboard into a workbook.

    When `wb` is None this creates a fresh workbook and returns a BytesIO buffer
    (the standalone dashboard download). When an existing `wb` is passed, the
    Dashboard sheet is ADDED to it and the workbook object is returned (combined
    export).
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    standalone = wb is None
    if standalone:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title
    else:
        ws = wb.create_sheet(sheet_title)

    BAND = "F4CCCC"; SUB = "FBE4E4"; BUDGET = "EDE7D3"; TOTAL = "F2F2F2"
    band_fill = PatternFill("solid", fgColor=BAND)
    sub_fill = PatternFill("solid", fgColor=SUB)
    bud_fill = PatternFill("solid", fgColor=BUDGET)
    tot_fill = PatternFill("solid", fgColor=TOTAL)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    thin = Border(*(Side(style="thin", color="D9D9D9"),) * 4)
    bold = Font(bold=True, size=9)
    reg = Font(size=9)

    freeze_label, otbvar_label = _freeze_labels(tables)
    has_week = bool(tables.get("_has_week"))
    show_week = has_week if show_week is None else (show_week and has_week)
    if not show_lr and not show_week:
        show_lr = True

    ncols = 0
    row = 1
    if header_line:
        ws.cell(row=row, column=1, value=header_line).font = Font(bold=True, size=12)
        row += 2

    for seg, title in [("Total", "Total OTB"), ("Transient", "Transient OTB"),
                       ("Group", "Group OTB")]:
        rows = tables[seg]
        cols = _dashboard_layout(title, freeze_label, otbvar_label,
                                 show_lr=show_lr, show_week=show_week)
        groups = _groups_from_cols(cols)
        ncols = max(ncols, len(cols) + 1)
        # banner row
        ws.cell(row=row, column=1, value="Month")
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=1)
        col = 2
        for name, span in groups:
            ws.cell(row=row, column=col, value=name)
            if span > 1:
                ws.merge_cells(start_row=row, start_column=col,
                               end_row=row, end_column=col + span - 1)
            col += span
        for cc in range(1, col):
            cell = ws.cell(row=row, column=cc)
            cell.fill = band_fill; cell.font = bold; cell.alignment = center; cell.border = thin
        # sub-header row
        sr = row + 1
        for i, c in enumerate(cols):
            cell = ws.cell(row=sr, column=2 + i, value=c["sub"])
            cell.fill = bud_fill if c["budget"] else sub_fill
            cell.font = bold; cell.alignment = center; cell.border = thin
        ws.cell(row=sr, column=1).border = thin
        # data
        dr = sr + 1
        for r in rows:
            is_total = r["month"] == "Total"
            mc = ws.cell(row=dr, column=1, value=r["month"])
            mc.alignment = left; mc.border = thin; mc.font = bold if is_total else reg
            if is_total:
                mc.fill = tot_fill
            for i, c in enumerate(cols):
                v = r.get(c["key"])
                cell = ws.cell(row=dr, column=2 + i,
                               value=(v if isinstance(v, (int, float)) else None))
                cell.number_format = c["numfmt"]
                cell.border = thin; cell.alignment = center
                cell.font = bold if is_total else reg
                if c["budget"]:
                    cell.fill = bud_fill
                elif is_total:
                    cell.fill = tot_fill
            dr += 1
        row = dr + 1

    ws.column_dimensions["A"].width = 10
    for j in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(j)].width = 11
    ws.sheet_view.showGridLines = False

    if not standalone:
        return wb

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_combined_excel(tables, report, header_line=None,
                         show_lr=True, show_week=None):
    """Assemble ONE workbook containing both the owner Dashboard and the DD
    Pick-Up Report. Either section is optional — pass None to skip it.
    Sheet order: Dashboard, Pick-Up Report."""
    from openpyxl import Workbook
    wb = Workbook()
    default_ws = wb.active                      # empty "Sheet" we'll drop at the end

    if tables is not None:
        dashboard_to_excel(tables, header_line=header_line, show_lr=show_lr,
                           show_week=show_week, wb=wb, sheet_title="Dashboard")
    if report is not None:
        to_excel(report, wb=wb)

    # remove the throwaway default sheet if it's still empty
    try:
        if (default_ws in wb.worksheets and default_ws.max_row == 1
                and default_ws.max_column == 1 and default_ws["A1"].value is None):
            wb.remove(default_ws)
    except Exception:
        pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def dashboard_tables_for_export(parts, as_of):
    """Compute the owner-Dashboard tables from `parts`, using the shared
    'Comparison files' uploads (parts['cmp_lr'] / parts['cmp_wk']) for the pickup
    baselines. Used to embed the Dashboard sheet in the DD page's combined export.
    Returns (tables, header_line, show_lr, show_week) or None when there's no
    Market Segment data to build from."""
    cur_raw = parts.get("mseg_raw")
    if cur_raw is None:
        return None
    cur_agg, years = parse_market_segment(cur_raw)
    if not cur_agg:
        return None

    data = parts.get("data")
    capacity = 142                              # The Compton is always 142 rooms
    year = as_of.year if as_of.year in years else (min(years) if years else as_of.year)
    report_date = pd.Timestamp(as_of).date()

    cmp_lr = parts.get("cmp_lr")
    cmp_wk = parts.get("cmp_wk")
    last_agg = (cmp_lr or {}).get("monthly") or cur_agg   # no file -> pickup 0
    week_agg = (cmp_wk or {}).get("monthly") or None

    budget_df, freeze_df = default_reference_frames()
    tables = compute_dashboard(cur_agg, last_agg, capacity, year,
                               data_extract=data, budget_df=budget_df,
                               freeze_df=freeze_df, week_agg=week_agg)
    header_line = f"The Compton  |  {pd.Timestamp(report_date):%m/%d/%Y}"
    return tables, header_line, True, bool(tables.get("_has_week"))


_MONTH_RE = None  # lazily compiled


def agg_from_compton_export(sheets):
    """Reconstruct the monthly/segment agg from a PREVIOUS Compton Streamlit
    export by reading its 'Dashboard' sheet. The Dashboard stacks three tables
    (Total OTB / Transient OTB / Group OTB); within each, column layout always
    starts Month | Occ% | RN OTB | ADR OTB | Revenue OTB, so RN OTB is col 3 and
    Revenue OTB is col 5 regardless of which pickup blocks were shown.

    Returns (agg, years) where agg[(‘Mon-YYYY’, seg)] = {'rooms', 'rev'} for
    seg in {Transient, Group}. Total is derived, so it's skipped. Returns
    (None, None) if no Dashboard sheet is present.
    """
    import re
    global _MONTH_RE
    if _MONTH_RE is None:
        _MONTH_RE = re.compile(r"^[A-Z][a-z]{2}-\d{4}$")

    dash_name = next((s for s in sheets if s.strip().lower() == "dashboard"), None)
    if dash_name is None:
        return None, None
    df = sheets[dash_name]

    agg = {}
    years = set()
    cur_seg = None
    for i in range(len(df)):
        rowvals = list(df.iloc[i])
        # detect a section banner: any cell reading 'Transient OTB' / 'Group OTB' / 'Total OTB'
        banners = [str(x).strip() for x in rowvals if isinstance(x, str)]
        for b in banners:
            if b in ("Total OTB", "Transient OTB", "Group OTB"):
                cur_seg = b.split(" ")[0]      # Total / Transient / Group
        label = str(rowvals[0]).strip() if len(rowvals) and rowvals[0] is not None else ""
        if cur_seg in ("Transient", "Group") and _MONTH_RE.match(label):
            try:
                rn = float(rowvals[2]); rev = float(rowvals[4])
            except (TypeError, ValueError, IndexError):
                continue
            agg[(label, cur_seg)] = {"rooms": rn, "rev": rev}
            try:
                years.add(int(label.split("-")[1]))
            except (ValueError, IndexError):
                pass
    return (agg, years) if agg else (None, None)


def date_otb_from_mseg(raw):
    """Per-occupancy-date OTB from a raw IDeaS Market Segment export.
    Returns { 'YYYY-MM-DD': [total_rms, rev, trans_rms, group_rms] }."""
    df = raw.copy()
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)

    def find(*names):
        for n in names:
            for c in df.columns:
                if str(c).strip().lower() == n.lower():
                    return c
        return None

    c_date = find("Occupancy Date")
    c_seg = find("Market Segment")
    c_rms = find("Occupancy On Books This Year", "Occupancy on Books This Year")
    c_rev = find("Booked Room Revenue This Year")
    if c_date is None or c_seg is None or c_rms is None:
        return {}
    dates = pd.to_datetime(df[c_date], errors="coerce")
    rooms = pd.to_numeric(df[c_rms], errors="coerce").fillna(0)
    rev = (pd.to_numeric(df[c_rev], errors="coerce").fillna(0)
           if c_rev is not None else pd.Series(0.0, index=df.index))
    seg = df[c_seg].map(classify_segment)
    out = {}
    for d, s, rm, rv in zip(dates, seg, rooms, rev):
        if pd.isna(d):
            continue
        ds = d.strftime("%Y-%m-%d")
        cur = out.setdefault(ds, [0.0, 0.0, 0.0, 0.0])
        if s in ("Transient", "Group"):
            cur[0] += float(rm)             # total (Transient + Group)
            cur[1] += float(rv)
            if s == "Transient":
                cur[2] += float(rm)
            else:
                cur[3] += float(rm)
    return out


def date_otb_from_compton_export(sheets):
    """Per-occupancy-date OTB read from a PREVIOUS Compton export's 'Pick-Up
    Report' sheet. Returns { 'YYYY-MM-DD': [total, 0, trans, group] } (revenue is
    not carried at the date level in the DD, so it's 0). Returns {} if absent."""
    name = next((s for s in sheets
                 if s.strip().lower() in ("pick-up report", "pick up report")), None)
    if name is None:
        return {}
    df = sheets[name]
    if len(df) < 4:
        return {}
    r1 = df.iloc[0].ffill()                 # level-1 headers (merged -> ffill across cols)
    r2 = df.iloc[1].ffill()                 # level-2 headers
    r3 = df.iloc[2]

    def col_where(l1, l2, l3):
        for c in range(df.shape[1]):
            if (str(r1.iloc[c]).strip() == l1 and str(r2.iloc[c]).strip() == l2
                    and str(r3.iloc[c]).strip() == l3):
                return c
        return None

    c_date = None
    for c in range(df.shape[1]):
        if str(r1.iloc[c]).strip() == "Date":
            c_date = c
            break
    c_tot = col_where("Rooms Sold", "Total Hotel", "Current")
    c_tr = col_where("Rooms Sold", "Total Transient", "Current")
    c_gr = col_where("Rooms Sold", "Total Group", "Current")
    if c_date is None or c_tot is None:
        return {}

    out = {}
    for i in range(3, len(df)):
        row = df.iloc[i]
        d = pd.to_datetime(row.iloc[c_date], errors="coerce")
        if pd.isna(d):
            continue
        def num(c):
            if c is None:
                return 0.0
            v = pd.to_numeric(row.iloc[c], errors="coerce")
            return float(v) if pd.notna(v) else 0.0
        out[d.strftime("%Y-%m-%d")] = [num(c_tot), 0.0, num(c_tr), num(c_gr)]
    return out


def restrictions_from_compton_export(sheets):
    """Read the 'Restrictions' column from a previous Compton export's Pick-Up
    Report sheet, stripping any '(Previous …)' annotation back to the base value.
    Returns { 'YYYY-MM-DD': 'MinLOS 3' } (only dates that carry a restriction)."""
    name = next((s for s in sheets
                 if s.strip().lower() in ("pick-up report", "pick up report")), None)
    if name is None:
        return {}
    df = sheets[name]
    if len(df) < 4:
        return {}
    r1 = df.iloc[0]
    c_date = c_restr = None
    for c in range(df.shape[1]):
        h = str(r1.iloc[c]).strip()
        if h == "Date":
            c_date = c
        elif h == "Restrictions":
            c_restr = c
    if c_date is None or c_restr is None:
        return {}
    out = {}
    for i in range(3, len(df)):
        row = df.iloc[i]
        d = pd.to_datetime(row.iloc[c_date], errors="coerce")
        if pd.isna(d):
            continue
        base = _strip_restr_annotation(row.iloc[c_restr])
        if base:
            out[d.strftime("%Y-%m-%d")] = base
    return out


def report_date_from_compton_export(sheets):
    """Read the report/as-of date printed in a previous Compton export's header
    line (e.g. 'The Compton | 08/28/2026'). Looks at the Dashboard sheet first,
    then the Pick-Up Report. Returns a datetime.date or None."""
    import re
    date_re = re.compile(r"(\d{1,2})/(\d{1,2})/(\d{4})")
    for want in ("dashboard", "pick-up report", "pick up report"):
        name = next((s for s in sheets if s.strip().lower() == want), None)
        if name is None:
            continue
        df = sheets[name]
        for i in range(min(4, len(df))):
            for v in df.iloc[i].tolist():
                if isinstance(v, str):
                    m = date_re.search(v)
                    if m:
                        try:
                            mo, da, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
                            return pd.Timestamp(year=yr, month=mo, day=da).date()
                        except ValueError:
                            pass
    return None


def load_comparison(upload):
    """Parse an uploaded comparison file into BOTH the monthly/segment agg (for the
    Dashboard) and the per-occupancy-date OTB (for the DD's day-level pickup).
    Accepts EITHER a raw IDeaS **Market Segment** export OR a previous **Compton
    export** from this app. Returns a dict {'monthly', 'by_date', 'years',
    'report_date'} or None (report_date is only known for Compton exports)."""
    if upload is None:
        return None
    try:
        sheets = _read_all_sheets(upload)
    except Exception as e:
        st.warning(f"Couldn't read **{getattr(upload, 'name', 'that file')}** ({e}).")
        return None
    report_date = None
    restrictions = {}
    ms = [s for s in sheets if "market segment" in s.lower()]
    if ms:
        monthly, years = parse_market_segment(sheets[ms[0]])
        by_date = date_otb_from_mseg(sheets[ms[0]])
    else:
        monthly, years = agg_from_compton_export(sheets)
        by_date = date_otb_from_compton_export(sheets)
        report_date = report_date_from_compton_export(sheets)
        restrictions = restrictions_from_compton_export(sheets)
        if not monthly and not by_date:
            # last resort: first sheet as a Market Segment layout
            monthly, years = parse_market_segment(sheets[list(sheets)[0]])
            by_date = date_otb_from_mseg(sheets[list(sheets)[0]])
    if not monthly and not by_date:
        return None
    return {"monthly": monthly or {}, "by_date": by_date or {},
            "years": years or set(), "report_date": report_date,
            "restrictions": restrictions or {}}


def _mseg_agg_from_upload(upload):
    """Back-compat helper: return just the monthly agg from a comparison upload."""
    cmp = load_comparison(upload)
    if cmp is None:
        return None, None
    return cmp["monthly"], cmp["years"]


def pickup_label_map(parts):
    """Map the DD's canonical pickup column names to DATED labels using the
    comparison files' embedded report dates (available from a prior Compton
    export). Only remaps when the date is known; a raw Market Segment file has no
    embedded report date, so those columns keep their generic label.
        'Rooms Sold | <seg> | Change' -> 'Rooms Sold | <seg> | P/U vs MM/DD'
        '7-Day P/U | <seg>'           -> '7-Day P/U vs MM/DD | <seg>'
    """
    m = {}
    lr = (parts.get("cmp_lr") or {}).get("report_date")
    wk = (parts.get("cmp_wk") or {}).get("report_date")
    segs = ("Total Hotel", "Total Transient", "Total Group")
    if lr:
        s = pd.Timestamp(lr).strftime("%m/%d")
        for seg in segs:
            m[f"Rooms Sold | {seg} | Change"] = f"Rooms Sold | {seg} | P/U vs {s}"
    if wk:
        s = pd.Timestamp(wk).strftime("%m/%d")
        for seg in segs:
            m[f"7-Day P/U | {seg}"] = f"7-Day P/U vs {s} | {seg}"
    return m


def render_dashboard_page(parts, as_of):
    """Streamlit page that reproduces the workbook 'Dashboard' tab."""
    st.title("The Compton — Owner Dashboard")
    st.caption("Total / Transient / Group on-the-books by month, versus last "
               "report, budget and the freeze forecast — the same view as the "
               "Dashboard tab of the Daily Report workbook.")

    cur_raw = parts.get("mseg_raw")
    if cur_raw is None:
        st.info("Upload your **Market Segment** export in the sidebar to build "
                "the dashboard (it supplies rooms and revenue by segment).")
        st.stop()

    cur_agg, years = parse_market_segment(cur_raw)
    if not cur_agg:
        st.error("Couldn't read rooms/revenue by segment from the Market Segment "
                 "file. Expected columns: 'Occupancy Date', 'Market Segment', "
                 "'Occupancy On Books This Year', 'Booked Room Revenue This Year'.")
        st.stop()

    # ---- capacity is fixed at 142 rooms for The Compton ----
    capacity = 142
    data = parts.get("data")

    # ---- pick the reporting year ----
    year = as_of.year if as_of.year in years else (min(years) if years else as_of.year)

    report_date = pd.Timestamp(as_of).date()

    with st.sidebar:
        st.markdown("---")
        st.header("Dashboard settings")
        if len(years) > 1:
            year = st.selectbox("Report year", sorted(years),
                                index=sorted(years).index(year))
        pickup_view = st.radio(
            "Pickup columns to show",
            ["Since last report", "7-day", "Both"], index=0, horizontal=True,
            help="Choose which pickup blocks to show. The '7-day' options need a "
                 "'7 days ago' comparison file in the main upload area.")
        dd_show_bar_chg = st.checkbox(
            "Include 'BAR Chg 7d' in the DD export", value=False,
            help="Controls the DD Pick-Up Report sheet inside the "
                 "'Dashboard + DD' workbook. Off by default so the day-by-day "
                 "pickup columns stay closer together.")

    # ---- pickup baselines come from the shared 'Comparison files' uploads ----
    cmp_lr = parts.get("cmp_lr")
    cmp_wk = parts.get("cmp_wk")

    last_agg, last_report_date, pickup_src = None, None, None
    if cmp_lr and cmp_lr.get("monthly"):
        last_agg = cmp_lr["monthly"]
        last_report_date = parts.get("cmp_lr_date") or (pd.Timestamp(report_date) - pd.Timedelta(days=1)).date()
        pickup_src = "uploaded file"
    if last_agg is None:                       # nothing to compare -> pickup 0
        last_agg = cur_agg
        last_report_date = report_date
        pickup_src = "none yet"

    week_agg, week_date, week_src = None, None, None
    if cmp_wk and cmp_wk.get("monthly"):
        week_agg = cmp_wk["monthly"]
        week_date = parts.get("cmp_wk_date") or (pd.Timestamp(report_date) - pd.Timedelta(days=7)).date()
        week_src = "uploaded file"

    # ---- Budget & Freeze: pull from the Data Extraction when it carries the
    #      Budget / Forecast columns, else fall back to the manual reference
    #      tables. Budget    = 'Budget …' columns.
    #                Freeze  = 'My Forecast …' (user) if present, else the
    #                          'Occupancy Forecast …' / 'Forecasted …' (system).
    bf_probe, _ = budget_freeze_from_extract(data, year)
    have_extract = bf_probe is not None

    budget_df, freeze_df = default_reference_frames()
    if not have_extract:
        with st.sidebar.expander("Budget & Freeze Forecast (manual)", expanded=False):
            st.caption("No Budget/Forecast columns were found in the Data "
                       "Extraction, so these fall back to editable reference "
                       "numbers. Re-export the Data Extraction with the Budget "
                       "and Forecast columns to pull them automatically.")
            st.markdown("**Budget**")
            budget_df = st.data_editor(budget_df, hide_index=True, key="dash_budget",
                                       use_container_width=True)
            st.markdown("**Freeze Forecast**")
            freeze_df = st.data_editor(freeze_df, hide_index=True, key="dash_freeze",
                                       use_container_width=True)

    tables = compute_dashboard(cur_agg, last_agg, capacity, year,
                               data_extract=data, budget_df=budget_df,
                               freeze_df=freeze_df, week_agg=week_agg)

    # ---- translate the pickup-view choice into render flags ----
    has_week = bool(tables.get("_has_week"))
    if pickup_view == "Since last report":
        show_lr, show_week = True, False
    elif pickup_view == "7-day":
        show_lr, show_week = False, True
    else:  # Both
        show_lr, show_week = True, True
    if (show_week and not show_lr) and not has_week:
        show_lr, show_week = True, False       # asked for 7-day but none available

    # ---- source banner so it's clear where Budget / Freeze came from ----
    if have_extract:
        fs = tables.get("_freeze_source", "System forecast")
        st.success(
            f"**Budget** pulled from the Data Extraction (Budget columns). "
            f"**Freeze Forecast** = {fs} "
            f"({'your My-Forecast values' if fs == 'Your forecast' else 'system forecast — no My-Forecast values found in the file'}). "
            "Note: the export only carries revenue at the total level, so within "
            "Budget/Freeze each segment shares that month's blended ADR while "
            "Transient + Group revenue still reconciles to the total.")
    else:
        st.warning("Budget & Freeze are using the manual reference tables "
                   "(sidebar). Re-export the Data Extraction with the Budget "
                   "and Forecast columns to pull them automatically.")

    # ---- pickup comparison banner(s) ----
    days_gap = (pd.Timestamp(report_date) - pd.Timestamp(last_report_date)).days
    day_word = "day" if days_gap == 1 else "days"
    if pickup_src == "none yet":
        st.info("No **Last report** comparison file uploaded, so **pickup since "
                "last report shows 0**. Add a prior day's file in the main upload "
                "area (Comparison files) to populate it.")
    elif show_lr:
        st.caption(f"📈 **Pickup since last report** compares "
                   f"**{pd.Timestamp(report_date):%b %d, %Y}** to "
                   f"**{pd.Timestamp(last_report_date):%b %d, %Y}** "
                   f"({days_gap} {day_word}, from your uploaded file).")

    if show_week:
        if has_week:
            wgap = (pd.Timestamp(report_date) - pd.Timestamp(week_date)).days
            st.caption(f"🗓️ **7-day pickup** compares "
                       f"**{pd.Timestamp(report_date):%b %d, %Y}** to "
                       f"**{pd.Timestamp(week_date):%b %d, %Y}** "
                       f"({wgap} days, from your uploaded file).")
        elif pickup_view != "Since last report":
            st.warning("No **7 days ago** comparison file uploaded, so the 7-day "
                       "pickup can't be shown. Add one in the main upload area.")

    gap_bits = []
    if show_lr:
        gap_bits.append(f"vs last report {pd.Timestamp(last_report_date):%m/%d/%Y}")
    if show_week and has_week:
        gap_bits.append(f"vs 7-days-ago {pd.Timestamp(week_date):%m/%d/%Y}")
    header_line = (f"The Compton&nbsp;&nbsp;|&nbsp;&nbsp;{pd.Timestamp(report_date):%m/%d/%Y}"
                   + ("&nbsp;&nbsp;|&nbsp;&nbsp;" + " · ".join(gap_bits) if gap_bits else ""))

    st.markdown(render_dashboard_html(tables, header_line,
                                      show_lr=show_lr, show_week=show_week),
                unsafe_allow_html=True)

    hdr_plain = header_line.replace("&nbsp;", " ")
    dl1, dl2 = st.columns(2)
    dl1.download_button(
        "⬇️ Dashboard only (Excel)",
        data=dashboard_to_excel(tables, hdr_plain,
                                show_lr=show_lr, show_week=show_week),
        file_name=f"Compton_Dashboard_{pd.Timestamp(report_date):%Y%m%d}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    # Combined workbook — Dashboard + DD Pick-Up Report in one file.
    # The DD report is built from the same uploaded files (needs the Data Extract).
    # Restrictions (entered on the DD page) live in session state and carry over.
    dd_report = None
    if parts.get("data") is not None:
        try:
            _restr_cur = st.session_state.get("dd_restr", {})
            _restr_prev = (parts.get("cmp_lr") or {}).get("restrictions") or {}
            dd_report = build_report(parts, as_of, show_own_change=dd_show_bar_chg,
                                     restrictions=_restr_cur,
                                     prev_restrictions=_restr_prev,
                                     restr_annotate=bool(parts.get("cmp_lr")))
            dd_report = dd_report.rename(columns=pickup_label_map(parts))
        except Exception:
            dd_report = None
    if dd_report is not None:
        combined = build_combined_excel(
            tables, dd_report,
            header_line=hdr_plain, show_lr=show_lr, show_week=show_week)
        dl2.download_button(
            "⬇️ Dashboard + DD (one Excel)",
            data=combined,
            file_name=f"Compton_Report_{pd.Timestamp(report_date):%Y%m%d}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        dl2.caption("Upload the Data Extract too to get a combined "
                    "Dashboard + DD workbook.")

st.title("The Compton — Daily Pick-Up Report Builder")
st.caption("Upload your four SynXis / rate-shop exports and get the full pick-up "
           "report instantly. No copy/paste required.")

with st.sidebar:
    page = st.radio("View", ["Pick-Up Report (DD)", "Dashboard"], index=0,
                    help="Switch between the daily pick-up report and the "
                         "owner dashboard (Total / Transient / Group by month).")
    st.markdown("---")
    st.header("1 - Upload your files")
    ups = st.file_uploader(
        "Drop in PCDC, Data Extract, Market Seg and the Rate-Shop file "
        "(all four at once is fine)",
        type=["xlsx"], accept_multiple_files=True,
    )

    st.markdown("---")
    st.subheader("Comparison files (pickup)")
    st.caption("Optional. Drop in a prior day's file to drive pickup on **both** "
               "the Dashboard and the DD. Each slot accepts **either** a raw "
               "IDeaS **Market Segment** export **or** a **previous Compton "
               "export** from this app.")
    cmp_lr_file = st.file_uploader("① Last report — Market Segment or Compton export",
                                   type=["xlsx"], accept_multiple_files=False,
                                   key="cmp_lr_up")
    cmp_wk_file = st.file_uploader("② 7 days ago — Market Segment or Compton export",
                                   type=["xlsx"], accept_multiple_files=False,
                                   key="cmp_wk_up")

if not ups:
    st.info("Upload your files in the sidebar to build the report.")
    st.stop()

parts = detect_and_load(ups)

# Parse the shared comparison files once; both pages read these from `parts`.
#   cmp_lr -> "since last report" pickup;  cmp_wk -> "7-day" pickup.
parts["cmp_lr"] = load_comparison(cmp_lr_file)
parts["cmp_wk"] = load_comparison(cmp_wk_file)

# The date range actually covered by the PCDC (pace/pickup) report, if uploaded.
# Data Extract typically spans the WHOLE YEAR, but PCDC only covers its shorter
# pace window (e.g. ~60-90 days). We use this as the default view so the report
# isn't cluttered with the full year of Data-Extract-only dates.
_pcdc_min = _pcdc_max = None
if parts.get("pcdc") is not None and "Occupancy Date" in parts["pcdc"].columns:
    _pcdc_dates = pd.to_datetime(parts["pcdc"]["Occupancy Date"], errors="coerce").dropna()
    if len(_pcdc_dates):
        _pcdc_min, _pcdc_max = _pcdc_dates.min().date(), _pcdc_dates.max().date()

# 'Days Left' always counts from the REAL current date. (The as-of date picker
# was removed — the app simply uses today, which is what daily uploads want.)
real_today = pd.Timestamp.today().date()
as_of = real_today

with st.sidebar:
    st.header("2 - Settings")
    if parts.get("pcdc_asof"):
        st.caption(f"ℹ️ For reference, this PCDC report was generated on "
                   f"**{pd.Timestamp(parts['pcdc_asof']):%b %d, %Y}**.")
    if page != "Dashboard":
        comp_method = st.radio("Comp Set Avg method", ["Average", "Median"], horizontal=True)
        show_change = st.radio(
            "Show competitor rate change vs.",
            ["None", "Yesterday", "3 days ago", "7 days ago"],
            help="Adds a 'Change' column beside each competitor, pulled from the "
                 "matching 'vs.' tab in the rate-shop file.",
        )
        show_own_change = st.checkbox(
            "Show our BAR change (7 days)", value=True,
            help="Adds a 'BAR Chg 7d' column showing how our own rate moved over the "
                 "last 7 days (from the rate-shop file).",
        )
    
        st.markdown("---")
        st.header("3 - Signal sensitivity")
        st.caption("The Compton prices as the market leader (#1 in the comp set). "
                   "The ⚡ Action column is purely about rate-leadership position.")
        with st.expander("🔴 Rate leadership (#1 vs #2)", expanded=False):
            st.caption("🔴 Not #1 = a competitor is at or above our BAR.  "
                       "🟠 Too close to #2 = we're #1, but by less than the cushion below.")
            lead_buffer = st.slider("Minimum $ cushion we want over the #2 (top competitor)",
                                    5, 150, 20, 5,
                                    help="If we're #1 but ahead by less than this, flag "
                                         "'Too close to #2'. If a competitor is at/above our "
                                         "BAR, that's always flagged as 'Not #1' regardless "
                                         "of this slider.")
        with st.expander("🔥 Spike alert", expanded=False):
            st.caption("This self-calibrates to this property's own pickup pattern instead of "
                       "a fixed number — it flags days with unusually high pickup relative to "
                       "everything else in the current report.")
            pickup_pctile = st.slider("Flag pickup at or above this percentile of all pickup days",
                                      50, 99, 75, 1,
                                      help="Lower = flags MORE days (e.g. 75 flags the busiest "
                                           "quarter of pickup days). Higher = flags fewer, more "
                                           "extreme outliers.")
            pickup_min = st.slider("...but never flag below this many rooms (floor)",
                                   1, 30, 3, 1,
                                   help="Keeps quiet reports from flagging tiny 1-room pickups "
                                        "just because they're 'relatively' high.")
    
        thr = {"lead_buffer": lead_buffer,
               "pickup_pctile": pickup_pctile, "pickup_min": pickup_min}
    
        st.markdown("---")
        st.caption("Tip: leave the file names as SynXis exports them — the app "
                   "auto-detects each report by its sheets.")

# --------------------------------------------------------------------------- #
#  Page router: render the Dashboard and stop before the pick-up report runs.
# --------------------------------------------------------------------------- #
if page == "Dashboard":
    render_dashboard_page(parts, as_of)
    st.stop()

c1, c2, c3, c4 = st.columns(4)
c1.metric("PCDC",         "OK" if parts["pcdc"] is not None else "-")
c2.metric("Data Extract", "OK" if parts["data"] is not None else "-")
c3.metric("Market Seg",   "OK" if parts["mseg"] is not None else "-")
c4.metric("Rate Shop",    "OK" if parts["shop"] is not None else "-")

st.caption(f"📅 'Days Left' counts from today, **{pd.Timestamp(as_of):%b %d, %Y}**.")

# =============================================================== #
#  Physical restrictions (manual entry + carry-forward)
# =============================================================== #
_prev_restr = (parts.get("cmp_lr") or {}).get("restrictions") or {}
_restr_has_prev = bool(parts.get("cmp_lr"))

# Carry-forward: seed the working set from the prior report the first time, and
# re-seed whenever a NEW comparison file is uploaded (so last report's
# restrictions roll forward automatically). Track the file signature to detect a
# fresh upload.
_cmp_sig = None
if "cmp_lr_file" in dir() and cmp_lr_file is not None:
    _cmp_sig = (cmp_lr_file.name, getattr(cmp_lr_file, "size", None))
if "dd_restr" not in st.session_state:
    st.session_state["dd_restr"] = dict(_prev_restr)
    st.session_state["dd_restr_sig"] = _cmp_sig
elif _cmp_sig is not None and st.session_state.get("dd_restr_sig") != _cmp_sig:
    st.session_state["dd_restr"] = dict(_prev_restr)   # new prior file -> roll forward
    st.session_state["dd_restr_sig"] = _cmp_sig

with st.expander("🔒 Physical restrictions (manual entry)", expanded=False):
    st.caption("Enter length-of-stay restrictions you can't export from IDeaS. "
               "They appear next to **Last Room Value** and, when a Last-report "
               "comparison file is loaded, flag changes as **(Previous …)**. "
               "Restrictions carry forward automatically from your prior export.")

    fc1, fc2, fc3, fc4 = st.columns([1, 1, 1.1, 0.8])
    _r_from = fc1.date_input("From", value=as_of, key="restr_from")
    _r_to = fc2.date_input("To", value=as_of, key="restr_to")
    _r_type = fc3.selectbox("Restriction", ["Open (remove)"] + RESTRICTION_TYPES,
                            key="restr_type")
    fc4.markdown("<div style='height:1.7em'></div>", unsafe_allow_html=True)
    if fc4.button("Apply", use_container_width=True):
        cur = dict(st.session_state["dd_restr"])
        d0, d1 = sorted([pd.Timestamp(_r_from), pd.Timestamp(_r_to)])
        for d in pd.date_range(d0, d1):
            ds = d.strftime("%Y-%m-%d")
            if _r_type == "Open (remove)":
                cur.pop(ds, None)
            else:
                cur[ds] = _r_type
        st.session_state["dd_restr"] = cur
        st.success(f"Applied **{_r_type}** to {d0:%b %d} – {d1:%b %d, %Y}.")

    # Editable list for fine control (add/delete rows, change type)
    _cur = st.session_state["dd_restr"]
    _restr_df = pd.DataFrame(
        [{"Date": pd.Timestamp(d).date(), "Restriction": v}
         for d, v in sorted(_cur.items())]
    )
    if _restr_df.empty:
        _restr_df = pd.DataFrame({"Date": pd.Series([], dtype="object"),
                                  "Restriction": pd.Series([], dtype="object")})
    _edited = st.data_editor(
        _restr_df, num_rows="dynamic", use_container_width=True, key="restr_editor",
        column_config={
            "Date": st.column_config.DateColumn("Date", format="YYYY-MM-DD"),
            "Restriction": st.column_config.SelectboxColumn(
                "Restriction", options=RESTRICTION_TYPES, required=True),
        })
    # rebuild the working set from the edited table
    _new = {}
    for _, _row in _edited.iterrows():
        _d, _v = _row.get("Date"), _row.get("Restriction")
        if pd.isna(_d) or _v in (None, "", float("nan")) or (isinstance(_v, float) and pd.isna(_v)):
            continue
        _new[pd.Timestamp(_d).strftime("%Y-%m-%d")] = str(_v)
    st.session_state["dd_restr"] = _new

    if st.button("↻ Reset to last report's restrictions"):
        st.session_state["dd_restr"] = dict(_prev_restr)
        st.rerun()

    _n_active = len(st.session_state["dd_restr"])
    _n_changed = sum(
        1 for ds, v in st.session_state["dd_restr"].items() if _prev_restr.get(ds, "") != v
    ) + sum(1 for ds in _prev_restr if ds not in st.session_state["dd_restr"])
    st.caption(f"**{_n_active}** date(s) with a restriction"
               + (f" · **{_n_changed}** changed vs last report" if _restr_has_prev else ""))

_dd_restr = st.session_state.get("dd_restr", {})

report = build_report(parts, as_of, comp_method, show_change, thr, show_own_change,
                      restrictions=_dd_restr, prev_restrictions=_prev_restr,
                      restr_annotate=_restr_has_prev)

# =============================================================== #
#  Filters
# =============================================================== #
st.subheader("Filters")
min_d, max_d = report["Date"].min(), report["Date"].max()

_window_options = ["Custom", "Next 7 days", "Next 14 days", "Next 30 days",
                  "Next 60 days", "Next 90 days", "All future dates", "All dates (incl. past)"]
_default_window_idx = 7
if _pcdc_min is not None:
    _window_options = ["PCDC report range"] + _window_options
    _default_window_idx = 0     # default to exactly the dates covered by PCDC

fa, fb, fc = st.columns([1.2, 1, 1])
window = fa.selectbox("Booking window", _window_options, index=_default_window_idx,
                      help="'PCDC report range' (default, when a PCDC file is uploaded) "
                           "shows exactly the dates covered by your PCDC pace report — "
                           f"{_pcdc_min:%b %d} to {_pcdc_max:%b %d, %Y}"
                           if _pcdc_min is not None else
                           "'All future dates' or any 'Next X days' option will HIDE "
                           "dates before your as-of date (negative Days Left) — use "
                           "the look-back option below to also see recent past dates.")
dow_pick = fb.multiselect("Day of week", ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"],
                          default=["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"])
action_pick = fc.multiselect("Action", ["🔴 Not #1", "🟠 Too close to #2", "🟠 At ceiling", "✅ Leading"],
                             default=["🔴 Not #1", "🟠 Too close to #2", "🟠 At ceiling", "✅ Leading"])

lookback = st.number_input(
    "Also include the last N days before the as-of date (negative Days Left)",
    min_value=0, max_value=90, value=0, step=1,
    help="0 = forward-looking only (default). Set this above 0 to also pull in recent "
         "past dates alongside 'Next X days' / 'All future dates', e.g. to review last "
         "week's pickup. Ignored when 'Custom' or 'All dates (incl. past)' is selected.")

fd, fe, ff, fg = st.columns([1, 1, 1, 1])
events_only = fd.checkbox("Events only", value=False)
occ_band = fe.select_slider("Occ Forecast % band",
                            options=[0, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100],
                            value=(0, 100))
var_thresh = ff.number_input("Min |variance vs STLY| (rooms)", min_value=0, value=0, step=1)
pickup_only = fg.checkbox("🔥 Spike days only", value=False)

# Pickup focus — filter rows to just the dates that moved, since last report or
# over the 7-day window (the latter needs a '7 days ago' comparison file).
_has_wk_col = "7-Day P/U | Total Hotel" in report.columns
_focus_opts = ["All dates", "Picked up since last report"]
if _has_wk_col:
    _focus_opts.append("7-day pickup only")
pickup_focus = st.selectbox(
    "Pickup focus", _focus_opts, index=0,
    help="'Picked up since last report' keeps dates whose Total-Hotel rooms "
         "changed since your last report. '7-day pickup only' (needs a '7 days "
         "ago' comparison file) keeps dates that moved over the last 7 days — "
         "handy for the 7-day recap.")

# custom date range (used when window == Custom)
start, end = min_d, max_d          # defaults so the filename always has a range
if window == "Custom":
    d1, d2 = st.columns(2)
    default_start = max(min_d, as_of)
    start = d1.date_input("From", value=default_start, min_value=min_d, max_value=max_d)
    end   = d2.date_input("To",   value=max_d,         min_value=min_d, max_value=max_d)

# ---- apply filters ---- #
v = report.copy()
win_map = {"Next 7 days": 7, "Next 14 days": 14, "Next 30 days": 30,
           "Next 60 days": 60, "Next 90 days": 90}
floor_days = -int(lookback)          # e.g. lookback=7 -> also show Days Left down to -7
if window == "PCDC report range" and _pcdc_min is not None:
    v = v[(v["Date"] >= _pcdc_min) & (v["Date"] <= _pcdc_max)]
elif window in win_map:
    v = v[(v["Days Left"] >= floor_days) & (v["Days Left"] <= win_map[window])]
elif window == "All future dates":
    v = v[v["Days Left"] >= floor_days]
elif window == "All dates (incl. past)":
    pass  # no Days Left filter at all — show everything in the uploaded range
else:  # Custom
    v = v[(v["Date"] >= start) & (v["Date"] <= end)]

dow_full = {"Mon": "Monday", "Tue": "Tuesday", "Wed": "Wednesday", "Thu": "Thursday",
            "Fri": "Friday", "Sat": "Saturday", "Sun": "Sunday"}
keep_dow = [dow_full[d] for d in dow_pick]
v = v[v["DOW"].isin(keep_dow)]

if events_only:
    v = v[v["Events"].astype(str).str.strip() != ""]

if "⚡ Action" in v:
    a = v["⚡ Action"].fillna("")
    sel = pd.Series(False, index=v.index)
    if "🔴 Not #1" in action_pick:          sel |= a.str.contains("Not #1")
    if "🟠 Too close to #2" in action_pick: sel |= a.str.contains("Too close")
    if "🟠 At ceiling" in action_pick:      sel |= a.str.contains("At ceiling")
    if "✅ Leading" in action_pick:         sel |= a.str.contains("Leading")
    sel |= (a.str.strip() == "")      # keep past/blank rows
    v = v[sel]

ofp_col = pd.to_numeric(v.get("Occ Forecast % | Total Hotel"), errors="coerce") * 100
lo, hi = occ_band
v = v[(ofp_col.isna()) | ((ofp_col >= lo) & (ofp_col <= hi))]

if var_thresh > 0 and "Rooms OTB STLY | Total Hotel | Change" in v:
    var_col = pd.to_numeric(v["Rooms OTB STLY | Total Hotel | Change"], errors="coerce").abs()
    v = v[var_col >= var_thresh]

if pickup_only and "🔥 Spike" in v:
    v = v[v["🔥 Spike"].astype(str).str.strip() != ""]

if pickup_focus == "Picked up since last report" and "Rooms Sold | Total Hotel | Change" in v:
    _pc = pd.to_numeric(v["Rooms Sold | Total Hotel | Change"], errors="coerce").fillna(0)
    v = v[_pc != 0]
elif pickup_focus == "7-day pickup only" and "7-Day P/U | Total Hotel" in v:
    _pc = pd.to_numeric(v["7-Day P/U | Total Hotel"], errors="coerce").fillna(0)
    v = v[_pc != 0]

view = v.reset_index(drop=True)

# Relabel the pickup columns with the comparison dates (e.g. "P/U vs 08/28",
# "7-Day P/U vs 08/24"). Done AFTER all filters (which use the canonical names)
# so both the on-screen table and the Excel export show the dated headers.
view = view.rename(columns=pickup_label_map(parts))

st.subheader(f"Pick-Up Report — {len(view)} dates")

pct_cols   = [c for c in view.columns if c.startswith("Occ Forecast %")]
shop_all   = [c for c in view.columns if c.startswith("Competitor Shops")]
shop_chg   = [c for c in shop_all if c.endswith("Change")]
shop_cur   = [c for c in shop_all if c not in shop_chg]
money_cols = (["BAR", "Comp Set Avg", "Last Room Value", "Estimated ADR"]
              + [c for c in view.columns if c.startswith("Booked ADR")])
rooms_chg  = [c for c in view.columns
              if (c.startswith("Rooms Sold") or c.startswith("Rooms OTB STLY"))
              and (c.endswith("Change") or "| P/U vs" in c)]
wk_pu_cols = [c for c in view.columns if c.startswith("7-Day P/U")]
var_cols   = ([c for c in view.columns if "Variance" in c] + shop_chg + rooms_chg
              + wk_pu_cols
              + [c for c in ("Lead Gap vs Comp", "BAR Chg 7d") if c in view.columns])
otb_sold   = "Rooms Left to Sell"      # heat-map column

def _money(v):
    if isinstance(v, (int, float)) and not pd.isna(v):
        return f"${v:,.0f}"
    return "" if v is None or (isinstance(v, float) and pd.isna(v)) else str(v)

def _signed(v):
    if isinstance(v, (int, float)) and not pd.isna(v) and v != 0:
        return f"{v:+,.0f}"
    return ""

fmt = {}
for c in pct_cols:
    fmt[c] = "{:.1%}"
for c in money_cols:
    fmt[c] = "${:,.0f}"
for c in shop_cur:
    fmt[c] = _money
for c in shop_chg:
    fmt[c] = _signed
for c in ("Lead Gap vs Comp", "BAR Chg 7d"):
    if c in view.columns:
        fmt[c] = _signed
for c in wk_pu_cols:
    fmt[c] = _signed
for c in view.columns:
    if c in ("Days Left", "OOO", "Ovrbk") or c.startswith(("Rooms Sold", "Rooms OTB STLY",
             "Remaining Demand", "Occ Forecast |")) and c not in pct_cols:
        fmt.setdefault(c, "{:,.0f}")

styler = (view.style
          .format(fmt, na_rep="")
          .map(lambda v: "color:#1a7f37;font-weight:600" if isinstance(v, (int, float)) and v > 0
               else ("color:#cf222e;font-weight:600" if isinstance(v, (int, float)) and v < 0 else ""),
               subset=var_cols))


# Heat map on the OTB rooms sold column — pure-Python red->yellow->green
# gradient (no matplotlib needed, so it works on Streamlit Cloud out of the box).
def _heat_bg(series):
    # For "Rooms Left to Sell": GREEN = lots of rooms left (wide open),
    # RED = few rooms left (near sell-out).  Low value -> red, high value -> green.
    vals = pd.to_numeric(series, errors="coerce")
    lo, hi = vals.min(), vals.max()
    span = (hi - lo) or 1
    out = []
    for v in vals:
        if pd.isna(v):
            out.append("")
            continue
        t = (v - lo) / span                    # low rooms left -> red, high -> green
        if t < 0.5:
            f = t / 0.5
            red, grn, blu = 248, int(105 + f * (235 - 105)), int(107 + f * (132 - 107))
        else:
            f = (t - 0.5) / 0.5
            red, grn, blu = int(255 - f * (255 - 99)), int(235 - f * (235 - 190)), int(132 - f * (132 - 123))
        out.append(f"background-color: #{red:02X}{grn:02X}{blu:02X}")
    return out

# Heat map is applied directly inside the pick-up report on the OTB column
if otb_sold in view.columns and view[otb_sold].notna().any():
    styler = styler.apply(_heat_bg, subset=[otb_sold])

# Colour the ⚡ Action column so recommendations pop
def _action_bg(series):
    out = []
    for v in series.astype(str):
        if "Not #1" in v:
            out.append("background-color:#F8D7DA;color:#842029;font-weight:600")
        elif "Too close" in v:
            out.append("background-color:#FFF3CD;color:#664D03;font-weight:600")
        elif "At ceiling" in v:
            out.append("background-color:#FFE5B4;color:#7A4A00;font-weight:600")
        elif "Leading" in v:
            out.append("background-color:#D1E7DD;color:#0F5132;font-weight:600")
        else:
            out.append("")
    return out

if "⚡ Action" in view.columns:
    styler = styler.apply(_action_bg, subset=["⚡ Action"])

# Colour the 🔥 Spike column so big pickup days pop
def _pickup_bg(series):
    return ["background-color:#FDE7CE;color:#B45309;font-weight:600" if "Spike" in str(v)
            else "" for v in series]

if "🔥 Spike" in view.columns:
    styler = styler.apply(_pickup_bg, subset=["🔥 Spike"])

# Highlight the manual Restrictions column (deeper amber when it changed)
def _restr_bg(series):
    out = []
    for v in series.astype(str):
        s = v.strip()
        if not s or s.lower() == "nan":
            out.append("")
        elif "(Previous" in s:
            out.append("background-color:#FFE08A;color:#7A4A00;font-weight:700")
        else:
            out.append("background-color:#FFF3CD;color:#7A4A00;font-weight:600")
    return out

if "Restrictions" in view.columns:
    styler = styler.apply(_restr_bg, subset=["Restrictions"])

st.dataframe(styler, use_container_width=True, height=560)

# filename reflects the actual filtered date range (falls back to the full range)
if len(view) > 0:
    _fn_start, _fn_end = view["Date"].min(), view["Date"].max()
else:
    _fn_start, _fn_end = start, end

_ddl1, _ddl2 = st.columns(2)
_ddl1.download_button(
    "⬇️ Pick-Up Report only (Excel)",
    data=to_excel(view),
    file_name=f"Compton_PickUp_{_fn_start:%Y%m%d}_{_fn_end:%Y%m%d}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)

# Combined workbook — Dashboard + DD Pick-Up Report in one file. The Dashboard is
# built from the uploaded Market Segment + comparison files (no extra uploads).
_dash_export = dashboard_tables_for_export(parts, as_of)
if _dash_export is not None:
    _dtables, _dhdr, _dlr, _dweek = _dash_export
    _combined = build_combined_excel(
        _dtables, view, header_line=_dhdr,
        show_lr=_dlr, show_week=_dweek)
    _ddl2.download_button(
        "⬇️ DD + Dashboard (one Excel)",
        data=_combined,
        file_name=f"Compton_Report_{_fn_start:%Y%m%d}_{_fn_end:%Y%m%d}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
else:
    _ddl2.caption("Upload the Market Segment export too to get a combined "
                  "DD + Dashboard workbook.")
